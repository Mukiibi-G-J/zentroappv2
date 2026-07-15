import uuid
import django_filters as filters
import os
import json
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import (
    Q,
    Case,
    When,
    Value,
    IntegerField,
    DecimalField,
    Sum,
    F,
    Prefetch,
    Exists,
    OuterRef,
    Subquery,
)
from django.db.models.functions import Coalesce
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, CreateView, View, UpdateView
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.cache import cache
from django_tenants.utils import schema_context
from django.shortcuts import get_object_or_404
from django.db import connection
from django.conf import settings
from celery.result import AsyncResult
from celery.states import PENDING, SUCCESS, FAILURE
import base64

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import ListAPIView, UpdateAPIView
from rest_framework.authentication import SessionAuthentication
from rest_framework.permissions import IsAuthenticated
from authentication.authentication import JWTAuthenticationWithRevocationChecks as JWTAuthentication
from rest_framework import viewsets, status
from rest_framework.decorators import action
from django.db import transaction
from django.core.exceptions import ValidationError

from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination


from items.enums import EntryType, InventoryType
from financials.enums import GeneralPostingType
from dimension.branch_filter import filter_queryset_by_branch, get_branch_for_request


from items.models import (
    ItemImages,
    ItemJournal,
    UnitOfMeasure,
    ItemLedgerEntries,
    ItemTrackingCodes,
    ItemUnitOfMeasure,
    TrackingSpecification,
    ItemAttribute,
    ItemAttributeValue,
    ItemAttributeEntry,
    Location,
    ValueEntry,
)
from items.forms import ItemForm, ItemJournalForm, UnitOfMeasureForm
from items.models import Item, ItemCategory, UnitOfMeasure, TrackingSpecification
from items.filters import ItemFilter
from items.serializers import ItemSerializer, ItemListSerializer, ItemLedgerEntriesSerializer
from purchases.models import PurchaseInvoiceLine

from financials.models import GeneralLedgerEntry, G_LAccount

from postings.models import GeneralPostingSetup

from setup.models import UploadTemplates
from setup.enums import UploadTemplateChoices

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .serializers import (
    ItemCategorySerializer,
    UnitOfMeasureSerializer,
    TrackingSpecificationSerializer,
    ItemTrackingCodeSerializer,
    ItemUnitOfMeasureSerializer,
    ItemJournalSerializer,
    ItemImagesSerializer,
    ItemAttributeSerializer,
    ItemAttributeValueSerializer,
    ItemAttributeEntrySerializer,
    LocationSerializer,
)
from rest_framework.pagination import PageNumberPagination

from items.admin import ItemJournalPreviewProcessor
from items.posting import ItemJournalFinalPoster
from items.tasks import export_items_task

from rest_framework.parsers import MultiPartParser, FormParser


# ---------------- API endpoints ---------------- #
# class ItemFilter(filters.FilterSet):
#     q = filters.CharFilter(method="filter_q")

#     def filter_q(self, queryset, name, value):
#         return queryset.filter(Q(item_name__icontains=value) | Q(bar_code_no__icontains=value) | Q(system_id__icontains=value)| Q(no__icontains=value))


# class ItemsFilter(ListAPIView):
#     permission_classes = [IsAuthenticated]
#     serializer_class = ItemSerializer
#     authentication_classes = [SessionAuthentication]
#     queryset = Item.objects.all()
#     filterset_fields = ["item_name", "bar_code_no", "system_id"]
#     filterset_class = ItemFilter
#     ordering_fields = ["item_name", "bar_code_no", "system_id"]
#     ordering = ["item_name"]


#     def get_queryset(self):
#         queryset = super().get_queryset()
#         return queryset


class ItemUpdateApiView(UpdateAPIView):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer


class TrackingSpecificationViewSet(viewsets.ModelViewSet):
    queryset = TrackingSpecification.objects.all()
    serializer_class = TrackingSpecificationSerializer
    filterset_fields = ["purchase_invoice_line", "item_journal"]
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def filter_queryset(self, queryset):
        """Filter by source_batch (journal batch) when item_journal is present.
        Ensures Item Journal shows only ITEM specs, Stock Taking only PHYS. INV. specs."""
        queryset = super().filter_queryset(queryset)
        item_journal_id = self.request.query_params.get("item_journal")
        if item_journal_id:
            try:
                journal = ItemJournal.objects.get(pk=item_journal_id)
                if journal.journal_batch_id:
                    queryset = queryset.filter(
                        source_batch=journal.journal_batch_id
                    )
            except ItemJournal.DoesNotExist:
                pass
        return queryset

    @action(detail=False, methods=["get"], url_path="summary/(?P<line_id>[^/.]+)")
    def summary(self, request, line_id=None):
        if not line_id:
            return Response({"error": "Line ID is required"}, status=400)

        # Check if this is a purchase invoice line or item journal
        context_type = request.query_params.get("context_type", "purchase")

        try:
            if context_type == "purchase":
                # Get the purchase line
                purchase_line = PurchaseInvoiceLine.objects.get(id=line_id)
                expected_quantity = (
                    purchase_line.quantity
                    * purchase_line.item_unit_of_measure.quantity_per_unit
                )

                # Get specifications for this line
                specifications = self.queryset.filter(purchase_invoice_line=line_id)
            elif context_type == "adjustment":
                # Get the item journal
                item_journal = ItemJournal.objects.get(id=line_id)
                expected_quantity = (
                    item_journal.quantity
                    * item_journal.item_unit_of_measure.quantity_per_unit
                )

                # Get specifications for this journal, filtered by source_batch
                # so Item Journal shows only ITEM specs, Stock Taking only PHYS. INV. specs
                specifications = self.queryset.filter(item_journal=line_id)
                if item_journal.journal_batch_id:
                    specifications = specifications.filter(
                        source_batch=item_journal.journal_batch_id
                    )
            else:
                return Response({"error": "Invalid context type"}, status=400)

            total_quantity = (
                specifications.aggregate(total=Sum("quantity_base"))["total"] or 0
            )

            return Response(
                {
                    "expected_quantity": expected_quantity,
                    "total_quantity": total_quantity,
                    "remaining_quantity": expected_quantity - total_quantity,
                    "specifications_count": specifications.count(),
                }
            )
        except (PurchaseInvoiceLine.DoesNotExist, ItemJournal.DoesNotExist):
            return Response({"error": "Line not found"}, status=404)

    @action(detail=False, methods=["get"])
    def check_lot(self, request):
        lot_no = (request.query_params.get("lot_no") or "").strip()
        item_id = (request.query_params.get("item") or "").strip()

        if not lot_no or not item_id:
            return Response(
                {"error": "Both lot_no and item parameters are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Prefer the most recent item ledger entry for this item/lot.
            existing_entry = (
                ItemLedgerEntries.objects.filter(
                    lot_no__iexact=lot_no,
                    item=item_id,
                )
                .order_by("-posting_date", "-created_at")
                .first()
            )

            if existing_entry:
                expiry = existing_entry.expiry_date
                return Response(
                    {
                        "exists": True,
                        "expiry_date": expiry.isoformat() if expiry else None,
                        "lot_no": existing_entry.lot_no,
                        "remaining_quantity": existing_entry.remaining_quantity,
                        "location": (
                            existing_entry.location.description
                            if existing_entry.location
                            else None
                        ),
                    }
                )

            # Fall back to tracking specs (e.g. lot captured on an unposted purchase).
            existing_spec = (
                TrackingSpecification.objects.filter(
                    lot_no__iexact=lot_no,
                    item=item_id,
                )
                .exclude(expiry_date__isnull=True)
                .order_by("-created_at")
                .first()
            )
            if existing_spec:
                expiry = existing_spec.expiry_date
                return Response(
                    {
                        "exists": True,
                        "expiry_date": expiry.isoformat() if expiry else None,
                        "lot_no": existing_spec.lot_no,
                    }
                )

            return Response({"exists": False})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _validate_negative_adjustment_lot(self, payload, instance=None):
        """
        For negative adjustment journals, enforce that lot exists in available stock
        for the same item/location/branch and has enough remaining quantity.
        """
        item_journal_id = payload.get("item_journal")
        lot_no = payload.get("lot_no")
        quantity_base = payload.get("quantity_base")

        if instance is not None:
            if item_journal_id in (None, ""):
                item_journal_id = getattr(instance, "item_journal_id", None)
            if not lot_no:
                lot_no = getattr(instance, "lot_no", None)
            if quantity_base in (None, ""):
                quantity_base = getattr(instance, "quantity_base", None)

        if not item_journal_id:
            return

        try:
            journal = ItemJournal.objects.select_related("item", "location_code").get(
                pk=item_journal_id
            )
        except ItemJournal.DoesNotExist:
            raise ValidationError("Item journal not found for this tracking specification.")

        negative_codes = {
            EntryType.NegativeAdjustment.name,
            EntryType.NegativeAdjustment.value,
            "NegativeAdjustment",
            "Negative Adjustment",
        }
        if journal.entry_type not in negative_codes:
            return

        if not lot_no:
            raise ValidationError(
                "Lot No. is required for negative adjustment tracking entries."
            )

        qty = int(quantity_base or 0)
        if qty <= 0:
            raise ValidationError(
                "Quantity (Base) must be greater than zero for negative adjustment tracking entries."
            )

        lot_entries = ItemLedgerEntries.objects.filter(
            item=journal.item,
            lot_no=lot_no,
            remaining_quantity__gt=0,
        )
        if journal.location_code_id:
            lot_entries = lot_entries.filter(location_id=journal.location_code_id)
        if journal.global_dimension_1_id:
            lot_entries = lot_entries.filter(
                global_dimension_1_id=journal.global_dimension_1_id
            )

        available_qty = lot_entries.aggregate(total=Coalesce(Sum("remaining_quantity"), 0))[
            "total"
        ]
        if available_qty <= 0:
            raise ValidationError(
                "Lot does not exist or has no available quantity in this location/branch."
            )

        requested_qs = TrackingSpecification.objects.filter(
            item_journal_id=journal.id,
            lot_no=lot_no,
        )
        if instance is not None and getattr(instance, "id", None):
            requested_qs = requested_qs.exclude(id=instance.id)
        requested_other_qty = requested_qs.aggregate(
            total=Coalesce(Sum("quantity_base"), 0)
        )["total"]
        requested_total = int(requested_other_qty or 0) + qty
        if requested_total > int(available_qty):
            raise ValidationError(
                f"Selected lot '{lot_no}' has insufficient quantity. "
                f"Available: {int(available_qty)}, requested: {requested_total}."
            )

    def create(self, request, *args, **kwargs):
        try:
            self._validate_negative_adjustment_lot(request.data)
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response(
                {"error": e.messages if hasattr(e, "messages") else [str(e)]},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def update(self, request, *args, **kwargs):
        try:
            self._validate_negative_adjustment_lot(
                request.data, instance=self.get_object()
            )
            return super().update(request, *args, **kwargs)
        except ValidationError as e:
            return Response(
                {"error": e.messages if hasattr(e, "messages") else [str(e)]},
                status=status.HTTP_400_BAD_REQUEST,
            )


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = "page_size"
    max_page_size = 1000


def _item_inventory_total_expr(branch=None):
    """Sum remaining quantity from item ledger entries, optionally scoped to branch."""
    if branch:
        return Coalesce(
            Sum(
                Case(
                    When(
                        itemledgerentries__global_dimension_1_id=branch.id,
                        then=F("itemledgerentries__remaining_quantity"),
                    ),
                    default=Value(0),
                )
            ),
            0,
            output_field=IntegerField(),
        )
    return Coalesce(
        Sum("itemledgerentries__remaining_quantity"),
        0,
        output_field=IntegerField(),
    )


class ItemsModalViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated]
    # SimpleJWT before session: Bearer requests authenticate as JWT (no CSRF on POST).
    # JWTAuthenticationWithRevocationChecks matches SimpleJWT behavior but enforces
    # token_valid_after (admin force-logout); CustomJWTAuthentication adds extra tenant checks.
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = ItemFilter
    search_fields = ["item_name", "description", "bar_code_no", "shelf_no", "no"]
    ordering_fields = [
        "item_name",
        "no",
        "item_category",
        "type",
        "unit_price",
        "inventory",
        "created_at",
        "updated_at",
    ]
    ordering = ["item_name"]
    lookup_field = "system_id"
    pagination_class = StandardResultsSetPagination

    def get_serializer_class(self):
        if getattr(self, "action", None) == "list":
            return ItemListSerializer
        return ItemSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        # Branch for inventory/expiry filters when multi-branch enabled
        branch = None
        branch_location = None
        try:
            from financials.models import GeneralLedgerSetup

            gl_setup = GeneralLedgerSetup.objects.first()
            if gl_setup and getattr(gl_setup, "enable_multiple_branches", False):
                branch = get_branch_for_request(self.request)
                if not branch:
                    branch = getattr(self.request.user, "global_dimension_1", None)
                if branch:
                    # Zentro convention: Location.code matches Branch DimensionValue.code
                    branch_location = Location.objects.filter(
                        code__iexact=getattr(branch, "code", "")
                    ).first()
        except ImportError:
            pass

        # Annotate a context-aware unit cost for list payloads to avoid N+1 queries.
        if getattr(self, "action", None) == "list":
            if branch:
                # Prefer branch+location latest cost, fallback to branch-wide latest cost.
                # Do not require on-hand stock; buying price should still show from the
                # latest cost entry even when remaining quantity is currently zero.
                ve_branch_qs = ValueEntry.objects.filter(
                    item=OuterRef("pk"),
                    global_dimension_1_id=branch.id,
                )

                if branch_location:
                    ve_loc_qs = ve_branch_qs.filter(location_code_id=branch_location.id)
                    queryset = queryset.annotate(
                        _unit_cost_context=Case(
                            When(
                                Exists(ve_loc_qs),
                                then=Subquery(
                                    ve_loc_qs.order_by("-created_at").values(
                                        "cost_per_unit"
                                    )[:1]
                                ),
                            ),
                            When(
                                Exists(ve_branch_qs),
                                then=Subquery(
                                    ve_branch_qs.order_by("-created_at").values(
                                        "cost_per_unit"
                                    )[:1]
                                ),
                            ),
                            default=Value(0),
                            output_field=IntegerField(),
                        )
                    )
                else:
                    queryset = queryset.annotate(
                        _unit_cost_context=Case(
                            When(
                                Exists(ve_branch_qs),
                                then=Subquery(
                                    ve_branch_qs.order_by("-created_at").values(
                                        "cost_per_unit"
                                    )[:1]
                                ),
                            ),
                            default=Value(0),
                            output_field=IntegerField(),
                        )
                    )
            else:
                latest_ve = ValueEntry.objects.filter(item=OuterRef("pk")).order_by(
                    "-created_at"
                )
                queryset = queryset.annotate(
                    _unit_cost_context=Coalesce(
                        Subquery(
                            latest_ve.values("cost_per_unit")[:1],
                            output_field=DecimalField(max_digits=10, decimal_places=2),
                        ),
                        Value(Decimal("0.00")),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    )
                )

        # Handle legacy search parameter
        search_query = self.request.query_params.get("q", None)
        if search_query:
            search_query = search_query.strip()
            queryset = queryset.filter(
                Q(item_name__icontains=search_query)
                | Q(bar_code_no__icontains=search_query)
                | Q(no__icontains=search_query)
            )

        # Handle advanced filtering for properties that can't be filtered directly
        unit_cost_min = self.request.query_params.get("unit_cost_min")
        unit_cost_max = self.request.query_params.get("unit_cost_max")
        inventory_min = self.request.query_params.get("inventory_min")
        inventory_max = self.request.query_params.get("inventory_max")
        profit_percentage_min = self.request.query_params.get("profit_percentage_min")
        profit_percentage_max = self.request.query_params.get("profit_percentage_max")

        # Inventory quantity filters (annotate once when any filter needs on-hand total)
        inventory_status = self.request.query_params.get("inventory_status")
        low_stock_param = self.request.query_params.get("low_stock")
        low_stock = str(low_stock_param).lower() in ("true", "1")
        needs_inventory_total = (
            inventory_status in ["IN_STOCK", "OUT_OF_STOCK"]
            or low_stock
            or inventory_min is not None
            or inventory_max is not None
        )
        if needs_inventory_total:
            queryset = queryset.annotate(
                inventory_total=_item_inventory_total_expr(branch)
            )
            if inventory_status == "IN_STOCK":
                queryset = queryset.filter(inventory_total__gt=0)
            elif inventory_status == "OUT_OF_STOCK":
                queryset = queryset.filter(inventory_total__lte=0)
            if low_stock:
                queryset = queryset.filter(
                    type=InventoryType.Inventory.value,
                    minimum_stock__isnull=False,
                    minimum_stock__gt=0,
                    inventory_total__gt=0,
                    inventory_total__lte=F("minimum_stock"),
                )
            if inventory_min is not None:
                try:
                    queryset = queryset.filter(
                        inventory_total__gte=int(inventory_min)
                    )
                except (ValueError, TypeError):
                    pass
            if inventory_max is not None:
                try:
                    queryset = queryset.filter(
                        inventory_total__lte=int(inventory_max)
                    )
                except (ValueError, TypeError):
                    pass

        # Expiry date range filtering from Item Ledger Entries
        expiry_from = self.request.query_params.get("expiry_from")
        expiry_to = self.request.query_params.get("expiry_to")
        if expiry_from or expiry_to:
            expiry_filter = Q()
            if branch:
                expiry_filter &= Q(itemledgerentries__global_dimension_1_id=branch.id)
            if expiry_from:
                expiry_filter &= Q(itemledgerentries__expiry_date__gte=expiry_from)
            if expiry_to:
                expiry_filter &= Q(itemledgerentries__expiry_date__lte=expiry_to)
            queryset = queryset.filter(expiry_filter).distinct()

        # Branch filter for multi-branch (Items are shared; filter is no-op for Item)
        queryset = filter_queryset_by_branch(
            queryset, self.request.user, Item, self.request
        )

        if getattr(self, "action", None) == "list":
            queryset = queryset.select_related(
                "item_category",
                "unit_of_measure",
                "tracking_code",
            ).prefetch_related(
                Prefetch(
                    "itemimages_set",
                    queryset=ItemImages.objects.order_by("-created_at"),
                ),
                Prefetch(
                    "itemunitofmeasure_set",
                    queryset=ItemUnitOfMeasure.objects.select_related(
                        "unit_of_measure"
                    ),
                ),
            )
            ann = getattr(queryset.query, "annotations", None) or {}
            if "inventory_total" not in ann:
                if branch:
                    inv = Coalesce(
                        Sum(
                            Case(
                                When(
                                    itemledgerentries__global_dimension_1_id=branch.id,
                                    then=F("itemledgerentries__remaining_quantity"),
                                ),
                                default=Value(0),
                            )
                        ),
                        0,
                        output_field=IntegerField(),
                    )
                else:
                    inv = Coalesce(
                        Sum("itemledgerentries__remaining_quantity"),
                        0,
                        output_field=IntegerField(),
                    )
                queryset = queryset.annotate(_list_inventory=inv)

        return queryset.order_by("item_name")

    @action(detail=False, methods=["post"])
    def export(self, request):
        """
        Trigger background task to export items to Excel or PDF format.
        Access matches list: any authenticated user who can list items may export
        (same data as the grid; column sensitivity still uses UserSetup flags in the task).
        Returns task_id for status checking.
        """
        export_format = request.data.get("format", "excel").lower()
        if export_format not in ["excel", "pdf"]:
            return Response(
                {"error": "Invalid format. Supported formats: excel, pdf"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get filter parameters from request
        filters_data = request.data.get("filters", {})

        # Get item IDs if specified (for selective export)
        item_ids = request.data.get("item_ids", None)

        # Get current schema name
        schema_name = connection.schema_name

        # Get user permissions for export
        from authentication.models import UserSetup

        user_permissions = {}
        try:
            user_setup = UserSetup.objects.get(user=request.user)
            user_permissions = {
                "can_see_buying_price": user_setup.can_see_buying_price,
                "can_see_profit_margin": user_setup.can_see_profit_margin,
                "can_see_item_cost": user_setup.can_see_item_cost,
            }
        except UserSetup.DoesNotExist:
            # Default to all permissions if no setup exists
            user_permissions = {
                "can_see_buying_price": True,
                "can_see_profit_margin": True,
                "can_see_item_cost": True,
            }

        # Trigger background task
        task = export_items_task.delay(
            item_ids=item_ids,
            export_format=export_format,
            filters_data=filters_data,
            schema_name=schema_name,
            user_permissions=user_permissions,
        )

        return Response(
            {
                "task_id": task.id,
                "message": "Export task started. Use the task_id to check status.",
                "status": "pending",
            },
            status=status.HTTP_202_ACCEPTED,
        )

    @action(detail=False, methods=["get"])
    def export_status(self, request):
        """
        Check the status of an export task and download file when ready
        Query params: task_id
        """
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response(
                {"error": "task_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        task = AsyncResult(task_id)

        if task.state == PENDING:
            response = {
                "state": task.state,
                "progress": 0,
                "message": "Task is pending...",
                "status": "pending",
            }
        elif task.state == FAILURE:
            response = {
                "state": task.state,
                "progress": 0,
                "message": str(task.info) if task.info else "Task failed",
                "status": "failed",
            }
        elif task.state == SUCCESS:
            # Task completed, check if file is ready
            result = task.info if task.info else {}

            # Get current schema name for cache key
            schema_name = connection.schema_name
            cache_key = f"export_file_{schema_name}_{task_id}"

            # Try to get file from cache
            file_data = cache.get(cache_key)

            if file_data:
                try:
                    import logging

                    logger = logging.getLogger(__name__)

                    logger.info(
                        f"Retrieved file data from cache. Base64 length: {len(file_data['file_data'])} characters"
                    )

                    # Decode base64 file data
                    file_bytes = base64.b64decode(file_data["file_data"])

                    logger.info(
                        f"Decoded file: {len(file_bytes)} bytes, First 10 bytes (hex): {file_bytes[:10].hex() if len(file_bytes) >= 10 else file_bytes.hex()}"
                    )

                    # Validate file data is not empty
                    if not file_bytes or len(file_bytes) == 0:
                        return Response(
                            {
                                "state": task.state,
                                "progress": 100,
                                "message": "Export file is empty",
                                "status": "error",
                            },
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        )

                    # Validate Excel file signature (first 4 bytes should be PK\x03\x04 for ZIP/Office files)
                    if (
                        file_data["content_type"]
                        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ):
                        if len(file_bytes) < 4 or file_bytes[:2] != b"PK":
                            import logging

                            logger = logging.getLogger(__name__)
                            logger.error(
                                f"Invalid Excel file signature. File size: {len(file_bytes)}, First bytes: {file_bytes[:10].hex() if len(file_bytes) >= 10 else file_bytes.hex()}"
                            )
                            return Response(
                                {
                                    "state": task.state,
                                    "progress": 100,
                                    "message": "Invalid Excel file format - file may be corrupted",
                                    "status": "error",
                                },
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            )

                    # Return file as download
                    response = HttpResponse(
                        file_bytes,
                        content_type=file_data["content_type"],
                    )
                    response["Content-Disposition"] = (
                        f'attachment; filename="{file_data["filename"]}"'
                    )
                    response["Content-Length"] = str(len(file_bytes))

                    # Clean up cache after download
                    cache.delete(cache_key)

                    return response
                except Exception as e:
                    # If decoding fails, return error
                    import logging

                    logger = logging.getLogger(__name__)
                    logger.error(
                        f"Error processing export file: {str(e)}", exc_info=True
                    )
                    return Response(
                        {
                            "state": task.state,
                            "progress": 100,
                            "message": f"Error processing file: {str(e)}",
                            "status": "error",
                        },
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )
            else:
                # File not in cache - check if task result has the info
                if result and result.get("cache_key"):
                    # Try alternative cache key from task result
                    alt_cache_key = result.get("cache_key")
                    file_data = cache.get(alt_cache_key)
                    if file_data:
                        try:
                            file_bytes = base64.b64decode(file_data["file_data"])
                            response = HttpResponse(
                                file_bytes,
                                content_type=file_data["content_type"],
                            )
                            response["Content-Disposition"] = (
                                f'attachment; filename="{file_data["filename"]}"'
                            )
                            cache.delete(alt_cache_key)
                            return response
                        except Exception as e:
                            pass

                # File not found in cache - try alternative keys
                import logging

                logger = logging.getLogger(__name__)

                # Try cache key from task result if available
                if result and result.get("cache_key"):
                    alt_cache_key = result.get("cache_key")
                    file_data = cache.get(alt_cache_key)
                    if file_data:
                        try:
                            file_bytes = base64.b64decode(file_data["file_data"])
                            response = HttpResponse(
                                file_bytes,
                                content_type=file_data["content_type"],
                            )
                            response["Content-Disposition"] = (
                                f'attachment; filename="{file_data["filename"]}"'
                            )
                            cache.delete(alt_cache_key)
                            return response
                        except Exception as e:
                            logger.error(
                                f"Error decoding file from alt cache key: {str(e)}"
                            )

                # Try without schema prefix (for backward compatibility)
                alt_cache_key = f"export_file_{task_id}"
                file_data = cache.get(alt_cache_key)
                if file_data:
                    try:
                        file_bytes = base64.b64decode(file_data["file_data"])
                        response = HttpResponse(
                            file_bytes,
                            content_type=file_data["content_type"],
                        )
                        response["Content-Disposition"] = (
                            f'attachment; filename="{file_data["filename"]}"'
                        )
                        cache.delete(alt_cache_key)
                        return response
                    except Exception as e:
                        logger.error(
                            f"Error decoding file from backward compat cache key: {str(e)}"
                        )
                file_data = cache.get(alt_cache_key)
                if file_data:
                    try:
                        file_bytes = base64.b64decode(file_data["file_data"])
                        response = HttpResponse(
                            file_bytes,
                            content_type=file_data["content_type"],
                        )
                        response["Content-Disposition"] = (
                            f'attachment; filename="{file_data["filename"]}"'
                        )
                        cache.delete(alt_cache_key)
                        logger.info(
                            f"Found file using alternative cache key: {alt_cache_key}"
                        )
                        return response
                    except Exception as e:
                        logger.error(
                            f"Error decoding file from alternative cache: {str(e)}"
                        )

                logger.warning(
                    f"Export file not found in cache for task_id: {task_id}, schema: {schema_name}, cache_key: {cache_key}"
                )

                response = {
                    "state": task.state,
                    "progress": 100,
                    "message": f"Export completed but file not found in cache. Please try exporting again.",
                    "status": "completed",
                }
        else:
            # PROGRESS or STARTED state
            if task.info:
                response = {
                    "state": task.state,
                    "progress": task.info.get("progress", 0),
                    "message": task.info.get("message", "Processing..."),
                    "status": task.info.get("status", "processing"),
                }
            else:
                response = {
                    "state": task.state,
                    "progress": 0,
                    "message": "Task is processing...",
                    "status": "processing",
                }

        return Response(response, status=status.HTTP_200_OK)

    def _export_to_excel(self, items, user_permissions=None):
        """Export items to Excel format with user permission filtering"""
        try:
            import xlsxwriter
            from authentication.models import UserSetup

            # Get user permissions
            can_see_buying_price = True
            can_see_profit = True

            if user_permissions:
                can_see_buying_price = user_permissions.get(
                    "can_see_buying_price", True
                )
                can_see_profit = user_permissions.get("can_see_profit_margin", True)

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet("Items")

            # Define formats
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#366092",
                    "font_color": "white",
                    "border": 1,
                    "align": "center",
                    "valign": "vcenter",
                }
            )
            data_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})

            # Define headers based on permissions
            headers = [
                "Item No",
                "Item Name",
                "Description",
                "Category",
                "Type",
                "Unit of Measure",
                "Unit Price",
            ]

            # Add Unit Cost column only if user has permission
            if can_see_buying_price:
                headers.append("Unit Cost")

            # Add Profit Margin column only if user has permission
            if can_see_profit:
                headers.append("Profit Margin %")

            headers.extend(
                [
                    "Inventory",
                    "Bar Code",
                    "Shelf No",
                    "Status",
                ]
            )

            # Write headers
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                worksheet.set_column(col, col, 15)

            # Write data
            for row_idx, item in enumerate(items, start=1):
                col = 0
                worksheet.write(row_idx, col, item.no or "", data_format)
                col += 1
                worksheet.write(row_idx, col, item.item_name or "", data_format)
                col += 1
                worksheet.write(row_idx, col, item.description or "", data_format)
                col += 1
                worksheet.write(
                    row_idx,
                    col,
                    item.item_category.description if item.item_category else "",
                    data_format,
                )
                col += 1
                worksheet.write(row_idx, col, item.type or "", data_format)
                col += 1
                worksheet.write(row_idx, col, item.unit_of_measure or "", data_format)
                col += 1
                worksheet.write(
                    row_idx,
                    col,
                    float(item.unit_price) if item.unit_price else 0,
                    number_format,
                )
                col += 1

                # Write Unit Cost only if user has permission
                if can_see_buying_price:
                    worksheet.write(
                        row_idx,
                        col,
                        float(item.unit_cost) if item.unit_cost else 0,
                        number_format,
                    )
                    col += 1

                # Write Profit Margin only if user has permission
                if can_see_profit:
                    profit = item.profit_percentage or 0
                    worksheet.write(
                        row_idx,
                        col,
                        float(profit),
                        number_format,
                    )
                    col += 1

                worksheet.write(
                    row_idx,
                    col,
                    float(item.inventory) if item.inventory else 0,
                    number_format,
                )
                col += 1
                worksheet.write(row_idx, col, item.bar_code_no or "", data_format)
                col += 1
                worksheet.write(row_idx, col, item.shelf_no or "", data_format)
                col += 1
                worksheet.write(
                    row_idx,
                    col,
                    "Active" if not item.blocked else "Blocked",
                    data_format,
                )

            workbook.close()
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="items_export.xlsx"'
            return response

        except ImportError as e:
            return Response(
                {"error": f"xlsxwriter library is not installed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as e:
            import traceback

            error_trace = traceback.format_exc()
            print(f"Excel export error: {error_trace}")  # Log for debugging
            return Response(
                {"error": f"Failed to export to Excel: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _export_to_pdf(self, items):
        """Export items to PDF format"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
            from reportlab.lib import colors
            from reportlab.lib.units import mm

            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph("<b>Items Export</b>", styles["Title"])
            elements.append(title)
            elements.append(Spacer(1, 12))

            # Create table data
            data = [
                [
                    "Item No",
                    "Item Name",
                    "Category",
                    "Type",
                    "Unit Price",
                    "Inventory",
                    "Status",
                ]
            ]

            for item in items[:500]:  # Limit to 500 items for PDF
                data.append(
                    [
                        item.no or "",
                        item.item_name or "",
                        item.item_category.description if item.item_category else "",
                        item.type or "",
                        (
                            f"UGX {item.unit_price:,.2f}"
                            if item.unit_price
                            else "UGX 0.00"
                        ),
                        str(item.inventory) if item.inventory else "0",
                        "Active" if not item.blocked else "Blocked",
                    ]
                )

            # Create table
            table = Table(data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ]
                )
            )

            elements.append(table)
            doc.build(elements)
            output.seek(0)

            response = HttpResponse(output.read(), content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="items_export.pdf"'
            return response

        except ImportError as e:
            return Response(
                {"error": f"reportlab library is not installed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception as e:
            import traceback

            error_trace = traceback.format_exc()
            print(f"PDF export error: {error_trace}")  # Log for debugging
            return Response(
                {"error": f"Failed to export to PDF: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def upsert(self, request):
        system_id = request.data.get("system_id")
        item_name = request.data.get("item_name")
        item_id = request.data.get("item_id")
        deleted = request.data.get("deleted", False)  # Add deleted flag

        try:
            with transaction.atomic():
                if system_id or item_id:
                    # Handle deletion
                    if deleted:
                        item = (
                            Item.objects.filter(system_id=system_id).first()
                            or Item.objects.filter(id=item_id).first()
                        )
                        if not item:
                            raise Item.DoesNotExist
                        item.delete()
                        return Response(
                            {"message": "Item deleted successfully"},
                            status=status.HTTP_200_OK,
                        )

                    # Update existing item
                    item = (
                        Item.objects.filter(system_id=system_id).first()
                        or Item.objects.filter(id=item_id).first()
                    )
                    if not item:
                        raise Item.DoesNotExist

                    # if unit_of_measure is included in the data sent, we will create a new item_unit_of_measure for the
                    # item if it does not exist. If it does exist, we will set it as the default and update all other
                    # related records to have default set to false at the database level
                    if request.data.get("unit_of_measure"):
                        unit_of_measure_code = request.data.get("unit_of_measure")
                        # Convert string code to UnitOfMeasure instance
                        # Handle both string code and instance
                        if isinstance(unit_of_measure_code, str):
                            unit_of_measure_obj, _ = (
                                UnitOfMeasure.objects.get_or_create(
                                    code=unit_of_measure_code
                                )
                            )
                        else:
                            # If it's already an instance or ID, use it directly
                            unit_of_measure_obj = unit_of_measure_code

                        item_unit_of_measure = ItemUnitOfMeasure.objects.filter(
                            item=item,
                            unit_of_measure=unit_of_measure_obj,
                        ).first()

                        if not item_unit_of_measure:
                            # Create new ItemUnitOfMeasure if it doesn't exist
                            item_unit_of_measure = ItemUnitOfMeasure.objects.create(
                                item=item,
                                unit_of_measure=unit_of_measure_obj,
                                quantity_per_unit=1,  # Default quantity per unit
                                default=True,
                            )
                            request.data["sales_unit_of_measure"] = (
                                item_unit_of_measure.id
                            )
                            request.data["purchase_unit_of_measure"] = (
                                item_unit_of_measure.id
                            )
                        else:
                            # If it exists, set it as default (this will automatically unset other defaults)
                            # Ensure quantity_per_unit is 1 for default ItemUnitOfMeasure
                            item_unit_of_measure.default = True
                            item_unit_of_measure.quantity_per_unit = 1
                            request.data["sales_unit_of_measure"] = (
                                item_unit_of_measure.id
                            )
                            request.data["purchase_unit_of_measure"] = (
                                item_unit_of_measure.id
                            )
                            item_unit_of_measure.save()

                    serializer = self.get_serializer(
                        item, data=request.data, partial=True
                    )
                else:
                    # Create new item - require item_name
                    if not item_name:
                        raise ValidationError({"item_name": "Item name is required"})
                    serializer = self.get_serializer(data=request.data)

                if serializer.is_valid(raise_exception=True):
                    item = serializer.save()
                    print("serializer.data", serializer.data)
                    return Response(
                        serializer.data,
                        status=(
                            status.HTTP_200_OK if system_id else status.HTTP_201_CREATED
                        ),
                    )

        except ValidationError as e:
            # Return the error as a string or dict, not e.detail
            if hasattr(e, "message_dict"):
                return Response(
                    {"detail": e.message_dict}, status=status.HTTP_400_BAD_REQUEST
                )
            else:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Item.DoesNotExist:
            return Response(
                {"detail": "Item not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"Error in upsert: {str(e)}")
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def import_template(self, request):
        """
        Generate a dynamic Excel template for item imports.
        Populates dropdown validations from the tenant's actual data
        (UOMs, categories, types) so lay users can pick valid values.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()

        # --- Reference sheet (hidden, powers dropdowns) ---
        ref_ws = wb.create_sheet("Reference")
        uom_codes = list(
            UnitOfMeasure.objects.values_list("code", flat=True).order_by("code")
        )
        category_codes = list(
            ItemCategory.objects.values_list("code", flat=True).order_by("code")
        )
        type_choices = ["Inventory", "Service", "Non-Inventory"]

        ref_ws["A1"] = "Unit of Measure"
        for i, code in enumerate(uom_codes, start=2):
            ref_ws[f"A{i}"] = code

        ref_ws["B1"] = "Item Category"
        for i, code in enumerate(category_codes, start=2):
            ref_ws[f"B{i}"] = code

        ref_ws["C1"] = "Type"
        for i, val in enumerate(type_choices, start=2):
            ref_ws[f"C{i}"] = val

        ref_ws.sheet_state = "hidden"

        query = getattr(request, "query_params", request.GET)
        import_mode = (query.get("import_mode", "standard") or "standard").strip().lower()
        include_opening_balance = import_mode == "opening_balance"

        # --- Items sheet (main data entry) ---
        ws = wb.active
        ws.title = "Items"

        columns = [
            {"header": "Item Name", "width": 30, "required": True},
            {"header": "Type", "width": 18, "required": True},
            {"header": "Unit of Measure", "width": 20, "required": True},
            {"header": "Unit Price", "width": 15, "required": True},
        ]
        if include_opening_balance:
            columns.extend(
                [
                    {"header": "Quantity", "width": 12, "required": False},
                    {
                        "header": "Unit Cost",
                        "width": 14,
                        "required": False,
                    },
                ]
            )
        columns.extend(
            [
                {"header": "Item Category", "width": 20, "required": False},
                {"header": "Bar Code No", "width": 20, "required": False},
                {"header": "Description", "width": 40, "required": False},
                {"header": "Shelf No", "width": 15, "required": False},
            ]
        )

        header_fill_req = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        header_fill_opt = PatternFill(
            start_color="E2E2E2", end_color="E2E2E2", fill_type="solid"
        )
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="999999")
        )
        example_font = Font(italic=True, color="999999")

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = header_font
            cell.fill = header_fill_req if col_def["required"] else header_fill_opt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Example rows
        if include_opening_balance:
            examples = [
                [
                    "Sugar 1kg",
                    "Inventory",
                    uom_codes[0] if uom_codes else "PCS",
                    5000,
                    100,
                    4200,
                    category_codes[0] if category_codes else "",
                    "",
                    "White sugar 1kg bag",
                    "A1",
                ],
                [
                    "Plumbing Service",
                    "Service",
                    uom_codes[0] if uom_codes else "PCS",
                    25000,
                    "",
                    "",
                    "",
                    "Basic plumbing service",
                    "",
                ],
            ]
        else:
            examples = [
                [
                    "Sugar 1kg",
                    "Inventory",
                    uom_codes[0] if uom_codes else "PCS",
                    5000,
                    category_codes[0] if category_codes else "",
                    "",
                    "White sugar 1kg bag",
                    "A1",
                ],
                [
                    "Plumbing Service",
                    "Service",
                    uom_codes[0] if uom_codes else "PCS",
                    25000,
                    "",
                    "",
                    "Basic plumbing service",
                    "",
                ],
            ]
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, val in enumerate(example, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = example_font

        # Data validations (dropdown lists from Reference sheet)
        max_rows = 5000
        if type_choices:
            dv_type = DataValidation(
                type="list",
                formula1=f"Reference!$C$2:$C${len(type_choices) + 1}",
                allow_blank=True,
            )
            dv_type.error = "Please select a valid Type"
            dv_type.errorTitle = "Invalid Type"
            ws.add_data_validation(dv_type)
            dv_type.add(f"B4:B{max_rows}")

        if uom_codes:
            dv_uom = DataValidation(
                type="list",
                formula1=f"Reference!$A$2:$A${len(uom_codes) + 1}",
                allow_blank=True,
            )
            dv_uom.error = "Please select a valid Unit of Measure"
            dv_uom.errorTitle = "Invalid UOM"
            ws.add_data_validation(dv_uom)
            dv_uom.add(f"C4:C{max_rows}")

        cat_col_letter = "G" if include_opening_balance else "E"
        if category_codes:
            dv_cat = DataValidation(
                type="list",
                formula1=f"Reference!$B$2:$B${len(category_codes) + 1}",
                allow_blank=True,
            )
            dv_cat.error = "Please select a valid Item Category"
            dv_cat.errorTitle = "Invalid Category"
            ws.add_data_validation(dv_cat)
            dv_cat.add(f"{cat_col_letter}4:{cat_col_letter}{max_rows}")

        # --- Instructions sheet ---
        instr_ws = wb.create_sheet("Instructions")
        template_title = (
            "ZentroApp - Item Import Template (Opening Balances)"
            if include_opening_balance
            else "ZentroApp - Item Import Template"
        )
        instructions = [
            [template_title],
            [""],
            ["How to use this template:"],
            ["1. Fill in your items starting from row 4 on the 'Items' sheet"],
            ["2. Green columns are REQUIRED - you must fill these in"],
            ["3. Grey columns are OPTIONAL - leave blank if not needed"],
            ["4. Use the dropdown lists where available (Type, Unit of Measure, Category)"],
            ["5. You may delete the example rows (rows 2-3) before importing"],
            [
                "6. Save the file and upload it on the Items page"
                + (
                    " (select 'Items with opening balances' when importing)"
                    if include_opening_balance
                    else ""
                )
            ],
            [""],
            ["Column Descriptions:"],
            ["Item Name", "The name of the item (must be unique)"],
            ["Type", "Inventory = physical stock, Service = labour/service, Non-Inventory = consumable"],
            ["Unit of Measure", "The base unit (e.g. PCS, KG, LTR). Must already exist in the system."],
            [
                "Unit Price",
                "Selling price on the item card (retail/list price).",
            ],
        ]
        if include_opening_balance:
            instructions.extend(
                [
                    [
                        "Quantity",
                        "Optional initial on-hand qty for Inventory items only. "
                        "Blank or zero = item created without a journal. "
                        "After import, review journals on Opening Balance Adjustment.",
                    ],
                    [
                        "Unit Cost",
                        "Buying/cost per unit for the opening balance journal (maps to "
                        "Buying Price / Unit Amount on the journal). If blank, Unit Price "
                        "is used, then the item's calculated unit cost if available.",
                    ],
                ]
            )
        instructions.extend(
            [
                ["Item Category", "Optional grouping category. Must already exist in the system."],
                ["Bar Code No", "Optional barcode. Leave blank to auto-generate."],
                ["Description", "Optional longer description of the item"],
                ["Shelf No", "Optional shelf/location reference for the item"],
            ]
        )
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                cell = instr_ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 3:
                    cell.font = Font(bold=True, size=12)
                elif row_idx == 11:
                    cell.font = Font(bold=True, size=12)
                elif row_idx >= 12 and col_idx == 1:
                    cell.font = Font(bold=True)
        instr_ws.column_dimensions["A"].width = 25
        instr_ws.column_dimensions["B"].width = 70

        # Write to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = (
            "item_import_template_opening_balance.xlsx"
            if include_opening_balance
            else "item_import_template.xlsx"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_items(self, request):
        """
        Import items from an uploaded Excel file (background task).
        Accepts multipart/form-data with a 'file' field.
        Returns task_id for status polling.
        """
        from items.tasks import process_item_import
        import uuid as _uuid

        has_permission, source = request.user.check_object_permission(10201, "insert")
        if not has_permission:
            return Response(
                {"error": "You do not have permission to import items"},
                status=status.HTTP_403_FORBIDDEN,
            )

        if "file" not in request.FILES:
            return Response(
                {"detail": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uploaded_file = request.FILES["file"]
            temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
            os.makedirs(temp_dir, exist_ok=True)

            import_id = _uuid.uuid4()
            file_ext = os.path.splitext(uploaded_file.name)[1]
            filename = f"item_import_{import_id}{file_ext}"
            file_path = os.path.join(temp_dir, filename)

            with open(file_path, "wb+") as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            schema_name = connection.schema_name
            import_mode = (
                request.data.get("import_mode", "standard") or "standard"
            ).strip().lower()
            if import_mode not in ("standard", "opening_balance"):
                import_mode = "standard"
            branch = get_branch_for_request(request)

            task = process_item_import.delay(
                df_json=None,
                schema_name=schema_name,
                file_path=file_path,
                user_id=request.user.id,
                branch_id=getattr(branch, "id", None),
                import_mode=import_mode,
            )

            return Response(
                {
                    "task_id": task.id,
                    "message": "Import started. Use the task_id to check status.",
                    "status": "pending",
                },
                status=status.HTTP_202_ACCEPTED,
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def import_items_status(self, request):
        """
        Poll the status of an item import Celery task.
        Query params: task_id
        """
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response(
                {"detail": "task_id is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            task_result = AsyncResult(task_id)

            if task_result.state == PENDING:
                response = {
                    "task_id": task_id,
                    "status": "pending",
                    "message": "Task is waiting to be processed",
                    "progress": 0,
                }
            elif task_result.state == "PROGRESS":
                meta = task_result.info or {}
                response = {
                    "task_id": task_id,
                    "status": "in_progress",
                    "message": meta.get("message", "Processing..."),
                    "progress": meta.get("progress", 0),
                }
            elif task_result.state == SUCCESS:
                result = task_result.result or {}
                response = {
                    "task_id": task_id,
                    "status": "success",
                    "message": result.get("message", "Import completed"),
                    "progress": 100,
                    "created_count": result.get("items_created", 0),
                    "updated_count": result.get("items_updated", 0),
                    "failed_count": result.get("failed_count", 0),
                    "total_rows": result.get("total_rows", 0),
                    "errors": result.get("errors", []),
                    "journals_created": result.get("journals_created", 0),
                    "journal_document_nos": result.get("journal_document_nos", []),
                    "warnings": result.get("warnings", []),
                }
            elif task_result.state == FAILURE:
                response = {
                    "task_id": task_id,
                    "status": "failure",
                    "message": str(task_result.info),
                    "error": str(task_result.info),
                    "progress": 0,
                }
            else:
                response = {
                    "task_id": task_id,
                    "status": task_result.state.lower(),
                    "message": f"Task state: {task_result.state}",
                    "progress": 0,
                }

            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"detail": f"Error checking task status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def generate_barcode(self):
        """Generate a unique 13-digit barcode"""
        import random

        while True:
            barcode = "".join([str(random.randint(0, 9)) for _ in range(13)])
            if not Item.objects.filter(bar_code_no=barcode).exists():
                return barcode


class ItemLedgerViewSet(viewsets.ModelViewSet):
    queryset = ItemLedgerEntries.objects.all()
    serializer_class = ItemLedgerEntriesSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    lookup_field = "item__no"  # Use the item's no field for lookup
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        # Get item_no from pk (URL parameter) since lookup_field is item__no
        item_no = self.kwargs.get("pk") or self.kwargs.get("item__no")
        if not item_no:
            return ItemLedgerEntries.objects.none()

        # Try to find by item__no first, if that fails, try by system_id
        queryset = ItemLedgerEntries.objects.filter(item__no=item_no)

        # If no results by item__no, try by system_id (in case frontend passes system_id)
        if not queryset.exists():
            try:
                from items.models import Item

                item = Item.objects.filter(system_id=item_no).first()
                if item:
                    queryset = ItemLedgerEntries.objects.filter(item=item)
            except Exception:
                pass

        queryset = filter_queryset_by_branch(
            queryset, self.request.user, ItemLedgerEntries, self.request
        )
        return queryset.order_by("-created_at")

    def retrieve(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Log for debugging
        item_no = self.kwargs.get("pk") or self.kwargs.get("item__no")
        total_count = queryset.count()
        entries_with_lot = (
            queryset.exclude(lot_no__isnull=True).exclude(lot_no="").count()
        )

        import logging

        logger = logging.getLogger(__name__)
        logger.info(
            f"ItemLedgerViewSet.retrieve: item_no={item_no}, total_entries={total_count}, entries_with_lot={entries_with_lot}"
        )

        # Get summary data from all entries (not just current page)
        summary = queryset.aggregate(
            total_quantity=Sum("quantity"), total_remaining=Sum("remaining_quantity")
        )

        # Paginate the queryset
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            # Add summary to paginated response
            response.data["summary"] = {
                "total_quantity": summary["total_quantity"] or 0,
                "total_remaining": summary["total_remaining"] or 0,
            }
            # Ensure both results and ledger_entries are available for frontend compatibility
            if "results" in response.data and "ledger_entries" not in response.data:
                response.data["ledger_entries"] = response.data["results"]
            return response

        # Fallback if pagination is not used
        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "ledger_entries": serializer.data,
                "results": serializer.data,  # Add results for consistency
                "summary": {
                    "total_quantity": summary["total_quantity"] or 0,
                    "total_remaining": summary["total_remaining"] or 0,
                },
            }
        )

    @action(detail=False, methods=["get"], url_path="availability-by-lot")
    def availability_by_lot(self, request):
        """
        Return item availability grouped by lot number (and expiry date).
        Used for Stock Taking / Physical Inventory to show "Item Availability by Lot No."
        Query params: item_id (required, item's no/code e.g. ITM-000001), location_id (optional).
        """
        item_no = request.query_params.get("item_id") or request.query_params.get("item_no")
        location_id = request.query_params.get("location_id")

        if not item_no:
            return Response(
                {"error": "item_id or item_no is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Item uses no (string) as primary key
        item = get_object_or_404(Item, no=item_no)

        qs = (
            ItemLedgerEntries.objects.filter(
                item=item,
                remaining_quantity__gt=0,
            )
            .exclude(lot_no__isnull=True)
            .exclude(lot_no="")
        )
        qs = filter_queryset_by_branch(
            qs, request.user, ItemLedgerEntries, request
        )

        if location_id:
            try:
                qs = qs.filter(location_id=int(location_id))
            except (ValueError, TypeError):
                pass

        from django.db.models import Count

        rows = (
            qs.values("lot_no", "expiry_date")
            .annotate(quantity=Sum("remaining_quantity"), entry_count=Count("id"))
            .order_by("expiry_date", "lot_no")
        )

        results = [
            {
                "lot_no": r["lot_no"],
                "expiry_date": r["expiry_date"].isoformat() if r["expiry_date"] else None,
                "quantity": r["quantity"],
                "entry_count": r["entry_count"],
            }
            for r in rows
        ]

        total = sum(r["quantity"] for r in results)

        return Response(
            {
                "item_id": item.no,
                "item_no": item.no,
                "item_name": getattr(item, "item_name", str(item)),
                "location_id": int(location_id) if location_id else None,
                "lots": results,
                "total_quantity": total,
            }
        )


class ItemListApiView(ListAPIView):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    # filter_backends = [DjangoFilterBackend]
    pagination_class = PageNumberPagination
    page_size = 10  # You can adjust this number

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get("q", None)

        if search_query:
            # Remove quotes and decode URL encoding
            search_query = search_query.replace('"', "").replace("%22", "").strip()
            queryset = queryset.filter(
                Q(item_name__icontains=search_query)
                | Q(bar_code_no__icontains=search_query)
            )

        queryset = filter_queryset_by_branch(
            queryset, self.request.user, Item, self.request
        )
        return queryset.order_by("item_name")

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context


class ItemsFilter(ListAPIView):

    permission_classes = [IsAuthenticated]
    serializer_class = ItemSerializer
    authentication_classes = [SessionAuthentication]

    def get_queryset(self):
        queryset = Item.objects.all()
        search_query = self.request.query_params.get("q", None)

        if search_query:
            search_query = search_query.strip().upper()
            queryset = queryset.filter(
                Q(item_name__icontains=search_query)
                | Q(bar_code_no__icontains=search_query)
                | Q(system_id__icontains=search_query)
                | Q(no__icontains=search_query)
            )

        queryset = filter_queryset_by_branch(
            queryset, self.request.user, Item, self.request
        )
        return queryset.order_by("item_name")


class ItemListView(ListView):
    model = Item
    context_object_name = "items"
    template_name = "items/item-list.html"


class ItemCreateView(CreateView):
    template_name = "items/new-item.html"
    model = Item
    form_class = ItemForm
    context_object_name = "items"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # context["unit_of_measure_url"] = reverse("items:unit-of-measure")
        return context

    def post(self, request, *args, **kwargs):
        print("POST data:", request.POST)
        print("FILES:", request.FILES)  # Debug print

        # Create a mutable copy of POST data
        mutable_post = request.POST.copy()
        mutable_post["unit_price"] = mutable_post["unit_price"].replace(",", "")
        mutable_post["item_name"] = mutable_post["item_name"].upper()

        # Create form with both POST and FILES data
        form = self.form_class(mutable_post, request.FILES)

        if form.is_valid():
            item = form.save()

            # Get all uploaded files

            files = request.FILES.getlist("image")
            for file in files:
                ItemImages.objects.create(
                    item=item,
                    url=file,
                    alt_text=item.item_name,
                )
            messages.success(request, "Item created successfully")
            return redirect("items:items-list")
        else:
            print("Form errors:", form.errors)
            messages.error(request, "Error: Invalid form data")

        return self.render_to_response({"form": form})

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        unit_of_measure_form = UnitOfMeasureForm()
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "unit_of_measure_form": unit_of_measure_form,
                "unit_of_measure_url": reverse("items:unit-of-measure"),
                "htmx_vals_unit_of_measure": '{"action": "check_code", "type": "unit"}',
            },
        )


class ItemUpdateView(UpdateView):
    template_name = "items/item-edit.html"
    model = Item
    form_class = ItemForm
    context_object_name = "item"
    success_url = reverse_lazy("items:items-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # add unit price comma
        context["unit_price"] = f"{context['item'].unit_price:,}"
        print(context["unit_price"])
        return context

    def post(self, request, *args, **kwargs):
        """Handle POST requests: instantiate form instance with request.POST and request.FILES"""
        self.object = self.get_object()

        # Create a mutable copy of POST data
        mutable_post = request.POST.copy()

        # Clean the unit price value
        if "unit_price" in mutable_post:
            mutable_post["unit_price"] = mutable_post["unit_price"].replace(",", "")

        # Initialize the form with current instance and cleaned data
        form = self.form_class(
            mutable_post,
            request.FILES,
            instance=self.object,  # This is crucial for updating existing instance
        )

        if form.is_valid():
            try:
                # Save the form
                form.save()
                messages.success(request, "Item updated successfully")
                return redirect(self.success_url)
            except Exception as e:
                messages.error(request, f"Error updating item: {str(e)}")
                return self.form_invalid(form)
        else:
            messages.error(request, "Please correct the errors below.")
            return self.form_invalid(form)

    def form_invalid(self, form):
        """If the form is invalid, render the invalid form."""
        return self.render_to_response(self.get_context_data(form=form))


class ItemDeleteView(View):
    def post(self, request, pk):
        item = Item.objects.get(id=pk)
        item.delete()
        messages.success(request, "Item deleted successfully")
        return redirect("items:items-list")


class ItemDeleteImageView(View):
    def post(self, request, pk):
        image = ItemImages.objects.get(id=pk)
        image.delete()
        messages.success(request, "Image deleted successfully")
        return redirect("items:edit-item", pk=image.item.id)


class ItemLedgerEntriesView(ListAPIView):
    serializer_class = ItemLedgerEntriesSerializer

    def get_queryset(self):
        item_id = self.kwargs["pk"]
        return ItemLedgerEntries.objects.filter(item_id=item_id).order_by(
            "-posting_date"
        )

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Get summary data
        summary = queryset.aggregate(
            total_quantity=Sum("quantity"), total_remaining=Sum("remaining_quantity")
        )

        # Serialize the ledger entries
        serializer = self.get_serializer(queryset, many=True)

        return Response(
            {
                "ledger_entries": serializer.data,
                "summary": {
                    "total_quantity": summary["total_quantity"] or 0,
                    "total_remaining": summary["total_remaining"] or 0,
                },
            }
        )


class ItemJournalView(View):
    template_name = "items/item-journal.html"
    form_class = ItemJournalForm

    def get(self, request):
        form = self.form_class()
        item_journals = ItemJournal.objects.all()
        return render(
            request, self.template_name, {"item_journals": item_journals, "form": form}
        )

    def post(self, request):
        journal_id = request.POST.get("journal_id_delete", "").split(",")
        journal_ids = [int(id) for id in journal_id if id.strip()]
        if journal_ids:
            journal = ItemJournal.objects.filter(id__in=journal_ids)
            journal.delete()
            messages.success(request, "Journal deleted successfully")
        elif all(
            field in request.POST
            for field in [
                "item_search",
                "entry_type",
                "quantity",
                "unit_cost",
                "date",
                "description",
            ]
        ):
            try:
                item = Item.objects.get(item_name=request.POST["item_search"])
                journal = ItemJournal.objects.get(item=item)
                journal.entry_type = request.POST["entry_type"]
                journal.quantity = int(request.POST["quantity"])
                journal.unit_cost = float(str(request.POST["unit_cost"]).replace(",", ""))
                journal.date = request.POST["date"]
                journal.description = request.POST["description"]
                journal.save()
                messages.success(request, "Journal updated successfully")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        return redirect("items:item-journal")


class PostItemJournalView(APIView):
    permission_classes = [IsAuthenticated]
    # Frontend posts with Bearer JWT; SessionAuthentication would enforce CSRF and can 403
    # when a session cookie is present but no CSRF token is provided.
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return Response(
                {"error": "No journal entry IDs provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        invalid_journals = []
        for journal_id in ids:
            journalentry = ItemJournal.objects.get(pk=journal_id)
            receipt_no = f"RCP-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

            # Create a custom message storage to capture validation errors
            from django.contrib.messages.storage.base import BaseStorage
            from django.contrib.messages import constants

            class ValidationMessageStorage(BaseStorage):
                def __init__(self, request):
                    super().__init__(request)
                    self.validation_errors = []

                def add(self, level, message, extra_tags=""):
                    if level == constants.ERROR:
                        self.validation_errors.append(message)

                def _get(self, *args, **kwargs):
                    # BaseStorage unpacks as messages, all_retrieved = self._get()
                    return [], True

                def _store(self, messages, response, *args, **kwargs):
                    return []

            # Store original messages and set our custom storage.
            # Do not use `if original_messages:` — MessageStorage.__bool__/__len__
            # loads session data and can raise ValueError for API/JWT requests
            # when the session has no message payload (expected 2 values from _get()).
            original_messages = getattr(request, "_messages", None)
            validation_storage = ValidationMessageStorage(request)
            request._messages = validation_storage
            try:
                previewer = ItemJournalPreviewProcessor(
                    journalentry, request, receipt_no=receipt_no
                )
                preview_data = previewer.process()
            finally:
                if original_messages is not None:
                    request._messages = original_messages
                elif hasattr(request, "_messages"):
                    delattr(request, "_messages")

            # Get validation errors from our custom storage
            validation_errors = validation_storage.validation_errors

            if not preview_data or (
                isinstance(preview_data, dict) and not any(preview_data.values())
            ):
                invalid_journals.append(
                    {
                        "document_no": journalentry.document_no,
                        "errors": validation_errors,
                    }
                )
                continue
            poster = ItemJournalFinalPoster(preview_data, journalentry, request.user)
            try:
                poster.post_to_tables()
            except Exception as e:
                # Convert posting-time failures (e.g. insufficient inventory) into a 400 payload
                # so the frontend can display the reason without a 500 crash.
                invalid_journals.append(
                    {
                        "document_no": journalentry.document_no,
                        "errors": [str(e)],
                    }
                )
                continue

        if invalid_journals:
            return Response(
                {
                    "error": "Some journals failed validation and were not posted.",
                    "invalid_journals": invalid_journals,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {"message": "Item Journals posted successfully"},
            status=status.HTTP_200_OK,
        )


class PostItemJournalAsyncView(APIView):
    """
    Queue posting of item journals to Celery (recommended for large batches).
    """

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return Response(
                {"error": "No journal entry IDs provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Tenant schema for celery
        schema_name = getattr(getattr(request, "tenant", None), "schema_name", None)
        if not schema_name:
            # Fallback to current connection tenant
            try:
                from django.db import connection

                schema_name = getattr(getattr(connection, "tenant", None), "schema_name", None)
            except Exception:
                schema_name = None

        # Optional branch context (keeps preview/post consistent with UI)
        branch_id = request.headers.get("X-Branch-Id") or request.META.get("HTTP_X_BRANCH_ID")

        from items.tasks import post_item_journals_async

        task = post_item_journals_async.delay(
            journal_ids=[int(x) for x in ids],
            schema_name=schema_name,
            user_id=request.user.id,
            branch_id=int(branch_id) if branch_id else None,
        )
        return Response(
            {
                "task_id": task.id,
                "message": "Posting started",
                "status": "processing",
            },
            status=status.HTTP_202_ACCEPTED,
        )

class PostJournalView(View):
    pass


class UnitOfMeasureView(View):
    template_name = ""
    form_class = UnitOfMeasureForm

    def post(self, request):
        print("POST data:", request.POST)
        if request.method == "POST":
            print("POST data:", request.POST)
            code = request.POST.get("code") or ""
            if code:
                if UnitOfMeasure.objects.filter(code=code).exists():
                    return JsonResponse(
                        {"success": False, "message": "Unit of measure already exists"},
                        status=400,
                    )
                else:
                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Unit of measure added successfully",
                        },
                        status=200,
                    )
                    # form = self.form_class(request.POST)
                    # if form.is_valid():
                    #     form.save()
                # messages.success(request, "Unit of measure added successfully")
                # return redirect("items:item-journal")
            else:
                messages.error(request, "Error: Invalid form data")
        return redirect("items:item-journal")


class ExcelUpload(View):
    template_name = "items/item-upload.html"

    def get(self, request):
        with schema_context("public"):
            item_upload_templates = UploadTemplates.objects.filter(
                name=UploadTemplateChoices.ITEMS.value
            ).first()
            return render(
                request,
                self.template_name,
                {"item_upload_templates": item_upload_templates},
            )

    def post(self, request, *args, **kwargs):
        try:
            uploaded_file = request.FILES["file"]

            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
            os.makedirs(temp_dir, exist_ok=True)

            # Generate unique filename
            import_id = uuid.uuid4()
            filename = f"import_{import_id}.xlsx"
            file_path = os.path.join(temp_dir, filename)

            # Save file with proper permissions
            with open(file_path, "wb+") as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            # Make sure file is readable by celery worker
            os.chmod(file_path, 0o644)

            # Start processing task
            from items.tasks import process_item_import

            task = process_item_import.delay(file_path)

            return JsonResponse({"task_id": task.id, "message": "File upload started"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


# --------------------------------------- HTMX --------------------------------------- #
def search_items(request):
    pass


def unit_of_measure(request):
    print("POST data:", request.POST)
    try:
        if request.method == "POST":
            code = request.POST.get("international_stnd_code") or ""
            description = request.POST.get("description") or ""
            action = request.POST.get("action") or ""
            if code or description:
                if action == "check_code":
                    if UnitOfMeasure.objects.filter(
                        international_stnd_code=code.upper()
                    ).exists():
                        return HttpResponse(
                            "<span class='text-red-500'>Unit of measure already exists</span>"
                        )
                    elif description:
                        if UnitOfMeasure.objects.filter(
                            description=description.upper()
                        ).exists():
                            return HttpResponse(
                                "<span class='text-red-500'>Unit of measure already exists</span>"
                            )
                    else:
                        return HttpResponse("<span class='text-green-500'></span>")

                elif action == "save":
                    # remove the action from the request.POST
                    formData = request.POST.copy()
                    del formData["action"]
                    formData["international_stnd_code"] = formData[
                        "international_stnd_code"
                    ].upper()
                    formData["description"] = formData["description"].upper()
                    form = UnitOfMeasureForm(formData)
                    if form.is_valid():
                        form.save()
                        messages.success(request, "Unit of measure added successfully")
                        return JsonResponse(
                            {
                                "success": True,
                                "message": "Unit of measure added successfully",
                            }
                        )
                    else:
                        print(form.errors)
                        return JsonResponse(
                            {"success": False, "message": "Error: Invalid form data"},
                            status=400,
                        )
        else:
            messages.error(request, "Error: Invalid form data")
            return JsonResponse(
                {"success": False, "message": "Error: Invalid form data"}, status=400
            )
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


def refresh_unit_options(request):
    form = ItemForm()
    unit_field = form["unit_of_measure"]

    # as_widget() is used to render the form field as HTML
    return HttpResponse(unit_field.as_widget())


# ---------------- Admin endpoints ---------------- #
@staff_member_required
def get_item_cost(request, item_id):
    try:
        item = Item.objects.get(id=item_id)
        return JsonResponse({"unit_cost": 1000})
    except Item.DoesNotExist:
        return JsonResponse({"error": "Item not found"}, status=404)


class ItemCategoryViewSet(viewsets.ModelViewSet):
    queryset = ItemCategory.objects.all()
    serializer_class = ItemCategorySerializer
    permission_classes = [IsAuthenticated]
    # pagination_class = StandardResultsSetPagination
    pagination_class = None
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    lookup_field = "code"

    def list(self, request, *args, **kwargs):
        """List all item categories - requires read permission"""
        has_permission, source = request.user.check_object_permission(10204, "read")
        if not has_permission:
            return Response(
                {"error": "You do not have permission to view item categories"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        """Create new item category - requires insert permission"""
        has_permission, source = request.user.check_object_permission(10204, "insert")
        if not has_permission:
            return Response(
                {"error": "You do not have permission to create item categories"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """Update item category - requires modify permission"""
        has_permission, source = request.user.check_object_permission(10204, "modify")
        if not has_permission:
            return Response(
                {"error": "You do not have permission to update item categories"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete item category - requires delete permission"""
        has_permission, source = request.user.check_object_permission(10204, "delete")
        if not has_permission:
            return Response(
                {"error": "You do not have permission to delete item categories"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        queryset = ItemCategory.objects.all()

        # Filter by parent (optional)
        parent_id = self.request.query_params.get("parent_id", None)
        if parent_id is not None:
            if parent_id == "0":  # Root categories
                queryset = queryset.filter(parent__isnull=True)
            else:
                queryset = queryset.filter(parent_id=parent_id)

        # Search by code or description
        search = self.request.query_params.get("search", None)
        if search is not None:
            queryset = queryset.filter(
                Q(code__icontains=search) | Q(description__icontains=search)
            )

        return queryset.order_by("tree_id", "lft")

    @action(detail=True, methods=["get"], url_path="attributes")
    def get_attributes(self, request, code=None):
        """Get all attributes assigned to this category"""
        category = self.get_object()
        attributes = category.attributes.filter(blocked=False).order_by("name")
        serializer = ItemAttributeSerializer(attributes, many=True)
        return Response({"attributes": serializer.data})


class ItemUnitOfMeasureViewSet(viewsets.ModelViewSet):
    queryset = ItemUnitOfMeasure.objects.all()
    serializer_class = ItemUnitOfMeasureSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None
    filterset_fields = ["item"]

    def destroy(self, request, *args, **kwargs):
        """Override destroy to handle validation errors gracefully"""
        instance = self.get_object()
        try:
            return super().destroy(request, *args, **kwargs)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # cause can send  iuoym as id or system_id


class UnitOfMeasureViewSet(viewsets.ModelViewSet):
    queryset = UnitOfMeasure.objects.all()
    serializer_class = UnitOfMeasureSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]

    # pagination_class = StandardResultsSetPagination
    pagination_class = None

    def get_queryset(self):
        """
        Optionally restricts the returned units of measure to a given user,
        by filtering against a `q` query parameter in the URL.
        """
        queryset = UnitOfMeasure.objects.all()
        q = self.request.query_params.get("q", None)
        if q is not None:
            queryset = queryset.filter(
                Q(code__icontains=q) | Q(description__icontains=q)
            )
        return queryset.order_by("code")


class ItemTrackingCodeViewSet(viewsets.ModelViewSet):
    queryset = ItemTrackingCodes.objects.all()
    serializer_class = ItemTrackingCodeSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    lookup_field = "system_id"
    pagination_class = None

    def _require_tracking_module(self):
        if not getattr(self.request, "has_module", lambda m: True)("item_tracking"):
            return Response(
                {"error": "Item tracking module required to manage tracking codes"},
                status=status.HTTP_403_FORBIDDEN,
            )
        return None

    def create(self, request, *args, **kwargs):
        denied = self._require_tracking_module()
        if denied:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._require_tracking_module()
        if denied:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._require_tracking_module()
        if denied:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._require_tracking_module()
        if denied:
            return denied
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        queryset = ItemTrackingCodes.objects.all()
        q = self.request.query_params.get("q", None)
        if q is not None:
            queryset = queryset.filter(
                Q(code__icontains=q) | Q(description__icontains=q)
            )
        return queryset.order_by("code")


class ItemJournalViewSet(viewsets.ModelViewSet):
    queryset = ItemJournal.objects.all()
    serializer_class = ItemJournalSerializer
    permission_classes = [IsAuthenticated]
    # JWT first for SPA/API calls; avoids SessionAuthentication CSRF 403 on POST
    # when a browser session cookie is present.
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = [
        "item",
        "entry_type",
        "document_no",
        "date",
        "status",
        "production_order",
        "global_dimension_1",
        "global_dimension_2",
        "adjustment_type",
    ]
    ordering_fields = [
        "document_no",
        "item__item_name",
        "entry_type",
        "quantity",
        "unit_amount",
        "amount",
        "date",
        "status",
        "created_at",
    ]
    ordering = ["-created_at"]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by journal template type if specified
        template_type = self.request.query_params.get("journal_template__type", None)
        if template_type:
            if template_type == "item":
                # Include ITEM journals and legacy journals with null template
                queryset = queryset.filter(
                    Q(journal_template__type="item")
                    | Q(journal_template__isnull=True)
                )
            else:
                queryset = queryset.filter(journal_template__type=template_type)

        # Handle search parameter
        search_query = self.request.query_params.get("q", None)
        if search_query:
            search_query = search_query.strip()
            queryset = queryset.filter(
                Q(document_no__icontains=search_query)
                | Q(description__icontains=search_query)
                | Q(item__item_name__icontains=search_query)
            )

        # Handle ordering - OrderingFilter will handle it, but we provide default
        ordering = self.request.query_params.get("ordering", None)
        if not ordering:
            queryset = queryset.order_by("-created_at")

        # Branch filter (same headers/user rules as other transaction viewsets).
        queryset = filter_queryset_by_branch(
            queryset, self.request.user, ItemJournal, request=self.request
        )

        return queryset

    @action(detail=False, methods=["post"])
    def upsert(self, request):
        system_id = request.data.get("system_id")
        id = request.data.get("id")
        document_no = request.data.get("document_no")
        deleted = request.data.get("deleted", False)

        try:
            with transaction.atomic():
                if system_id or id or document_no:
                    # Handle deletion
                    if deleted:
                        item_journal = (
                            ItemJournal.objects.filter(system_id=system_id).first()
                            or ItemJournal.objects.filter(id=id).first()
                        )
                        if not item_journal:
                            raise ItemJournal.DoesNotExist
                        item_journal.delete()
                        return Response(
                            {"message": "ItemJournal deleted successfully"},
                            status=status.HTTP_200_OK,
                        )

                    # Update existing item journal
                    item_journal = (
                        ItemJournal.objects.filter(system_id=system_id).first()
                        or ItemJournal.objects.filter(id=id).first()
                        or ItemJournal.objects.filter(document_no=document_no).first()
                    )
                    if not item_journal:
                        raise ItemJournal.DoesNotExist

                    serializer = self.get_serializer(
                        item_journal, data=request.data, partial=True
                    )
                else:
                    # Create new item journal
                    serializer = self.get_serializer(data=request.data)

                if serializer.is_valid(raise_exception=True):
                    item_journal = serializer.save()
                    return Response(
                        serializer.data,
                        status=(
                            status.HTTP_200_OK
                            if (system_id or id)
                            else status.HTTP_201_CREATED
                        ),
                    )

        except ValidationError as e:
            # Return the error as a string or dict, not e.detail
            if hasattr(e, "message_dict"):
                return Response(
                    {"detail": e.message_dict}, status=status.HTTP_400_BAD_REQUEST
                )
            else:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except ItemJournal.DoesNotExist:
            return Response(
                {"detail": "ItemJournal not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"Error in upsert: {str(e)}")
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        """
        Bulk delete item journals by ids.
        Expected payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            return Response(
                {"detail": "ids must be a non-empty list"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Use get_queryset() so branch-level filtering/rules are respected.
            journals = self.get_queryset().filter(id__in=ids)
            found_ids = set(journals.values_list("id", flat=True))
            missing_ids = [journal_id for journal_id in ids if journal_id not in found_ids]

            with transaction.atomic():
                deleted_count, _ = journals.delete()

            return Response(
                {
                    "success": True,
                    "deleted_count": deleted_count,
                    "missing_ids": missing_ids,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def export_adjustments(self, request):
        """
        Export open inventory adjustment journals (ITEM template) to Excel.
        Respects existing queryset filters (search/order/branch visibility).
        """
        try:
            import xlsxwriter

            # Start from permission-scoped queryset and force Adjust Inventory scope.
            queryset = (
                self.get_queryset()
                .filter(status="Open")
                .filter(
                    Q(journal_template__type="item")
                    | Q(journal_template__isnull=True)
                )
                .select_related("user", "item")
            )

            adjustment_type = (request.query_params.get("adjustment_type") or "").strip()
            if adjustment_type in ("operational", "opening_balance"):
                queryset = queryset.filter(adjustment_type=adjustment_type)

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {"strings_to_urls": False})
            worksheet = workbook.add_worksheet("Inventory Adjustments")

            headers = [
                "Document No",
                "Item No",
                "Item Name",
                "Entry Type",
                "Adjustment Category",
                "Quantity",
                "Unit Amount",
                "Amount",
                "Date",
                "Description",
                "Status",
                "Adjusted By",
                "Created At",
            ]

            header_format = workbook.add_format(
                {"bold": True, "bg_color": "#F3F4F6", "border": 1}
            )
            text_format = workbook.add_format({"border": 1})
            number_format = workbook.add_format({"border": 1, "num_format": "#,##0.00"})
            date_format = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd"})
            datetime_format = workbook.add_format(
                {"border": 1, "num_format": "yyyy-mm-dd hh:mm:ss"}
            )

            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)

            worksheet.set_column(0, 0, 16)   # Document No
            worksheet.set_column(1, 1, 14)   # Item No
            worksheet.set_column(2, 2, 28)   # Item Name
            worksheet.set_column(3, 4, 20)   # Entry/Adjustment
            worksheet.set_column(5, 7, 14)   # Qty/amounts
            worksheet.set_column(8, 8, 14)   # Date
            worksheet.set_column(9, 9, 30)   # Description
            worksheet.set_column(10, 10, 12) # Status
            worksheet.set_column(11, 11, 20) # Adjusted By
            worksheet.set_column(12, 12, 22) # Created At

            row_idx = 1
            for journal in queryset.iterator():
                user_name = ""
                if getattr(journal, "user_id", None) and journal.user:
                    u = journal.user
                    user_name = (getattr(u, "full_name", "") or "").strip()
                    if not user_name:
                        user_name = getattr(u, "username", "") or getattr(
                            u, "email", ""
                        ) or ""

                worksheet.write(row_idx, 0, journal.document_no or "", text_format)
                worksheet.write(row_idx, 1, getattr(journal.item, "no", "") or "", text_format)
                worksheet.write(
                    row_idx, 2, getattr(journal.item, "item_name", "") or "", text_format
                )
                worksheet.write(row_idx, 3, journal.get_entry_type_display(), text_format)
                worksheet.write(
                    row_idx,
                    4,
                    (journal.adjustment_type or "").replace("_", " ").title(),
                    text_format,
                )
                worksheet.write_number(row_idx, 5, float(journal.quantity or 0), number_format)
                worksheet.write_number(
                    row_idx, 6, float(journal.unit_amount or 0), number_format
                )
                worksheet.write_number(row_idx, 7, float(journal.amount or 0), number_format)

                if journal.date:
                    worksheet.write_datetime(
                        row_idx,
                        8,
                        datetime.combine(journal.date, datetime.min.time()),
                        date_format,
                    )
                else:
                    worksheet.write(row_idx, 8, "", text_format)

                worksheet.write(row_idx, 9, journal.description or "", text_format)
                worksheet.write(row_idx, 10, journal.status or "", text_format)
                worksheet.write(row_idx, 11, user_name, text_format)

                if journal.created_at:
                    worksheet.write_datetime(
                        row_idx,
                        12,
                        journal.created_at.replace(tzinfo=None),
                        datetime_format,
                    )
                else:
                    worksheet.write(row_idx, 12, "", text_format)

                row_idx += 1

            workbook.close()
            output.seek(0)

            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            export_label = (
                "opening_balance"
                if adjustment_type == "opening_balance"
                else "inventory_adjustments"
            )
            filename = f"{export_label}_export_{timestamp}.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response(
                {"detail": f"Failed to export inventory adjustments: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"])
    def calculate_inventory(self, request):
        """
        Calculate inventory and populate journal lines (background task).
        Request body:
        {
            "posting_date": "2026-01-05",
            "location_filter": 1,  # optional location ID
            "item_filter": "1000",  # optional item no
            "journal_template": "PHYS. INV.",  # template name
            "journal_batch": "DEFAULT"  # batch name
        }
        Returns task_id immediately for status checking.
        """
        from items.tasks import calculate_inventory_task
        from django.db import connection

        posting_date = request.data.get("posting_date")
        location_filter = request.data.get("location_filter")
        item_filter = request.data.get("item_filter")
        include_zero_quantity = str(
            request.data.get("include_zero_quantity", "false")
        ).lower() in ("1", "true", "yes", "y", "on")
        template_name = request.data.get("journal_template", "PHYS. INV.")
        batch_name = request.data.get("journal_batch", "DEFAULT")
        branch = get_branch_for_request(request)

        if not posting_date:
            return Response(
                {"detail": "posting_date is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get schema name for tenant context
            schema_name = connection.schema_name

            # Trigger background task
            task = calculate_inventory_task.delay(
                posting_date=posting_date,
                location_filter=location_filter,
                item_filter=item_filter,
                journal_template=template_name,
                journal_batch=batch_name,
                include_zero_quantity=include_zero_quantity,
                user_id=request.user.id,
                schema_name=schema_name,
                branch_id=getattr(branch, "id", None),
            )

            return Response(
                {
                    "task_id": task.id,
                    "message": "Inventory calculation started. Use the task_id to check status.",
                    "status": "pending",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            print(f"Error starting calculate_inventory task: {str(e)}")
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def calculate_inventory_status(self, request):
        """
        Check the status of a calculate inventory task.
        Query params:
            task_id: The task ID returned from calculate_inventory
        """
        task_id = request.query_params.get("task_id")

        if not task_id:
            return Response(
                {"detail": "task_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            task_result = AsyncResult(task_id)

            if task_result.state == PENDING:
                response = {
                    "task_id": task_id,
                    "status": "pending",
                    "message": "Task is waiting to be processed",
                }
            elif task_result.state == "PROGRESS":
                response = {
                    "task_id": task_id,
                    "status": "in_progress",
                    "message": task_result.info.get("status", "Processing..."),
                    "progress": {
                        "current": task_result.info.get("current", 0),
                        "total": task_result.info.get("total", 0),
                    },
                }
            elif task_result.state == SUCCESS:
                result = task_result.result
                response = {
                    "task_id": task_id,
                    "status": "success",
                    "message": result.get("message", "Inventory calculation completed"),
                    "deleted_count": result.get("deleted_count", 0),
                    "created_count": result.get("created_count", 0),
                    "journal_ids": result.get("journal_ids", []),
                }
            elif task_result.state == FAILURE:
                response = {
                    "task_id": task_id,
                    "status": "failure",
                    "message": str(task_result.info),
                    "error": str(task_result.info),
                }
            else:
                response = {
                    "task_id": task_id,
                    "status": task_result.state.lower(),
                    "message": f"Task state: {task_result.state}",
                }

            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"detail": f"Error checking task status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"])
    def export_stock_taking(self, request):
        """
        Export stock taking journals to Excel (background task).
        Request body:
        {
            "show_calculated_qty": true,
            "show_tracking_numbers": true,
            "show_location_code": true,
            "location_filter": 1  # optional location ID
        }
        Returns task_id immediately for status checking.
        """
        from items.tasks import export_stock_taking_task
        from django.db import connection

        show_calculated_qty = request.data.get("show_calculated_qty", False)
        show_tracking_numbers = request.data.get("show_tracking_numbers", False)
        show_location_code = request.data.get("show_location_code", False)
        include_unit_cost = str(
            request.data.get("include_unit_cost", "false")
        ).lower() in ("1", "true", "yes", "y", "on")
        include_zero_quantity = str(
            request.data.get("include_zero_quantity", "true")
        ).lower() in ("1", "true", "yes", "y", "on")
        location_filter = request.data.get("location_filter", None)

        try:
            # Get schema name for tenant context
            schema_name = connection.schema_name

            # Trigger background task
            task = export_stock_taking_task.delay(
                show_calculated_qty=show_calculated_qty,
                show_tracking_numbers=show_tracking_numbers,
                show_location_code=show_location_code,
                include_unit_cost=include_unit_cost,
                include_zero_quantity=include_zero_quantity,
                location_filter=location_filter,
                schema_name=schema_name,
                user_id=request.user.id,
            )

            return Response(
                {
                    "task_id": task.id,
                    "message": "Export started. Use the task_id to check status.",
                    "status": "pending",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def export_stock_taking_status(self, request):
        """
        Check the status of an export task.
        Query params:
            task_id: The task ID returned from export_stock_taking
        """
        task_id = request.query_params.get("task_id")

        if not task_id:
            return Response(
                {"detail": "task_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            task_result = AsyncResult(task_id)

            if task_result.state == PENDING:
                response = {
                    "task_id": task_id,
                    "status": "pending",
                    "message": "Task is waiting to be processed",
                    "progress": 0,
                }
            elif task_result.state == "PROGRESS":
                meta = task_result.info if task_result.info else {}
                response = {
                    "task_id": task_id,
                    "status": "in_progress",
                    "message": meta.get("message", "Processing..."),
                    "progress": meta.get("progress", 0),
                }
            elif task_result.state == SUCCESS:
                result = task_result.result
                response = {
                    "task_id": task_id,
                    "status": "success",
                    "message": result.get("message", "Export completed"),
                    "progress": 100,
                    "filename": result.get("filename"),
                    "cache_key": result.get("cache_key"),
                    "total_journals": result.get("total_journals", 0),
                }
            elif task_result.state == FAILURE:
                response = {
                    "task_id": task_id,
                    "status": "failure",
                    "message": str(task_result.info),
                    "error": str(task_result.info),
                    "progress": 0,
                }
            else:
                response = {
                    "task_id": task_id,
                    "status": task_result.state.lower(),
                    "message": f"Task state: {task_result.state}",
                    "progress": 0,
                }

            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"detail": f"Error checking task status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"])
    def export_stock_taking_download(self, request):
        """
        Download the exported Excel file.
        Query params:
            task_id: The task ID returned from export_stock_taking
        """
        from django.http import FileResponse
        from django.db import connection

        task_id = request.query_params.get("task_id")

        if not task_id:
            return Response(
                {"detail": "task_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            schema_name = connection.schema_name
            cache_key = f"export_stock_taking_{schema_name}_{task_id}"

            # Retrieve file from cache
            file_info = cache.get(cache_key)

            if not file_info:
                return Response(
                    {"detail": "Export file not found or expired"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Decode base64 file data
            file_data = base64.b64decode(file_info["file_data"])

            # Create BytesIO object
            file_buffer = BytesIO(file_data)
            file_buffer.seek(0)

            # Return file response
            response = FileResponse(
                file_buffer,
                content_type=file_info["content_type"],
                as_attachment=True,
                filename=file_info["filename"],
            )

            return response

        except Exception as e:
            return Response(
                {"detail": f"Error downloading file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"])
    def import_stock_taking(self, request):
        """
        Import physical quantities from Excel/CSV file (background task).
        Request: multipart/form-data with 'file' field
        Returns task_id immediately for status checking.
        """
        from items.tasks import import_stock_taking_task
        from django.db import connection
        from django.conf import settings
        import uuid

        if "file" not in request.FILES:
            return Response(
                {"detail": "No file uploaded"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            uploaded_file = request.FILES["file"]

            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
            os.makedirs(temp_dir, exist_ok=True)

            # Generate unique filename
            import_id = uuid.uuid4()
            file_ext = os.path.splitext(uploaded_file.name)[1]
            filename = f"stock_taking_import_{import_id}{file_ext}"
            file_path = os.path.join(temp_dir, filename)

            # Save file with proper permissions
            with open(file_path, "wb+") as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            # Make sure file is readable by celery worker
            os.chmod(file_path, 0o644)

            # Get schema name for tenant context
            schema_name = connection.schema_name

            # Trigger background task
            task = import_stock_taking_task.delay(
                file_path=file_path,
                schema_name=schema_name,
                user_id=request.user.id,
            )

            return Response(
                {
                    "task_id": task.id,
                    "message": "Import started. Use the task_id to check status.",
                    "status": "pending",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def import_stock_taking_status(self, request):
        """
        Check the status of an import task.
        Query params:
            task_id: The task ID returned from import_stock_taking
        """
        task_id = request.query_params.get("task_id")

        if not task_id:
            return Response(
                {"detail": "task_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            task_result = AsyncResult(task_id)

            if task_result.state == PENDING:
                response = {
                    "task_id": task_id,
                    "status": "pending",
                    "message": "Task is waiting to be processed",
                    "progress": 0,
                }
            elif task_result.state == "PROGRESS":
                meta = task_result.info if task_result.info else {}
                response = {
                    "task_id": task_id,
                    "status": "in_progress",
                    "message": meta.get("message", "Processing..."),
                    "progress": meta.get("progress", 0),
                }
            elif task_result.state == SUCCESS:
                result = task_result.result
                response = {
                    "task_id": task_id,
                    "status": "success",
                    "message": result.get("message", "Import completed"),
                    "progress": 100,
                    "updated_count": result.get("updated_count", 0),
                    "failed_count": result.get("failed_count", 0),
                    "total_rows": result.get("total_rows", 0),
                    "errors": result.get("errors", []),
                }
            elif task_result.state == FAILURE:
                response = {
                    "task_id": task_id,
                    "status": "failure",
                    "message": str(task_result.info),
                    "error": str(task_result.info),
                    "progress": 0,
                }
            else:
                response = {
                    "task_id": task_id,
                    "status": task_result.state.lower(),
                    "message": f"Task state: {task_result.state}",
                    "progress": 0,
                }

            return Response(response, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"detail": f"Error checking task status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # ---- Inventory Adjustment Import (simple Excel flow) ----

    @action(detail=False, methods=["get"])
    def import_adjustment_template(self, request):
        """
        Generate a dynamic Excel template for inventory adjustment imports.
        Populates dropdowns from the tenant's items, UOMs, and locations.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        adjustment_type = (request.query_params.get("adjustment_type") or "").strip()
        if adjustment_type not in ("operational", "opening_balance"):
            adjustment_type = "operational"
        default_category_label = (
            "Opening Balance"
            if adjustment_type == "opening_balance"
            else "Operational Adjustment"
        )
        template_filename = (
            "opening_balance_adjustment_template.xlsx"
            if adjustment_type == "opening_balance"
            else "inventory_adjustment_template.xlsx"
        )

        wb = Workbook()

        # --- Reference sheet (hidden, powers dropdowns) ---
        ref_ws = wb.create_sheet("Reference")

        item_names = list(
            Item.objects.values_list("item_name", flat=True).order_by("item_name")
        )
        uom_codes = list(
            UnitOfMeasure.objects.values_list("code", flat=True).order_by("code")
        )
        location_names = list(
            Location.objects.values_list("description", flat=True).order_by("description")
        )
        entry_types = ["Increase Inventory", "Decrease Inventory"]
        adjustment_categories = ["Operational Adjustment", "Opening Balance"]

        ref_ws["A1"] = "Item"
        for i, name in enumerate(item_names, start=2):
            ref_ws[f"A{i}"] = name

        ref_ws["B1"] = "Entry Type"
        for i, val in enumerate(entry_types, start=2):
            ref_ws[f"B{i}"] = val

        ref_ws["C1"] = "Adjustment Category"
        for i, val in enumerate(adjustment_categories, start=2):
            ref_ws[f"C{i}"] = val

        ref_ws["D1"] = "Unit of Measure"
        for i, code in enumerate(uom_codes, start=2):
            ref_ws[f"D{i}"] = code

        ref_ws["E1"] = "Location"
        for i, name in enumerate(location_names, start=2):
            ref_ws[f"E{i}"] = name

        ref_ws.sheet_state = "hidden"

        # --- Adjustments sheet (main data entry) ---
        ws = wb.active
        ws.title = "Adjustments"

        columns = [
            {"header": "Document No", "width": 16, "required": False},
            {"header": "Item", "width": 35, "required": True},
            # Increase/Decrease (maps to ItemJournal.entry_type)
            {"header": "Entry Type", "width": 22, "required": True},
            # Operational vs Opening Balance (maps to ItemJournal.adjustment_type)
            {"header": "Adjustment Category", "width": 22, "required": False},
            {"header": "Quantity", "width": 12, "required": True},
            {"header": "Unit Cost", "width": 14, "required": False},
            {"header": "Unit of Measure", "width": 20, "required": False},
            {"header": "Location", "width": 20, "required": False},
            {"header": "Lot No", "width": 18, "required": False},
            {"header": "Expiry Date", "width": 14, "required": False},
            {"header": "Date", "width": 15, "required": False},
            {"header": "Description", "width": 40, "required": False},
        ]

        header_fill_req = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        header_fill_opt = PatternFill(
            start_color="E2E2E2", end_color="E2E2E2", fill_type="solid"
        )
        header_font = Font(bold=True, size=11)
        thin_border = Border(bottom=Side(style="thin", color="999999"))
        example_font = Font(italic=True, color="999999")

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = header_font
            cell.fill = header_fill_req if col_def["required"] else header_fill_opt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # Example rows
        from datetime import date
        examples = [
            [
                "",  # Document No — fill when updating an exported line
                item_names[0] if item_names else "Sugar 1kg",
                "Increase Inventory",
                default_category_label,
                10,
                "",  # Unit Cost - optional, leave blank to use item default
                uom_codes[0] if uom_codes else "",
                location_names[0] if location_names else "",
                "",  # Lot No - optional (required for lot-tracked items)
                "",  # Expiry Date - optional (YYYY-MM-DD or Excel date)
                date.today().isoformat(),  # Current date as sample
                "Opening stock" if adjustment_type == "opening_balance" else "Stock adjustment",
            ],
        ]
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, val in enumerate(example, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = example_font

        max_rows = 5000

        if item_names:
            dv_item = DataValidation(
                type="list",
                formula1=f"Reference!$A$2:$A${len(item_names) + 1}",
                allow_blank=True,
            )
            dv_item.error = "Please select a valid Item"
            dv_item.errorTitle = "Invalid Item"
            ws.add_data_validation(dv_item)
            dv_item.add(f"B4:B{max_rows}")

        dv_entry = DataValidation(
            type="list",
            formula1=f"Reference!$B$2:$B${len(entry_types) + 1}",
            allow_blank=True,
        )
        dv_entry.error = "Please select a valid Entry Type"
        dv_entry.errorTitle = "Invalid Type"
        ws.add_data_validation(dv_entry)
        dv_entry.add(f"C4:C{max_rows}")

        dv_adj_category = DataValidation(
            type="list",
            formula1=f"Reference!$C$2:$C${len(adjustment_categories) + 1}",
            allow_blank=True,
        )
        dv_adj_category.error = "Please select a valid Adjustment Category"
        dv_adj_category.errorTitle = "Invalid Category"
        ws.add_data_validation(dv_adj_category)
        dv_adj_category.add(f"D4:D{max_rows}")

        if uom_codes:
            dv_uom = DataValidation(
                type="list",
                formula1=f"Reference!$D$2:$D${len(uom_codes) + 1}",
                allow_blank=True,
            )
            dv_uom.error = "Please select a valid Unit of Measure"
            dv_uom.errorTitle = "Invalid UOM"
            ws.add_data_validation(dv_uom)
            dv_uom.add(f"G4:G{max_rows}")

        if location_names:
            dv_loc = DataValidation(
                type="list",
                formula1=f"Reference!$E$2:$E${len(location_names) + 1}",
                allow_blank=True,
            )
            dv_loc.error = "Please select a valid Location"
            dv_loc.errorTitle = "Invalid Location"
            ws.add_data_validation(dv_loc)
            dv_loc.add(f"H4:H{max_rows}")

        # --- Instructions sheet ---
        instr_ws = wb.create_sheet("Instructions")
        instructions = [
            ["ZentroApp - Inventory Adjustment Import Template"],
            [""],
            ["How to use this template:"],
            ["1. Fill in your adjustments starting from row 4 on the 'Adjustments' sheet"],
            ["2. Green columns are REQUIRED - you must fill these in"],
            ["3. Grey columns are OPTIONAL - leave blank if not needed"],
            ["4. Use the dropdown lists where available"],
            ["5. You may delete the example row (row 2) before importing"],
            ["6. Save the file and upload it on the Adjust Inventory page"],
            [
                "7. To update existing open lines, export from the page, keep Document No, edit values, then re-import",
            ],
            [""],
            ["Column Descriptions:"],
            [
                "Document No",
                "Optional. Include when updating a line exported from ZentroApp. Leave blank for new lines.",
            ],
            ["Item", "The name of the item to adjust (must exist in the system)"],
            ["Entry Type", "'Increase Inventory' to add stock, 'Decrease Inventory' to remove stock"],
            ["Adjustment Category", "'Operational Adjustment' for normal inventory changes, 'Opening Balance' for initial stock setup"],
            ["Quantity", "The number of units to add or remove (always a positive number)"],
            ["Unit Cost", "Buying/unit cost per item. Leave blank to use the item's default cost."],
            ["Unit of Measure", "The unit of measure. Leave blank to use the item's default."],
            ["Location", "Warehouse/branch location. Leave blank if single-location."],
            [
                "Lot No",
                "Batch/lot number for lot-tracked items. Required when the item's tracking code requires a lot.",
            ],
            [
                "Expiry Date",
                "Expiry date for lot-tracked items (YYYY-MM-DD). Required when the item's tracking code requires expiry.",
            ],
            ["Date", "Date of the adjustment (YYYY-MM-DD). Defaults to today."],
            ["Description", "Optional reason or note for the adjustment"],
        ]
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                cell = instr_ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 3:
                    cell.font = Font(bold=True, size=12)
                elif row_idx == 11:
                    cell.font = Font(bold=True, size=12)
                elif row_idx >= 12 and col_idx == 1:
                    cell.font = Font(bold=True)
        instr_ws.column_dimensions["A"].width = 25
        instr_ws.column_dimensions["B"].width = 70

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{template_filename}"'
        )
        return response

    @action(
        detail=False,
        methods=["post"],
        parser_classes=[MultiPartParser, FormParser],
        authentication_classes=[JWTAuthentication],
        permission_classes=[IsAuthenticated],
    )
    def import_adjustments_preflight(self, request):
        """
        Preflight an adjustment import file and report missing items before importing.

        Accepts multipart/form-data with a 'file' field.
        Returns:
          - missing_items: list[str] (item names not found)
          - total_rows: int
          - has_tracking: bool (any Lot/Expiry values present)
        """
        if "file" not in request.FILES:
            return Response(
                {"detail": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Keep aliases in sync with items.tasks.process_journal_import
        COLUMN_ALIASES = {
            "Adjustment Type": "Entry Type",
            "Entry Type": "Entry Type",
            "Unit of Measure": "Unit Of Measure",
            "Location": "Location Code",
            "Item Name": "Item",
            "Lot No": "Lot No",
            "Lot No.": "Lot No",
            "Batch": "Lot No",
            "Batch / Lot": "Lot No",
            "Batch/Lot": "Lot No",
            "Expiry Date": "Expiry Date",
            "Expiration Date": "Expiry Date",
        }

        try:
            import pandas as pd

            uploaded_file = request.FILES["file"]
            name = (uploaded_file.name or "").lower()
            if name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                # Default to Excel for .xlsx/.xls and unknown extensions.
                df = pd.read_excel(uploaded_file, header=0)

            df.columns = df.columns.str.strip()
            df.rename(columns=COLUMN_ALIASES, inplace=True)
            df = df.fillna("")

            total_rows = int(len(df))

            # Extract distinct item names from the "Item" column (template uses Item Name).
            items_series = df.get("Item", None)
            if items_series is None and "Item Name" in df.columns:
                items_series = df["Item Name"]
            if items_series is None:
                return Response(
                    {"detail": "Required column 'Item' not found in file"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            raw_names = (
                items_series.astype(str).map(lambda v: v.strip()).tolist()
                if total_rows
                else []
            )
            unique_names = sorted({n for n in raw_names if n})

            existing_qs = Item.objects.filter(item_name__in=unique_names).values_list(
                "item_name", flat=True
            )
            existing_names = set(existing_qs)

            # Case-insensitive existence check for any that didn't match exactly by case.
            missing = []
            for n in unique_names:
                if n in existing_names:
                    continue
                if Item.objects.filter(item_name__iexact=n).exists():
                    continue
                missing.append(n)

            has_lot = False
            has_exp = False
            if "Lot No" in df.columns:
                has_lot = df["Lot No"].astype(str).map(lambda v: v.strip()).ne("").any()
            if "Expiry Date" in df.columns:
                has_exp = (
                    df["Expiry Date"].astype(str).map(lambda v: v.strip()).ne("").any()
                )

            return Response(
                {
                    "missing_items": missing,
                    "total_rows": total_rows,
                    "has_tracking": bool(has_lot or has_exp),
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=["post"],
        parser_classes=[MultiPartParser, FormParser],
        authentication_classes=[JWTAuthentication],
        permission_classes=[IsAuthenticated],
    )
    def import_adjustments(self, request):
        """
        Import inventory adjustments from an uploaded Excel file (background task).
        Accepts multipart/form-data with a 'file' field.
        Returns task_id for status polling.
        """
        from items.tasks import process_journal_import
        import uuid as _uuid

        if "file" not in request.FILES:
            return Response(
                {"detail": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            uploaded_file = request.FILES["file"]
            temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
            os.makedirs(temp_dir, exist_ok=True)

            import_id = _uuid.uuid4()
            file_ext = os.path.splitext(uploaded_file.name)[1]
            filename = f"adjustment_import_{import_id}{file_ext}"
            file_path = os.path.join(temp_dir, filename)

            with open(file_path, "wb+") as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            schema_name = connection.schema_name
            # Capture current branch context (X-Branch-Id / user.global_dimension_1) for the background task.
            branch = get_branch_for_request(request)

            create_missing_items = str(
                request.data.get("create_missing_items", "false")
            ).lower() in ("1", "true", "yes", "y", "on")
            default_tracking_code = (
                request.data.get("default_tracking_code") or "ALL LOT"
            )
            create_missing_lot_expiry_for_missing = str(
                request.data.get("create_missing_lot_expiry_for_missing", "false")
            ).lower() in ("1", "true", "yes", "y", "on")
            missing_lot_expiry_date = (
                str(request.data.get("missing_lot_expiry_date", "")).strip() or None
            )

            def _normalize_debug_key(value):
                return (
                    str(value or "")
                    .strip()
                    .lower()
                    .replace("-", "_")
                    .replace(" ", "_")
                )

            def _is_debug_admin_user(user):
                if not user:
                    return False
                if getattr(user, "is_superuser", False):
                    return True
                role_names = set(
                    _normalize_debug_key(v)
                    for v in user.roles.values_list("name", flat=True)
                )
                group_codes = set(
                    _normalize_debug_key(v)
                    for v in user.user_groups.values_list("code", flat=True)
                )
                group_names = set(
                    _normalize_debug_key(v)
                    for v in user.user_groups.values_list("name", flat=True)
                )
                debug_keys = {"debug_admin", "debugadmin"}
                return bool(
                    (role_names & debug_keys)
                    or (group_codes & debug_keys)
                    or (group_names & debug_keys)
                )

            if create_missing_lot_expiry_for_missing and not _is_debug_admin_user(
                request.user
            ):
                return Response(
                    {
                        "detail": (
                            "You are not allowed to enable missing lot/expiry auto-creation. "
                            "This option is restricted to debug admin users."
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
            auto_pick_lot_for_negative = str(
                request.data.get("auto_pick_lot_for_negative", "true")
            ).lower() in ("1", "true", "yes", "y", "on")
            auto_pick_lot_for_positive = str(
                request.data.get("auto_pick_lot_for_positive", "true")
            ).lower() in ("1", "true", "yes", "y", "on")
            lot_pick_strategy = (
                str(request.data.get("lot_pick_strategy", "fifo")).strip().lower()
                or "fifo"
            )
            auto_fill_unit_amount = str(
                request.data.get("auto_fill_unit_amount", "true")
            ).lower() in ("1", "true", "yes", "y", "on")
            default_adjustment_type = (
                str(request.data.get("default_adjustment_type", "operational"))
                .strip()
                .lower()
            )
            if default_adjustment_type not in ("operational", "opening_balance"):
                default_adjustment_type = "operational"

            task = process_journal_import.delay(
                df_json=None,
                schema_name=schema_name,
                user_id=request.user.id,
                file_path=file_path,
                branch_id=getattr(branch, "id", None),
                create_missing_items=create_missing_items,
                default_tracking_code=default_tracking_code,
                auto_pick_lot_for_negative=auto_pick_lot_for_negative,
                auto_pick_lot_for_positive=auto_pick_lot_for_positive,
                lot_pick_strategy=lot_pick_strategy,
                auto_fill_unit_amount=auto_fill_unit_amount,
                create_missing_lot_expiry_for_missing=create_missing_lot_expiry_for_missing,
                missing_lot_expiry_date=missing_lot_expiry_date,
                default_adjustment_type=default_adjustment_type,
            )

            return Response(
                {
                    "task_id": task.id,
                    "message": "Import started. Use the task_id to check status.",
                    "status": "pending",
                },
                status=status.HTTP_202_ACCEPTED,
            )
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=["get"],
        authentication_classes=[JWTAuthentication],
        permission_classes=[IsAuthenticated],
    )
    def import_adjustments_status(self, request):
        """
        Poll the status of an adjustment import Celery task.
        Query params: task_id
        """
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response(
                {"detail": "task_id is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            task_result = AsyncResult(task_id)

            if task_result.state == PENDING:
                response = {
                    "task_id": task_id,
                    "status": "pending",
                    "message": "Task is waiting to be processed",
                    "progress": 0,
                }
            elif task_result.state == "PROGRESS":
                meta = task_result.info or {}
                response = {
                    "task_id": task_id,
                    "status": "in_progress",
                    "message": meta.get("message", "Processing..."),
                    "progress": meta.get("progress", 0),
                }
            elif task_result.state == SUCCESS:
                result = task_result.result or {}
                meta = result.get("meta", result)
                response = {
                    "task_id": task_id,
                    "status": "success",
                    "message": meta.get("message", "Import completed"),
                    "progress": 100,
                    "created_count": meta.get("journals_created", 0),
                    "updated_count": meta.get("journals_updated", 0),
                    "failed_count": meta.get("failed_count", 0),
                    "total_rows": meta.get("total", 0),
                    "errors": meta.get("errors", []),
                    "errors_download_key": meta.get("errors_download_key"),
                }
            elif task_result.state == FAILURE:
                response = {
                    "task_id": task_id,
                    "status": "failure",
                    "message": str(task_result.info),
                    "error": str(task_result.info),
                    "progress": 0,
                }
            else:
                response = {
                    "task_id": task_id,
                    "status": task_result.state.lower(),
                    "message": f"Task state: {task_result.state}",
                    "progress": 0,
                }

            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"detail": f"Error checking task status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(
        detail=False,
        methods=["get"],
        authentication_classes=[JWTAuthentication],
        permission_classes=[IsAuthenticated],
    )
    def import_adjustments_errors_download(self, request):
        """
        Download the full error list for an adjustment import as a text file.

        Query params:
          - key: cache key returned as `errors_download_key`
        """
        key = request.query_params.get("key")
        if not key:
            return Response(
                {"detail": "key is required"}, status=status.HTTP_400_BAD_REQUEST
            )
        errors = cache.get(key)
        if not errors:
            return Response(
                {"detail": "No errors found (link expired or invalid key)"},
                status=status.HTTP_404_NOT_FOUND,
            )
        if not isinstance(errors, list):
            errors = [str(errors)]

        content = "\n".join(str(e) for e in errors)
        resp = HttpResponse(content, content_type="text/plain; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="adjustment_import_errors.txt"'
        return resp


class ItemImagesViewSet(viewsets.ModelViewSet):
    queryset = ItemImages.objects.all()
    serializer_class = ItemImagesSerializer
    permission_classes = [IsAuthenticated]
    # Bearer JWT first; Session first would enforce CSRF on unsafe methods when admin cookie is present.
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    parser_classes = [MultiPartParser, FormParser]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    def get_queryset(self):
        qs = ItemImages.objects.all().order_by("-created_at")
        item_no = self.request.query_params.get("item")
        if item_no:
            qs = qs.filter(item__no=item_no)
        return qs

    def perform_create(self, serializer):
        serializer.save()


def validate_quantity_balance(item_journal_id, request=None):
    """
    Standalone function to validate that the total quantity from tracking specifications
    equals the expected quantity for an item journal.

    Args:
        item_journal_id: The ID of the ItemJournal to validate
        request: Optional request object for adding error messages

    Returns:
        dict: {
            'is_valid': bool,
            'expected_quantity': int,
            'total_quantity': int,
            'remaining_quantity': int,
            'error_message': str (if validation fails)
        }
    """
    try:
        # Get the item journal
        item_journal = ItemJournal.objects.get(id=item_journal_id)

        # Get the unit of measure
        if not item_journal.item_unit_of_measure:
            error_msg = (
                f"Unit of measure is required for document {item_journal.document_no}"
            )
            if request:
                messages.error(request, error_msg)
            return {
                "is_valid": False,
                "expected_quantity": 0,
                "total_quantity": 0,
                "remaining_quantity": 0,
                "error_message": error_msg,
            }

        item_unit_of_measure = ItemUnitOfMeasure.objects.get(
            id=item_journal.item_unit_of_measure.id
        )

        # Calculate expected quantity
        expected_quantity = int(item_journal.quantity) * int(
            item_unit_of_measure.quantity_per_unit
        )

        # Get all tracking specifications for this item journal
        specifications = TrackingSpecification.objects.filter(
            item_journal=item_journal.id, item=item_journal.item
        )

        # Calculate total quantity from specifications
        total_quantity = (
            specifications.aggregate(total=Sum("quantity_base"))["total"] or 0
        )

        # Calculate remaining quantity
        remaining_quantity = expected_quantity - total_quantity

        # Only validate quantities if the item has tracking code
        if item_journal.item.tracking_code:
            # Check if quantities match
            print(item_journal.item.tracking_code)
            is_valid = total_quantity == expected_quantity

            # If validation fails, create error message
            if not is_valid:
                error_msg = (
                    f"Quantity mismatch for document {item_journal.document_no}: "
                    f"Expected {expected_quantity} (from {item_journal.quantity} × {item_unit_of_measure.quantity_per_unit}), "
                    f"but tracking specifications total {total_quantity}. "
                    f"Please ensure all items have proper tracking specifications."
                )
                if request:
                    messages.error(request, error_msg)
            else:
                error_msg = None
        else:
            # If item doesn't have tracking code, validation is always valid
            is_valid = True
            error_msg = None

        return {
            "is_valid": is_valid,
            "expected_quantity": expected_quantity,
            "total_quantity": total_quantity,
            "remaining_quantity": remaining_quantity,
            "error_message": error_msg,
        }

    except ItemJournal.DoesNotExist:
        error_msg = f"Item journal with ID {item_journal_id} not found"
        if request:
            messages.error(request, error_msg)
        return {
            "is_valid": False,
            "expected_quantity": 0,
            "total_quantity": 0,
            "remaining_quantity": 0,
            "error_message": error_msg,
        }
    except ItemUnitOfMeasure.DoesNotExist:
        error_msg = f"Unit of measure not found for document {item_journal.document_no if 'item_journal' in locals() else 'Unknown'}"
        if request:
            messages.error(request, error_msg)
        return {
            "is_valid": False,
            "expected_quantity": 0,
            "total_quantity": 0,
            "remaining_quantity": 0,
            "error_message": error_msg,
        }
    except Exception as e:
        error_msg = f"Error validating quantity balance: {str(e)}"
        if request:
            messages.error(request, error_msg)
        return {
            "is_valid": False,
            "expected_quantity": 0,
            "total_quantity": 0,
            "remaining_quantity": 0,
            "error_message": error_msg,
        }


# ========== Item Attribute ViewSets ==========


class ItemAttributeValueViewSet(viewsets.ModelViewSet):
    queryset = ItemAttributeValue.objects.filter(blocked=False)
    serializer_class = ItemAttributeValueSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        queryset = ItemAttributeValue.objects.all()
        # Filter by blocked status if needed
        blocked = self.request.query_params.get("blocked", None)
        if blocked is not None:
            queryset = queryset.filter(blocked=blocked.lower() == "true")
        return queryset.order_by("value")


class ItemAttributeViewSet(viewsets.ModelViewSet):
    queryset = ItemAttribute.objects.filter(blocked=False)
    serializer_class = ItemAttributeSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        queryset = ItemAttribute.objects.all()
        # Filter by blocked status if needed
        blocked = self.request.query_params.get("blocked", None)
        if blocked is not None:
            queryset = queryset.filter(blocked=blocked.lower() == "true")
        # Search by name
        search = self.request.query_params.get("search", None)
        if search is not None:
            queryset = queryset.filter(name__icontains=search)
        return queryset.order_by("name")

    @action(detail=True, methods=["get"])
    def values(self, request, pk=None):
        """Get all values for a specific attribute"""
        attribute = self.get_object()
        values = attribute.values.filter(blocked=False).order_by("value")
        serializer = ItemAttributeValueSerializer(values, many=True)
        return Response({"values": serializer.data})


class ItemAttributeEntryViewSet(viewsets.ModelViewSet):
    queryset = ItemAttributeEntry.objects.all()
    serializer_class = ItemAttributeEntrySerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None
    filterset_fields = [
        "attribute"
    ]  # Removed "item" - handled manually in get_queryset

    def get_queryset(self):
        queryset = ItemAttributeEntry.objects.select_related(
            "item", "attribute"
        ).prefetch_related("selected_values")
        # Filter by item system_id
        item_system_id = self.request.query_params.get("item", None)
        if item_system_id:
            queryset = queryset.filter(item__system_id=item_system_id)
        return queryset.order_by("attribute__name")

    @action(detail=False, methods=["post"])
    def upsert(self, request):
        """Create or update an attribute entry"""
        try:
            item_system_id = request.data.get("item")
            attribute_id = request.data.get("attribute")

            if not item_system_id or not attribute_id:
                return Response(
                    {"error": "item and attribute are required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Get item by system_id
            try:
                item = Item.objects.get(system_id=item_system_id)
            except Item.DoesNotExist:
                return Response(
                    {"error": "Item not found"}, status=status.HTTP_404_NOT_FOUND
                )

            # Get or create entry
            entry, created = ItemAttributeEntry.objects.get_or_create(
                item=item, attribute_id=attribute_id
            )

            # Update entry with provided data
            serializer = ItemAttributeEntrySerializer(
                entry, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class LocationViewSet(viewsets.ModelViewSet):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        """
        Optionally restricts the returned locations to a given search term,
        by filtering against a `q` query parameter in the URL.
        """
        queryset = Location.objects.all()
        q = self.request.query_params.get("q", None)
        if q is not None:
            queryset = queryset.filter(
                Q(code__icontains=q) | Q(description__icontains=q)
            )
        return queryset.order_by("code")


# ── Item Variant ViewSets ─────────────────────────────────────────────────────

from itertools import product as iterproduct
from items.models import ItemVariantOption, ItemVariantOptionValue, ItemVariant
from items.serializers import (
    ItemVariantOptionSerializer,
    ItemVariantOptionValueSerializer,
    ItemVariantSerializer,
)


class ItemVariantOptionViewSet(viewsets.ModelViewSet):
    serializer_class = ItemVariantOptionSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        qs = ItemVariantOption.objects.prefetch_related("values").all()
        item_no = self.request.query_params.get("item")
        if item_no:
            qs = qs.filter(item__no=item_no)
        return qs.order_by("display_order", "name")


class ItemVariantOptionValueViewSet(viewsets.ModelViewSet):
    serializer_class = ItemVariantOptionValueSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        qs = ItemVariantOptionValue.objects.all()
        option_id = self.request.query_params.get("option")
        if option_id:
            qs = qs.filter(option_id=option_id)
        return qs.order_by("display_order", "value")


class ItemVariantViewSet(viewsets.ModelViewSet):
    serializer_class = ItemVariantSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    pagination_class = None

    def get_queryset(self):
        qs = ItemVariant.objects.prefetch_related("option_values").all()
        item_no = self.request.query_params.get("item")
        if item_no:
            qs = qs.filter(item__no=item_no)
        return qs.order_by("code")

    @action(detail=False, methods=["post"], url_path="generate")
    def generate_variants(self, request):
        """
        Auto-generate all option-value combinations as ItemVariant rows.
        POST body: { "item": "<item_no>" }
        Returns the list of created (or already-existing) variants.
        """
        item_no = request.data.get("item")
        if not item_no:
            return Response({"error": "item is required"}, status=400)

        try:
            from items.models import Item as ItemModel
            item = ItemModel.objects.get(no=item_no)
        except ItemModel.DoesNotExist:
            return Response({"error": "Item not found"}, status=404)

        options = list(
            ItemVariantOption.objects.filter(item=item).prefetch_related("values")
        )
        if not options:
            return Response({"error": "No variant options defined for this item"}, status=400)

        # Build all combinations: e.g. [(XL-val, Red-val), (XL-val, Blue-val), ...]
        value_lists = [list(opt.values.order_by("display_order", "value")) for opt in options]
        combinations = list(iterproduct(*value_lists))

        created = []
        for combo in combinations:
            code = "-".join(v.value.upper().replace(" ", "_") for v in combo)
            description = " / ".join(v.value for v in combo)
            variant, _ = ItemVariant.objects.get_or_create(
                item=item,
                code=code,
                defaults={"description": description},
            )
            variant.option_values.set(combo)
            created.append(variant)

        serializer = ItemVariantSerializer(created, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="availability")
    def availability_by_variant(self, request):
        """
        Returns per-variant inventory for an item.
        GET ?item=<item_no>
        """
        item_no = request.query_params.get("item")
        if not item_no:
            return Response({"error": "item query param required"}, status=400)

        variants = ItemVariant.objects.filter(item__no=item_no).prefetch_related("option_values")
        data = [
            {
                "id": v.id,
                "code": v.code,
                "description": v.description,
                "unit_price": v.unit_price,
                "effective_price": v.effective_price,
                "inventory": v.inventory,
                "blocked": v.blocked,
                "option_value_labels": [str(ov) for ov in v.option_values.all()],
            }
            for v in variants
        ]
        return Response(data)
