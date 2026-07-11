from io import BytesIO
from unittest.mock import MagicMock, patch

import pandas as pd
from django.test import RequestFactory, SimpleTestCase
from openpyxl import load_workbook


class ParseRowQuantityTests(SimpleTestCase):
    def test_blank_quantity_is_zero(self):
        from items.tasks import _parse_row_quantity

        row = pd.Series({"Quantity": ""})
        self.assertEqual(_parse_row_quantity(row), 0)

    def test_numeric_quantity(self):
        from items.tasks import _parse_row_quantity

        row = pd.Series({"Quantity": "12.5"})
        self.assertEqual(_parse_row_quantity(row), 12)


class ItemImportTemplateTests(SimpleTestCase):
    @patch("items.views.ItemCategory")
    @patch("items.views.UnitOfMeasure")
    def test_standard_template_excludes_quantity(self, mock_uom, mock_cat):
        from items.views import ItemsModalViewSet

        mock_uom.objects.values_list.return_value.order_by.return_value = ["PCS"]
        mock_cat.objects.values_list.return_value.order_by.return_value = []

        request = RequestFactory().get(
            "/api/items/import_template/", {"import_mode": "standard"}
        )
        request.user = MagicMock()

        response = ItemsModalViewSet().import_template(request)
        self.assertEqual(response.status_code, 200)

        wb = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in wb["Items"][1]]
        self.assertNotIn("Quantity", headers)
        self.assertIn("item_import_template.xlsx", response["Content-Disposition"])

    @patch("items.views.ItemCategory")
    @patch("items.views.UnitOfMeasure")
    def test_opening_balance_template_includes_quantity(self, mock_uom, mock_cat):
        from items.views import ItemsModalViewSet

        mock_uom.objects.values_list.return_value.order_by.return_value = ["PCS"]
        mock_cat.objects.values_list.return_value.order_by.return_value = []

        request = RequestFactory().get(
            "/api/items/import_template/", {"import_mode": "opening_balance"}
        )
        request.user = MagicMock()

        response = ItemsModalViewSet().import_template(request)
        self.assertEqual(response.status_code, 200)

        wb = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in wb["Items"][1]]
        self.assertIn("Quantity", headers)
        self.assertIn("Unit Cost", headers)
        self.assertIn(
            "item_import_template_opening_balance.xlsx",
            response["Content-Disposition"],
        )


class ProcessItemsOpeningBalanceTests(SimpleTestCase):
    @patch("items.tasks._create_opening_balance_journal_from_item_import")
    @patch("items.tasks._process_single_item")
    @patch("items.tasks.User")
    def test_opening_balance_mode_creates_journal_when_quantity_set(
        self, mock_user_cls, mock_process_item, mock_create_journal
    ):
        from items.tasks import _process_items

        item = MagicMock()
        item.pk = 1
        mock_process_item.return_value = item
        mock_user_cls.objects.get.return_value = MagicMock(id=1)
        def _fake_create_journal(**kwargs):
            kwargs["stats"]["journals_created"] = kwargs["stats"].get(
                "journals_created", 0
            ) + 1
            kwargs["stats"].setdefault("journal_document_nos", []).append("DOC-001")
            return "DOC-001"

        mock_create_journal.side_effect = _fake_create_journal

        df = pd.DataFrame(
            [
                {
                    "Item Name": "Test Sugar",
                    "Type": "Inventory",
                    "Unit of Measure": "PCS",
                    "Unit Price": 1000,
                    "Quantity": 5,
                }
            ]
        )
        task = MagicMock()
        result = _process_items(
            task,
            df,
            user_id=1,
            branch_id=None,
            import_mode="opening_balance",
        )

        self.assertEqual(result["journals_created"], 1)
        mock_create_journal.assert_called_once()
        self.assertEqual(result["journal_document_nos"], ["DOC-001"])

    @patch("items.tasks._create_opening_balance_journal_from_item_import")
    @patch("items.tasks._process_single_item")
    @patch("items.tasks.User")
    def test_standard_mode_ignores_quantity_with_warning(
        self, mock_user_cls, mock_process_item, mock_create_journal
    ):
        from items.tasks import _process_items

        mock_process_item.return_value = MagicMock(pk=1)
        mock_user_cls.objects.get.return_value = MagicMock(id=1)

        df = pd.DataFrame(
            [
                {
                    "Item Name": "Test Sugar",
                    "Type": "Inventory",
                    "Quantity": 5,
                }
            ]
        )
        task = MagicMock()
        result = _process_items(
            task,
            df,
            user_id=1,
            import_mode="standard",
        )

        self.assertEqual(result["journals_created"], 0)
        mock_create_journal.assert_not_called()
        self.assertTrue(
            any("ignored" in w.lower() for w in result.get("warnings", []))
        )
