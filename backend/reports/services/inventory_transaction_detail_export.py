"""Excel and PDF export for Inventory Transaction Detail report."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

PAGE_SIZE = landscape(A4)

_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_ITEM_HDR = PatternFill("solid", fgColor="404040")
_FILL_OPENING = PatternFill("solid", fgColor="F2F2F2")
_FILL_SUBTOTAL = PatternFill("solid", fgColor="D9E2F3")
_FILL_GRAND = PatternFill("solid", fgColor="1F3864")
_FILL_GREEN = PatternFill("solid", fgColor="E8F5E9")
_FILL_RED = PatternFill("solid", fgColor="FFEBEE")
_FILL_AMBER = PatternFill("solid", fgColor="FFF8E1")

_FONT_WHITE = Font(color="FFFFFF", bold=True)
_FONT_BOLD = Font(bold=True)
_FONT_ITALIC = Font(italic=True)
_NUM_FMT = "#,##0.##"

_COL_WIDTHS = {"A": 14, "B": 18, "C": 18, "D": 35, "E": 12, "F": 12, "G": 12, "H": 14, "I": 14}

_HEADERS = [
    "Posting Date",
    "Entry Type",
    "Document No.",
    "Description",
    "Increases",
    "Decreases",
    "Inventory",
    "Cost Amount",
    "Running Cost",
]

_NUM_COLS = 9


def _num(val):
    if val is None or val == "" or val == "—":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _write_row(ws, row_idx, values, fills=None, fonts=None, number_cols=None, strike_cols=None):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if fills and col_idx <= len(fills) and fills[col_idx - 1]:
            cell.fill = fills[col_idx - 1]
        if fonts and col_idx <= len(fonts) and fonts[col_idx - 1]:
            cell.font = fonts[col_idx - 1]
        if number_cols and col_idx in number_cols and _num(val) is not None:
            cell.value = _num(val)
            cell.number_format = _NUM_FMT
        if strike_cols and col_idx in strike_cols:
            f = cell.font or Font()
            cell.font = Font(
                name=f.name,
                size=f.size,
                bold=f.bold,
                italic=f.italic,
                color=f.color,
                strike=True,
            )


def build_inventory_transaction_detail_excel(
    report_data: dict, company_name: str = "", generated_by: str = ""
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Detail"

    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    start_date = report_data.get("start_date", "")
    end_date = report_data.get("end_date", "")
    generated_at = report_data.get("generated_at", datetime.now().isoformat())
    items = report_data.get("items", [])
    summary = report_data.get("summary", {})

    ws.cell(row=1, column=1, value=company_name or "")
    ws.cell(row=1, column=_NUM_COLS, value=generated_at[:10] if generated_at else "")
    ws.cell(row=2, column=1, value="Inventory - Transaction Detail").font = _FONT_BOLD
    ws.cell(row=3, column=1, value=f"Period: {start_date} to {end_date}")
    header_row = 5
    for col_idx, h in enumerate(_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = _FONT_WHITE
        c.fill = _FILL_HEADER

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:I{header_row}"

    row = header_row + 1
    number_cols = {5, 6, 7, 8, 9}

    for item in items:
        item_label = (
            f"{item.get('item_no', '')}  {item.get('item_name', '')}  "
            f"[{item.get('unit_of_measure', 'PCS')}]"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
        hdr_cell = ws.cell(row=row, column=1, value=item_label)
        hdr_cell.font = _FONT_WHITE
        hdr_cell.fill = _FILL_ITEM_HDR
        row += 1

        _write_row(
            ws,
            row,
            [
                "",
                "Opening Balance",
                "",
                "",
                "",
                "",
                _num(item.get("opening_qty")) or item.get("opening_qty"),
                "",
                _num(item.get("opening_cost")) or item.get("opening_cost"),
            ],
            fills=[_FILL_OPENING] * _NUM_COLS,
            fonts=[_FONT_ITALIC] * _NUM_COLS,
            number_cols={7, 9},
        )
        row += 1

        for entry in item.get("entries", []):
            inc = _num(entry.get("increases"))
            dec = _num(entry.get("decreases"))
            fills = [None] * _NUM_COLS
            strike = set()
            if entry.get("reversed"):
                fills = [_FILL_AMBER] * _NUM_COLS
                strike = set(range(1, _NUM_COLS + 1))
            elif inc and inc > 0:
                fills = [_FILL_GREEN] * _NUM_COLS
            elif dec and dec > 0:
                fills = [_FILL_RED] * _NUM_COLS

            _write_row(
                ws,
                row,
                [
                    entry.get("posting_date", ""),
                    entry.get("entry_type", ""),
                    entry.get("document_no", ""),
                    entry.get("description", ""),
                    entry.get("increases") if inc else "",
                    entry.get("decreases") if dec else "",
                    entry.get("running_qty", ""),
                    entry.get("cost_amount", ""),
                    entry.get("running_cost", ""),
                ],
                fills=fills,
                number_cols=number_cols,
                strike_cols=strike,
            )
            row += 1

        net_cost = ""
        try:
            cin = float(item.get("total_cost_in") or 0)
            cout = float(item.get("total_cost_out") or 0)
            net_cost = round(cin + cout, 2)
        except (TypeError, ValueError):
            pass

        _write_row(
            ws,
            row,
            [
                "",
                item.get("item_name", ""),
                "",
                "",
                item.get("total_increases", ""),
                item.get("total_decreases", ""),
                item.get("closing_qty", ""),
                net_cost,
                item.get("closing_cost", ""),
            ],
            fills=[_FILL_SUBTOTAL] * _NUM_COLS,
            fonts=[_FONT_BOLD] * _NUM_COLS,
            number_cols=number_cols,
        )
        row += 2

    grand_net = ""
    try:
        g_open = float(summary.get("grand_opening_cost") or 0)
        g_close = float(summary.get("grand_closing_cost") or 0)
        grand_net = round(g_close - g_open, 2)
    except (TypeError, ValueError):
        pass

    _write_row(
        ws,
        row,
        [
            "GRAND TOTAL",
            "",
            "",
            "",
            summary.get("grand_total_increases", ""),
            summary.get("grand_total_decreases", ""),
            "",
            grand_net,
            summary.get("grand_closing_cost", ""),
        ],
        fills=[_FILL_GRAND] * _NUM_COLS,
        fonts=[_FONT_WHITE] * _NUM_COLS,
        number_cols=number_cols,
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pdf_page_callbacks(company_name, period_label, generated_by):
    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(15 * mm, PAGE_SIZE[1] - 12 * mm, company_name or "")
        canvas.drawCentredString(
            PAGE_SIZE[0] / 2, PAGE_SIZE[1] - 12 * mm, "Inventory - Transaction Detail"
        )
        canvas.drawRightString(
            PAGE_SIZE[0] - 15 * mm,
            PAGE_SIZE[1] - 12 * mm,
            f"{datetime.now().strftime('%Y-%m-%d')}  p.{doc.page}",
        )
        canvas.drawString(15 * mm, 10 * mm, period_label)
        canvas.drawRightString(
            PAGE_SIZE[0] - 15 * mm, 10 * mm, f"Generated by: {generated_by or ''}"
        )
        canvas.restoreState()

    return on_page


def build_inventory_transaction_detail_pdf(
    report_data: dict, company_name: str = "", generated_by: str = ""
) -> BytesIO:
    buf = BytesIO()
    start_date = report_data.get("start_date", "")
    end_date = report_data.get("end_date", "")
    period_label = f"Period: {start_date} to {end_date}"
    items = report_data.get("items", [])
    summary = report_data.get("summary", {})

    doc = SimpleDocTemplate(
        buf,
        pagesize=PAGE_SIZE,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=18 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7, leading=8)
    story = []

    pdf_headers = [
        "Date",
        "Entry Type",
        "Doc No.",
        "Description",
        "Increases",
        "Decreases",
        "Inventory",
        "Cost Amt",
    ]
    col_widths = [22 * mm, 24 * mm, 22 * mm, 58 * mm, 18 * mm, 18 * mm, 18 * mm, 24 * mm]

    def _fmt_cell(v):
        if v is None or v == "":
            return "—"
        return str(v)

    table_data = [[Paragraph(h, small) for h in pdf_headers]]
    row_styles = []

    for item in items:
        item_label = (
            f"<b>{item.get('item_no', '')} {item.get('item_name', '')} "
            f"[{item.get('unit_of_measure', 'PCS')}]</b>"
        )
        table_data.append(
            [Paragraph(item_label, small)]
            + [""] * (len(pdf_headers) - 1)
        )
        row_styles.append(("BACKGROUND", (0, len(table_data) - 1), (-1, len(table_data) - 1), colors.HexColor("#E0E0E0")))

        table_data.append(
            [
                "",
                "Opening Balance",
                "",
                "",
                "",
                "",
                _fmt_cell(item.get("opening_qty")),
                _fmt_cell(item.get("opening_cost")),
            ]
        )

        for entry in item.get("entries", []):
            table_data.append(
                [
                    _fmt_cell(entry.get("posting_date")),
                    _fmt_cell(entry.get("entry_type")),
                    _fmt_cell(entry.get("document_no")),
                    _fmt_cell(entry.get("description"))[:60],
                    _fmt_cell(entry.get("increases") if _num(entry.get("increases")) else ""),
                    _fmt_cell(entry.get("decreases") if _num(entry.get("decreases")) else ""),
                    _fmt_cell(entry.get("running_qty")),
                    _fmt_cell(entry.get("cost_amount")),
                ]
            )
            idx = len(table_data) - 1
            if entry.get("reversed"):
                row_styles.append(
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FFF8E1"))
                )
            elif _num(entry.get("increases")):
                row_styles.append(
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#E8F5E9"))
                )
            elif _num(entry.get("decreases")):
                row_styles.append(
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FFEBEE"))
                )

        net_cost = ""
        try:
            net_cost = round(
                float(item.get("total_cost_in") or 0)
                + float(item.get("total_cost_out") or 0),
                2,
            )
        except (TypeError, ValueError):
            pass

        table_data.append(
            [
                "",
                item.get("item_name", ""),
                "",
                "",
                _fmt_cell(item.get("total_increases")),
                _fmt_cell(item.get("total_decreases")),
                _fmt_cell(item.get("closing_qty")),
                _fmt_cell(net_cost),
            ]
        )
        idx = len(table_data) - 1
        row_styles.append(("LINEABOVE", (0, idx), (-1, idx), 0.5, colors.grey))
        row_styles.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))

    table_data.append(
        [
            "GRAND TOTAL",
            "",
            "",
            "",
            _fmt_cell(summary.get("grand_total_increases")),
            _fmt_cell(summary.get("grand_total_decreases")),
            "",
            _fmt_cell(summary.get("grand_closing_cost")),
        ]
    )
    idx = len(table_data) - 1
    row_styles.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#1F3864")))
    row_styles.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.white))
    row_styles.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))

    t_style = [
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#4472C4")),
    ]
    t_style.extend(row_styles)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(t_style))
    story.append(table)
    story.append(Spacer(1, 4 * mm))

    doc.build(
        story,
        onFirstPage=_pdf_page_callbacks(company_name, period_label, generated_by),
        onLaterPages=_pdf_page_callbacks(company_name, period_label, generated_by),
    )
    buf.seek(0)
    return buf
