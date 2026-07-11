"""PDF and Excel export for Inventory Value Movement report."""

from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PAGE_SIZE = landscape(A4)
PAGE_WIDTH, PAGE_HEIGHT = PAGE_SIZE
USABLE_WIDTH = PAGE_WIDTH - 30 * mm

# Brand palette
_BLUE = "#1F3FFF"
_INDIGO_LIGHT = "#EEF2FF"
_INDIGO_BORDER = "#C7D2FE"
_GRAY_700 = "#374151"
_GRAY_400 = "#9CA3AF"
_GRAY_200 = "#E5E7EB"
_GREEN = "#059669"
_RED = "#DC2626"


from financials.currency import get_local_currency_code
from reports.utils.formatters import format_currency


def _cur(label: str) -> str:
    return f"{label} ({get_local_currency_code()})"


def _format_ugx(amount) -> str:
    return format_currency(amount)


def _dash_if_zero(amount: float, formatted: bool = True) -> str:
    if not amount:
        return "—"
    return _format_ugx(amount) if formatted else str(amount)


def _movement_rows(report_payload: dict) -> list:
    return report_payload.get("data", {}).get("buckets", [])


def _summary(report_payload: dict) -> dict:
    return report_payload.get("data", {}).get("summary", {})


def _period_meta(report_payload: dict) -> dict:
    return report_payload.get("period", {})


def _export_details(report_payload: dict) -> dict:
    return report_payload.get("export_details") or {}


def _value_entry_reconciliation_story(summary: dict, styles) -> list:
    """PDF block comparing headline G/L figures to ValueEntry costing."""
    ve = summary.get("value_entry_reconciliation") or {}
    if not ve.get("available"):
        return []

    account = (
        f"{summary.get('gl_account_no', '2110')} "
        f"{summary.get('gl_account_name', 'Resale Items')}"
    ).strip()

    def _variance_cell(v: float) -> str:
        if abs(v) < 0.01:
            return "—"
        sign = "+" if v > 0 else ""
        return f"{sign}{_format_ugx(v)}"

    rows = [
        ["", "G/L (report)", "ValueEntry", "Variance (G/L − VE)"],
        [
            "Opening",
            _format_ugx(summary.get("opening_value", 0)),
            _format_ugx(ve.get("opening_balance", 0)),
            _variance_cell(ve.get("opening_variance", 0)),
        ],
        [
            "Closing",
            _format_ugx(summary.get("closing_value", 0)),
            _format_ugx(ve.get("closing_balance", 0)),
            _variance_cell(ve.get("closing_variance", 0)),
        ],
        [
            "Stock in (period)",
            _format_ugx(summary.get("inbound_value", 0)),
            _format_ugx(ve.get("inbound_balance", 0)),
            _variance_cell(ve.get("period_inbound_variance", 0)),
        ],
    ]
    for line in ve.get("stock_in_breakdown") or []:
        rows.append(
            [
                f"    {line.get('label', '')} ({line.get('transaction_count', 0)})",
                "—",
                _format_ugx(line.get("amount", 0)),
                "—",
            ]
        )
    rows.append(
        [
            "Stock out (period)",
            _format_ugx(summary.get("outbound_value", 0)),
            _format_ugx(ve.get("outbound_balance", 0)),
            _variance_cell(ve.get("period_outbound_variance", 0)),
        ]
    )
    for line in ve.get("stock_out_breakdown") or []:
        rows.append(
            [
                f"    {line.get('label', '')} ({line.get('transaction_count', 0)})",
                "—",
                _format_ugx(line.get("amount", 0)),
                "—",
            ]
        )
    table = Table(
        rows,
        colWidths=[42 * mm, 40 * mm, 40 * mm, 44 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            _table_style_base(8, header_bg=_INDIGO_LIGHT, header_fg=_BLUE) + [
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    return [
        Spacer(1, 4 * mm),
        _pdf_section_header("ValueEntry Comparison"),
        Spacer(1, 2 * mm),
        Paragraph(
            f"Headline amounts are from G/L account <b>{account}</b> "
            "(same as Sales dashboard). ValueEntry shows item-level costing detail.",
            styles["Normal"],
        ),
        Spacer(1, 2 * mm),
        table,
        Spacer(1, 2 * mm),
        Paragraph(
            "Indented rows under stock in/out are ValueEntry by type (purchase, sales, "
            "adjustments). Positive variance = G/L higher than ValueEntry.",
            styles["Normal"],
        ),
        Spacer(1, 4 * mm),
    ]


def _net_change_color(net_change: float):
    if net_change > 0:
        return colors.HexColor("#276221")
    if net_change < 0:
        return colors.HexColor("#9C0006")
    return colors.black


def _table_style_base(font_size: int, header_bg=_BLUE, header_fg="#FFFFFF"):
    return [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_bg)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(header_fg)),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(_GRAY_200)),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(_GRAY_400)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]


def _make_page_callbacks(title: str, subtitle: str, branch_label: str):
    def _draw_page(canvas, doc):
        canvas.saveState()
        # Header bar
        canvas.setFillColor(colors.HexColor(_BLUE))
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(15 * mm, PAGE_HEIGHT - 12 * mm, title)
        canvas.setFillColor(colors.HexColor(_GRAY_700))
        canvas.setFont("Helvetica", 8)
        y = PAGE_HEIGHT - 17 * mm
        canvas.drawString(15 * mm, y, subtitle)
        if branch_label:
            y -= 4 * mm
            canvas.drawString(15 * mm, y, f"Branch: {branch_label}")
        canvas.setStrokeColor(colors.HexColor(_BLUE))
        canvas.setLineWidth(0.75)
        canvas.line(15 * mm, PAGE_HEIGHT - 21 * mm, PAGE_WIDTH - 15 * mm, PAGE_HEIGHT - 21 * mm)
        # Footer bar
        canvas.setStrokeColor(colors.HexColor(_GRAY_200))
        canvas.setLineWidth(0.5)
        canvas.line(15 * mm, 14 * mm, PAGE_WIDTH - 15 * mm, 14 * mm)
        canvas.setFillColor(colors.HexColor(_GRAY_400))
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(PAGE_WIDTH - 15 * mm, 9 * mm, f"Page {canvas.getPageNumber()}")
        canvas.drawString(15 * mm, 9 * mm, "Inventory Value Movement Report")
        canvas.restoreState()

    return _draw_page, _draw_page


def _pdf_title_banner(title: str, subtitle: str, branch_label: str, generated: str) -> list:
    title_table = Table([[title]], colWidths=[USABLE_WIDTH])
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(_BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 18),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 11),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 11),
    ]))
    meta_parts = [subtitle]
    if branch_label:
        meta_parts.append(f"Branch: {branch_label}")
    meta_parts.append(f"Generated: {generated}")
    meta_table = Table([[" • ".join(meta_parts)]], colWidths=[USABLE_WIDTH])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(_INDIGO_LIGHT)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(_GRAY_700)),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(_INDIGO_BORDER)),
    ]))
    return [title_table, meta_table]


def _pdf_kpi_cards(summary: dict) -> Table:
    col_w = USABLE_WIDTH / 5
    items = [
        ("Opening Value", "opening_value", "#6B7280"),
        ("Stock In  (+)", "inbound_value", _GREEN),
        ("Stock Out  (−)", "outbound_value", _RED),
        ("Net Change", "net_change", "#3B82F6"),
        ("Closing Value", "closing_value", "#7C3AED"),
    ]
    label_row = [item[0] for item in items]
    value_row = [_format_ugx(summary.get(item[1], 0)) for item in items]
    style_cmds = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 1), (-1, 1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 7),
        ("BACKGROUND", (0, 1), (-1, 1), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#111827")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(_GRAY_400)),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor(_GRAY_200)),
    ]
    for idx, (_, _, color) in enumerate(items):
        style_cmds.extend([
            ("BACKGROUND", (idx, 0), (idx, 0), colors.HexColor(color)),
            ("TEXTCOLOR", (idx, 0), (idx, 0), colors.white),
        ])
    net_val = float(summary.get("net_change", 0))
    style_cmds.append(("TEXTCOLOR", (3, 1), (3, 1), _net_change_color(net_val)))
    table = Table([label_row, value_row], colWidths=[col_w] * 5, rowHeights=[None, 14])
    table.setStyle(TableStyle(style_cmds))
    return table


def _pdf_section_header(title: str) -> Table:
    table = Table([[title]], colWidths=[USABLE_WIDTH])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(_INDIGO_LIGHT)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(_BLUE)),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 1.5, colors.HexColor(_BLUE)),
    ]))
    return table


def _pdf_formula_visual(summary: dict) -> Table:
    opening = summary.get("opening_value", 0)
    inbound = summary.get("inbound_value", 0)
    outbound = summary.get("outbound_value", 0)
    net = summary.get("net_change", 0)
    closing = summary.get("closing_value", 0)

    val_w = 45 * mm
    op_w = (USABLE_WIDTH - 5 * val_w) / 4

    net_bg = "#DCFCE7" if float(net) > 0 else "#FEE2E2" if float(net) < 0 else "#F3F4F6"
    box_fills = ["#F3F4F6", None, "#ECFDF5", None, "#FEF2F2", None, net_bg, None, "#FAF5FF"]

    labels = ["Opening\nValue", "", "Stock In\n(+)", "", "Stock Out\n(−)", "", "Net Change", "", "Closing\nValue"]
    values = [_format_ugx(opening), "+", _format_ugx(inbound), "−", _format_ugx(outbound), "=", _format_ugx(net), "→", _format_ugx(closing)]

    col_widths = [val_w, op_w, val_w, op_w, val_w, op_w, val_w, op_w, val_w]
    style_cmds = [
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for idx in [0, 2, 4, 6, 8]:
        bg = box_fills[idx] or "#F3F4F6"
        style_cmds.extend([
            ("BACKGROUND", (idx, 0), (idx, 1), colors.HexColor(bg)),
            ("BOX", (idx, 0), (idx, 1), 0.5, colors.HexColor(_GRAY_200)),
            ("FONTNAME", (idx, 1), (idx, 1), "Helvetica-Bold"),
            ("TEXTCOLOR", (idx, 0), (idx, 0), colors.HexColor("#6B7280")),
        ])
    for idx in [1, 3, 5, 7]:
        style_cmds.extend([
            ("FONTNAME", (idx, 1), (idx, 1), "Helvetica-Bold"),
            ("FONTSIZE", (idx, 1), (idx, 1), 12),
            ("TEXTCOLOR", (idx, 0), (idx, 1), colors.HexColor("#9CA3AF")),
        ])
    net_val = float(net)
    if net_val > 0:
        style_cmds.append(("TEXTCOLOR", (6, 1), (6, 1), colors.HexColor(_GREEN)))
    elif net_val < 0:
        style_cmds.append(("TEXTCOLOR", (6, 1), (6, 1), colors.HexColor(_RED)))
    table = Table([labels, values], colWidths=col_widths)
    table.setStyle(TableStyle(style_cmds))
    return table


def _movement_table_data(buckets: list, summary: dict) -> tuple[list, list]:
    headers = [
        "Period",
        "Opening Value",
        "Stock In (+)",
        "Stock Out (−)",
        "Net Change",
        "Closing Value",
    ]
    table_data = [headers]
    row_nets = []

    for bucket in buckets:
        net = float(bucket.get("net_change", 0))
        row_nets.append(net)
        table_data.append(
            [
                bucket.get("period", ""),
                _format_ugx(bucket.get("opening_value", 0)),
                _format_ugx(bucket.get("inbound_value", 0)),
                _format_ugx(bucket.get("outbound_value", 0)),
                _format_ugx(net),
                _format_ugx(bucket.get("closing_value", 0)),
            ]
        )

    if len(table_data) == 1:
        table_data.append(["No movement data", "-", "-", "-", "-", "-"])
        row_nets.append(0)
    else:
        table_data.append(
            [
                "TOTAL",
                _format_ugx(summary.get("opening_value", 0)),
                _format_ugx(summary.get("inbound_value", 0)),
                _format_ugx(summary.get("outbound_value", 0)),
                _format_ugx(summary.get("net_change", 0)),
                _format_ugx(summary.get("closing_value", 0)),
            ]
        )
        row_nets.append(float(summary.get("net_change", 0)))

    return table_data, row_nets


def _item_tables_story(buckets: list, styles) -> list:
    story = []
    heading_style = ParagraphStyle(
        "ItemHeading",
        parent=styles["Heading4"],
        fontSize=9,
        leading=11,
    )

    for bucket in buckets:
        items = bucket.get("items") or []
        if not items:
            continue
        period = bucket.get("period", "")
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(f"Items moved — {period}", heading_style))
        item_data = [
            ["Item No", "Description", "Stock In", "Stock Out", "Net Qty"],
        ]
        for item in items:
            item_data.append(
                [
                    item.get("item_no", ""),
                    (item.get("item_description", "") or "")[:40],
                    _format_ugx(item.get("inbound_value", 0)),
                    _format_ugx(item.get("outbound_value", 0)),
                    str(item.get("net_qty", 0)),
                ]
            )
        item_table = Table(
            item_data,
            colWidths=[22 * mm, 55 * mm, 32 * mm, 32 * mm, 18 * mm],
        )
        item_table.setStyle(TableStyle(_table_style_base(8) + [("ALIGN", (2, 1), (-1, -1), "RIGHT")]))
        story.append(item_table)
    return story


def _entry_type_table_story(export_details: dict, styles) -> list:
    story = []
    rows = export_details.get("movement_by_entry_type") or []
    if not rows:
        return story

    story.append(Spacer(1, 6 * mm))
    story.append(_pdf_section_header("Movement by Transaction Type"))

    table_data = [
        [
            "Entry Type",
            "Category",
            "Transactions",
            "Stock In",
            "Stock Out",
            "Net Value",
            "Sales Revenue",
        ]
    ]
    sales_row_idx = None
    for idx, row in enumerate(rows, start=1):
        if row.get("entry_type") == "Sales":
            sales_row_idx = idx
        table_data.append(
            [
                row.get("entry_type_label", ""),
                row.get("category", ""),
                str(row.get("transaction_count", 0)),
                _dash_if_zero(row.get("stock_in_value", 0)),
                _dash_if_zero(row.get("stock_out_value", 0)),
                _format_ugx(row.get("net_value", 0)),
                _dash_if_zero(row.get("sales_revenue", 0)),
            ]
        )

    col_widths = [32 * mm, 18 * mm, 16 * mm, 24 * mm, 24 * mm, 24 * mm, 24 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = _table_style_base(8) + [
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
    ]
    if sales_row_idx is not None:
        style_cmds.append(
            ("BACKGROUND", (0, sales_row_idx), (-1, sales_row_idx), colors.HexColor("#fff3cd"))
        )
        style_cmds.append(("FONTNAME", (0, sales_row_idx), (-1, sales_row_idx), "Helvetica-Bold"))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    return story


def _detail_sections_story(export_details: dict, styles) -> list:
    """One full transaction table per entry type (sales, purchases, adjustments, etc.)."""
    story = []
    sections = export_details.get("detail_sections") or []
    if not sections:
        return story

    heading = ParagraphStyle(
        "DetailSectionHeading",
        parent=styles["Heading2"],
        fontSize=11,
        leading=13,
        spaceAfter=4,
    )
    subheading = ParagraphStyle(
        "DetailSectionSub",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#4b5563"),
    )

    story.append(PageBreak())
    story.append(
        Paragraph(
            "Transaction detail by type",
            ParagraphStyle(
                "DetailPartTitle",
                parent=styles["Heading2"],
                fontSize=12,
                leading=14,
            ),
        )
    )
    story.append(
        Paragraph(
            "Every Value Entry in the period is listed below, grouped by transaction type. "
            "Sales show COGS and revenue; purchases and adjustments show how stock value changed.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 4 * mm))

    col_widths = [
        17 * mm,
        20 * mm,
        22 * mm,
        14 * mm,
        28 * mm,
        9 * mm,
        14 * mm,
        20 * mm,
        20 * mm,
        22 * mm,
        32 * mm,
    ]

    for section_idx, section in enumerate(sections):
        if section_idx > 0:
            story.append(PageBreak())

        entry_type = section.get("entry_type", "")
        title = section.get("title", entry_type)
        count = section.get("transaction_count", 0)
        story.append(Paragraph(f"{title} — {count} line(s)", heading))

        summary_parts = [
            f"Category: {section.get('category', '')}",
            f"Stock In: {_format_ugx(section.get('stock_in_total', 0))}",
            f"Stock Out: {_format_ugx(section.get('stock_out_total', 0))}",
            f"Net: {_format_ugx(section.get('net_cost', 0))}",
        ]
        if entry_type == "Sales":
            summary_parts.append(
                f"Total revenue: {_format_ugx(section.get('sales_revenue', 0))}"
            )
        story.append(Paragraph(" | ".join(summary_parts), subheading))
        story.append(Spacer(1, 2 * mm))

        if entry_type == "Sales":
            headers = [
                "Date",
                "Invoice",
                "Item",
                "Description",
                "Qty",
                "COGS",
                "Revenue",
                "Margin",
                "Balance",
                "Doc Type",
                "Notes",
            ]
        else:
            headers = [
                "Date",
                "Document",
                "Item",
                "Description",
                "Qty",
                "Direction",
                "Cost",
                "Balance",
                "Doc Type",
                "Entry",
                "Notes",
            ]

        table_data = [headers]
        for row in section.get("rows") or []:
            item_no = row.get("item_no", "")
            desc = (row.get("item_description") or "")[:24]
            if entry_type == "Sales":
                cogs = row.get("cost_amount", 0)
                rev = row.get("sales_amount", 0)
                table_data.append(
                    [
                        row.get("posting_date", ""),
                        row.get("document_no", ""),
                        item_no,
                        desc,
                        str(abs(row.get("quantity", 0))),
                        _format_ugx(cogs),
                        _format_ugx(rev),
                        _format_ugx(round(rev - cogs, 2)),
                        _format_ugx(row.get("running_balance", 0)),
                        row.get("document_type", "") or "—",
                        (row.get("description") or "")[:36],
                    ]
                )
            else:
                table_data.append(
                    [
                        row.get("posting_date", ""),
                        row.get("document_no", ""),
                        item_no,
                        desc,
                        str(row.get("quantity", 0)),
                        row.get("direction", ""),
                        _format_ugx(row.get("cost_amount", 0)),
                        _format_ugx(row.get("running_balance", 0)),
                        row.get("document_type", "") or "—",
                        row.get("entry_type_label", ""),
                        (row.get("description") or "")[:36],
                    ]
                )

        if len(table_data) == 1:
            table_data.append(["No lines"] + [""] * (len(headers) - 1))

        if entry_type == "Sales":
            totals_row = [
                "TOTAL",
                "",
                "",
                "",
                "",
                "",
                _format_ugx(section.get("stock_out_total", 0)),
                _format_ugx(section.get("sales_revenue", 0)),
                _format_ugx(
                    round(
                        section.get("sales_revenue", 0)
                        - section.get("stock_out_total", 0),
                        2,
                    )
                ),
                "",
                "",
                "",
            ]
        else:
            totals_row = [
                "TOTAL",
                "",
                "",
                "",
                "",
                "",
                _format_ugx(section.get("net_cost", 0)),
                "",
                "",
                "",
                "",
            ]
        table_data.append(totals_row)

        widths = col_widths[: len(headers)]
        table = Table(table_data, colWidths=widths, repeatRows=1)
        style_cmds = _table_style_base(7) + [
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")),
        ]
        if entry_type == "Sales":
            style_cmds.append(
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3cd"))
            )
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

    return story


def _sales_detail_story(export_details: dict, styles) -> list:
    story = []
    sales_rows = export_details.get("sales_detail") or []
    totals = export_details.get("sales_totals") or {}

    story.append(Spacer(1, 6 * mm))
    story.append(_pdf_section_header("Sales Detail"))

    if not sales_rows:
        story.append(Paragraph("No sales transactions in selected period.", styles["Normal"]))
        return story

    table_data = [
        [
            "Date",
            "Invoice No",
            "Item",
            "Qty Sold",
            "COGS",
            "Revenue",
            "Gross Margin",
        ]
    ]
    for row in sales_rows:
        item_label = row.get("item_no", "")
        if row.get("item_description"):
            item_label = f"{item_label} — {(row.get('item_description') or '')[:28]}"
        table_data.append(
            [
                row.get("posting_date", ""),
                row.get("document_no", ""),
                item_label,
                str(row.get("quantity", 0)),
                _format_ugx(row.get("cogs", 0)),
                _format_ugx(row.get("revenue", 0)),
                _format_ugx(row.get("gross_margin", 0)),
            ]
        )
    table_data.append(
        [
            "TOTAL",
            "",
            "",
            "",
            _format_ugx(totals.get("cogs", 0)),
            _format_ugx(totals.get("revenue", 0)),
            _format_ugx(totals.get("gross_margin", 0)),
        ]
    )

    col_widths = [20 * mm, 24 * mm, 42 * mm, 14 * mm, 24 * mm, 24 * mm, 24 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = _table_style_base(7) + [
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")),
    ]
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    return story


def _grouped_ledger_story(export_details: dict, styles) -> list:
    story = []
    groups = export_details.get("ledger_groups") or []

    separator_style = ParagraphStyle(
        "PeriodSeparator",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        backColor=colors.HexColor(_INDIGO_LIGHT),
        textColor=colors.HexColor(_BLUE),
    )
    story.append(Spacer(1, 6 * mm))
    story.append(_pdf_section_header("Transaction Detail — All Value Entries"))

    if not groups:
        story.append(Paragraph("No transactions in selected period.", styles["Normal"]))
        return story

    ledger_headers = [
        "Date",
        "Document",
        "Entry Type",
        "Item",
        "Qty",
        "Direction",
        _cur("Cost"),
        _cur("Revenue"),
        _cur("Balance"),
        "Description",
    ]
    col_widths = [
        18 * mm,
        22 * mm,
        24 * mm,
        30 * mm,
        10 * mm,
        16 * mm,
        22 * mm,
        22 * mm,
        24 * mm,
        38 * mm,
    ]

    for group in groups:
        sep_text = (
            f"─── {group.get('period_label', group.get('period', ''))} ─── "
            f"Opening: {_format_ugx(group.get('opening_value', 0))} │ "
            f"In: +{_format_ugx(group.get('stock_in', 0))} │ "
            f"Out: −{_format_ugx(group.get('stock_out', 0))} │ "
            f"Closing: {_format_ugx(group.get('closing_value', 0))} ───"
        )
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(sep_text, separator_style))

        table_data = [ledger_headers]
        for txn in group.get("transactions") or []:
            item_label = txn.get("item_no", "")
            if txn.get("item_description"):
                item_label = f"{item_label} {(txn.get('item_description') or '')[:20]}"
            revenue_display = (
                _format_ugx(txn.get("revenue", 0))
                if txn.get("revenue", 0)
                else "—"
            )
            table_data.append(
                [
                    txn.get("posting_date", ""),
                    txn.get("document_no", ""),
                    txn.get("entry_type_label", ""),
                    item_label,
                    str(txn.get("quantity", 0)),
                    txn.get("direction", ""),
                    _format_ugx(txn.get("cost_amount", 0)),
                    revenue_display,
                    _format_ugx(txn.get("running_balance", 0)),
                    (txn.get("description") or "")[:50],
                ]
            )

        if len(table_data) == 1:
            table_data.append(["No transactions", "", "", "", "", "", "", "", "", ""])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = _table_style_base(7) + [
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (5, 1), (5, -1), "CENTER"),
        ]
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

    return story


def build_inventory_value_movement_pdf(
    report_payload: dict, branch_label: str = ""
) -> HttpResponse:
    summary = _summary(report_payload)
    buckets = _movement_rows(report_payload)
    period = _period_meta(report_payload)
    export_details = _export_details(report_payload)

    response = HttpResponse(content_type="application/pdf")
    start = period.get("start_date", "")
    end = period.get("end_date", "")
    filename = f"inventory-value-movement-{start}-{end}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    subtitle = f"Period: {start} to {end} ({period.get('period_type', '')})"
    on_first, on_later = _make_page_callbacks(
        "Inventory Value Movement Report",
        subtitle,
        branch_label,
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=PAGE_SIZE,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=22 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    generated = timezone.now().strftime("%Y-%m-%d %H:%M")
    story = _pdf_title_banner("Inventory Value Movement Report", subtitle, branch_label, generated)
    story.append(Spacer(1, 5 * mm))

    # KPI cards
    story.append(_pdf_kpi_cards(summary))
    story.append(Spacer(1, 4 * mm))

    # Formula visual
    story.append(_pdf_formula_visual(summary))
    story.append(Spacer(1, 4 * mm))

    story.extend(_value_entry_reconciliation_story(summary, styles))

    txn_count = export_details.get("transaction_count", 0)
    if txn_count:
        note_style = ParagraphStyle(
            "ExportNote",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.HexColor(_GRAY_700),
            backColor=colors.HexColor(_INDIGO_LIGHT),
            borderPad=4,
        )
        story.append(
            Paragraph(
                f"This export documents <b>{txn_count}</b> inventory value entries, "
                "including every sale, purchase, and adjustment in the period.",
                note_style,
            )
        )
        story.append(Spacer(1, 4 * mm))

    # Main movement table
    story.append(_pdf_section_header("Inventory Value Movement by Period"))
    story.append(Spacer(1, 2 * mm))
    table_data, row_nets = _movement_table_data(buckets, summary)
    table = Table(
        table_data,
        colWidths=[40 * mm, 45 * mm, 45 * mm, 45 * mm, 42 * mm, 45 * mm],
        repeatRows=1,
    )
    style_commands = _table_style_base(9) + [
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.HexColor("#F9FAFB"), colors.white]),
    ]
    if len(table_data) > 1:
        totals_row = len(table_data) - 1
        style_commands.extend([
            ("FONTNAME", (0, totals_row), (-1, totals_row), "Helvetica-Bold"),
            ("BACKGROUND", (0, totals_row), (-1, totals_row), colors.HexColor(_INDIGO_LIGHT)),
            ("LINEABOVE", (0, totals_row), (-1, totals_row), 1.0, colors.HexColor(_BLUE)),
        ])
    for row_idx, net_val in enumerate(row_nets, start=1):
        style_commands.append(
            ("TEXTCOLOR", (4, row_idx), (4, row_idx), _net_change_color(net_val))
        )
    table.setStyle(TableStyle(style_commands))
    story.append(table)
    story.extend(_entry_type_table_story(export_details, styles))
    story.extend(_detail_sections_story(export_details, styles))
    story.extend(_grouped_ledger_story(export_details, styles))
    story.extend(_item_tables_story(buckets, styles))

    doc.build(story, onFirstPage=on_first, onLaterPages=on_later)
    return response


def _apply_header_style(cell, header_font, header_fill):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_net_change_cell(cell, value: float):
    cell.number_format = "#,##0.00"
    cell.alignment = Alignment(horizontal="right")
    if value > 0:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="276221")
    elif value < 0:
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(color="9C0006")


def _write_entry_type_sheet(ws, export_details, header_font, header_fill, right_align):
    headers = [
        "Entry Type",
        "Category",
        "Transactions",
        _cur("Stock In"),
        _cur("Stock Out"),
        _cur("Net Value"),
        _cur("Sales Revenue"),
    ]
    for col_idx, header in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx, value=header), header_font, header_fill)

    rows = export_details.get("movement_by_entry_type") or []
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row.get("entry_type_label", ""))
        ws.cell(row=row_idx, column=2, value=row.get("category", ""))
        ws.cell(row=row_idx, column=3, value=row.get("transaction_count", 0))
        for col, key in (
            (4, "stock_in_value"),
            (5, "stock_out_value"),
            (6, "net_value"),
            (7, "sales_revenue"),
        ):
            val = float(row.get(key, 0))
            cell = ws.cell(row=row_idx, column=col)
            if val and key in ("stock_in_value", "stock_out_value", "sales_revenue"):
                cell.value = val
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            elif key == "net_value":
                cell.value = val
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            elif val:
                cell.value = val
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            else:
                cell.value = "—"

    widths = [28, 14, 14, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _safe_excel_sheet_name(name: str) -> str:
    invalid = set("[]:*?/\\")
    cleaned = "".join(c for c in name if c not in invalid).strip()
    return (cleaned or "Detail")[:31]


def _write_type_detail_sheet(ws, section, header_font, header_fill, right_align):
    """Full line listing for one entry type (sales, purchase, adjustment, etc.)."""
    entry_type = section.get("entry_type", "")
    is_sales = entry_type == "Sales"

    if is_sales:
        headers = [
            "Date",
            "Invoice No",
            "Item No",
            "Description",
            "Qty Sold",
            _cur("COGS"),
            _cur("Revenue"),
            _cur("Gross Margin"),
            _cur("Running Balance"),
            "Doc Type",
            "Notes",
        ]
    else:
        headers = [
            "Date",
            "Document No",
            "Item No",
            "Description",
            "Qty",
            "Direction",
            _cur("Cost"),
            _cur("Running Balance"),
            "Doc Type",
            "Notes",
        ]

    for col_idx, header in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx, value=header), header_font, header_fill)

    row_idx = 2
    for row in section.get("rows") or []:
        ws.cell(row=row_idx, column=1, value=row.get("posting_date", ""))
        ws.cell(row=row_idx, column=2, value=row.get("document_no", ""))
        ws.cell(row=row_idx, column=3, value=row.get("item_no", ""))
        ws.cell(row=row_idx, column=4, value=row.get("item_description", ""))
        ws.cell(row=row_idx, column=5, value=float(abs(row.get("quantity", 0)) if is_sales else row.get("quantity", 0)))
        if is_sales:
            cogs = float(row.get("cost_amount", 0))
            rev = float(row.get("sales_amount", 0))
            for col, val in (
                (6, cogs),
                (7, rev),
                (8, rev - cogs),
                (9, row.get("running_balance", 0)),
            ):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.number_format = "#,##0.00"
                c.alignment = right_align
            ws.cell(row=row_idx, column=10, value=row.get("document_type", "") or "")
            ws.cell(row=row_idx, column=11, value=(row.get("description") or "")[:120])
        else:
            c = ws.cell(row=row_idx, column=6, value=row.get("direction", ""))
            cost_cell = ws.cell(row=row_idx, column=7, value=float(row.get("cost_amount", 0)))
            cost_cell.number_format = "#,##0.00"
            cost_cell.alignment = right_align
            bal = ws.cell(row=row_idx, column=8, value=float(row.get("running_balance", 0)))
            bal.number_format = "#,##0.00"
            bal.alignment = right_align
            ws.cell(row=row_idx, column=9, value=row.get("document_type", "") or "")
            ws.cell(row=row_idx, column=10, value=(row.get("description") or "")[:120])
        row_idx += 1

    if row_idx > 2:
        ws.cell(row=row_idx, column=1, value="TOTAL")
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        if is_sales:
            for col, val in (
                (6, section.get("stock_out_total", 0)),
                (7, section.get("sales_revenue", 0)),
                (8, round(section.get("sales_revenue", 0) - section.get("stock_out_total", 0), 2)),
            ):
                c = ws.cell(row=row_idx, column=col, value=float(val))
                c.number_format = "#,##0.00"
                c.font = Font(bold=True)
                c.alignment = right_align
        else:
            c = ws.cell(row=row_idx, column=7, value=float(section.get("net_cost", 0)))
            c.number_format = "#,##0.00"
            c.font = Font(bold=True)
            c.alignment = right_align
    else:
        ws.cell(row=2, column=1, value="No transactions for this type in period")

    width_letters = (
        "ABCDEFGHIJK"
        if is_sales
        else "ABCDEFGHIJ"
    )
    widths = (
        [12, 16, 14, 36, 10, 14, 14, 14, 16, 14, 40]
        if is_sales
        else [12, 16, 14, 36, 10, 12, 14, 16, 14, 40]
    )
    for letter, width in zip(width_letters, widths):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"


def _write_sales_sheet(ws, export_details, header_font, header_fill, right_align):
    sales_section = next(
        (
            s
            for s in export_details.get("detail_sections") or []
            if s.get("entry_type") == "Sales"
        ),
        None,
    )
    if sales_section:
        _write_type_detail_sheet(ws, sales_section, header_font, header_fill, right_align)
        return

    headers = [
        "Date",
        "Invoice No",
        "Item No",
        "Description",
        "Qty Sold",
        _cur("COGS"),
        _cur("Revenue"),
        _cur("Gross Margin"),
    ]
    for col_idx, header in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx, value=header), header_font, header_fill)

    row_idx = 2
    for row in export_details.get("sales_detail") or []:
        ws.cell(row=row_idx, column=1, value=row.get("posting_date", ""))
        ws.cell(row=row_idx, column=2, value=row.get("document_no", ""))
        ws.cell(row=row_idx, column=3, value=row.get("item_no", ""))
        ws.cell(row=row_idx, column=4, value=row.get("item_description", ""))
        ws.cell(row=row_idx, column=5, value=float(row.get("quantity", 0)))
        for col, key in ((6, "cogs"), (7, "revenue"), (8, "gross_margin")):
            c = ws.cell(row=row_idx, column=col, value=float(row.get(key, 0)))
            c.number_format = "#,##0.00"
            c.alignment = right_align
        row_idx += 1

    totals = export_details.get("sales_totals") or {}
    if row_idx > 2:
        ws.cell(row=row_idx, column=1, value="TOTAL")
        for col, key in ((6, "cogs"), (7, "revenue"), (8, "gross_margin")):
            c = ws.cell(row=row_idx, column=col, value=float(totals.get(key, 0)))
            c.number_format = "#,##0.00"
            c.font = Font(bold=True)
            c.alignment = right_align
    else:
        ws.cell(row=2, column=1, value="No sales in selected period")

    for letter, width in zip("ABCDEFGH", [12, 16, 14, 36, 10, 16, 16, 16]):
        ws.column_dimensions[letter].width = width


class _InventoryExcelLayout:
    """Single-sheet layout mirroring PDF section order."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="1F3FFF")
        self.right_align = Alignment(horizontal="right", vertical="center")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.title_font = Font(bold=True, size=18, color="FFFFFF")
        self.section_font = Font(bold=True, size=12, color="1F3FFF")
        self.subsection_font = Font(bold=True, size=11, color="374151")
        self.subtle_fill = PatternFill("solid", fgColor="E8ECF8")
        self.totals_fill = PatternFill("solid", fgColor="EEF2FF")
        self.sales_highlight = PatternFill("solid", fgColor="FFF3CD")
        self.title_fill = PatternFill("solid", fgColor="1F3FFF")
        self.subtitle_fill = PatternFill("solid", fgColor="EEF2FF")
        self.section_fill = PatternFill("solid", fgColor="EEF2FF")
        _thin = Side(style="thin", color="D1D5DB")
        _medium = Side(style="medium", color="9CA3AF")
        self.inner_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        self.outer_h_border = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)

    def skip(self, lines: int = 1) -> None:
        self.row += lines

    def write_line(self, text: str, *, font: Font | None = None, merge_to_col: int = 8) -> None:
        cell = self.ws.cell(row=self.row, column=1, value=text)
        if font:
            cell.font = font
        if merge_to_col > 1:
            self.ws.merge_cells(
                start_row=self.row,
                start_column=1,
                end_row=self.row,
                end_column=merge_to_col,
            )
        self.row += 1

    def write_section(self, title: str, *, subsection: bool = False) -> None:
        cell = self.ws.cell(row=self.row, column=1, value=title)
        cell.font = self.subsection_font if subsection else self.section_font
        if not subsection:
            cell.fill = self.section_fill
            self.ws.row_dimensions[self.row].height = 20
        self.ws.merge_cells(
            start_row=self.row, start_column=1, end_row=self.row, end_column=8
        )
        self.row += 1

    def write_table_headers(self, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            _apply_header_style(
                self.ws.cell(row=self.row, column=col_idx, value=header),
                self.header_font,
                self.header_fill,
            )
        self.row += 1

    def _write_amount(self, row: int, col: int, value: float, *, bold: bool = False) -> None:
        cell = self.ws.cell(row=row, column=col, value=float(value))
        cell.number_format = "#,##0.00"
        cell.alignment = self.right_align
        if bold:
            cell.font = Font(bold=True)

    def _dash_or_amount(self, row: int, col: int, value: float) -> None:
        if value and abs(value) >= 0.005:
            self._write_amount(row, col, value)
        else:
            self.ws.cell(row=row, column=col, value="—")


def _excel_write_report_header(
    layout: _InventoryExcelLayout,
    period: dict,
    branch_label: str,
    summary: dict,
) -> None:
    start = period.get("start_date", "")
    end = period.get("end_date", "")
    subtitle = f"Period: {start} to {end} ({period.get('period_type', '')})"

    # Title banner row
    title_cell = layout.ws.cell(row=layout.row, column=1, value="Inventory Value Movement Report")
    title_cell.font = layout.title_font
    title_cell.fill = layout.title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    layout.ws.merge_cells(start_row=layout.row, start_column=1, end_row=layout.row, end_column=11)
    layout.ws.row_dimensions[layout.row].height = 38
    layout.row += 1

    # Subtitle / meta row
    meta_parts = [subtitle]
    if branch_label:
        meta_parts.append(f"Branch: {branch_label}")
    meta_parts.append(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    sub_cell = layout.ws.cell(row=layout.row, column=1, value="  •  ".join(meta_parts))
    sub_cell.font = Font(size=10, color="374151")
    sub_cell.fill = layout.subtitle_fill
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    layout.ws.merge_cells(start_row=layout.row, start_column=1, end_row=layout.row, end_column=11)
    layout.ws.row_dimensions[layout.row].height = 22
    layout.row += 1
    layout.skip()

    # KPI cards — 5 coloured columns (A–E)
    kpi_items = [
        ("Opening Value", "opening_value", "6B7280"),
        ("Stock In  (+)", "inbound_value", "059669"),
        ("Stock Out  (−)", "outbound_value", "DC2626"),
        ("Net Change", "net_change", "3B82F6"),
        ("Closing Value", "closing_value", "7C3AED"),
    ]
    for col_idx, (label, _, color) in enumerate(kpi_items, start=1):
        cell = layout.ws.cell(row=layout.row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = layout.center_align
    layout.ws.row_dimensions[layout.row].height = 18
    layout.row += 1

    for col_idx, (_, key, color) in enumerate(kpi_items, start=1):
        val = float(summary.get(key, 0))
        cell = layout.ws.cell(row=layout.row, column=col_idx, value=val)
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True, size=11)
        cell.alignment = layout.center_align
        cell.fill = PatternFill("solid", fgColor="F9FAFB")
    layout.ws.row_dimensions[layout.row].height = 24
    layout.row += 1
    layout.skip()

    # Formula summary row
    opening = summary.get("opening_value", 0)
    inbound = summary.get("inbound_value", 0)
    outbound = summary.get("outbound_value", 0)
    net = summary.get("net_change", 0)
    closing = summary.get("closing_value", 0)
    formula_text = (
        f"Opening ({_format_ugx(opening)})  +  Stock In ({_format_ugx(inbound)})  "
        f"−  Stock Out ({_format_ugx(outbound)})  =  Net Change ({_format_ugx(net)})  →  "
        f"Closing ({_format_ugx(closing)})"
    )
    formula_cell = layout.ws.cell(row=layout.row, column=1, value=formula_text)
    formula_cell.font = Font(size=9, italic=True, color="374151")
    formula_cell.fill = PatternFill("solid", fgColor="EEF2FF")
    formula_cell.alignment = Alignment(horizontal="center", vertical="center")
    layout.ws.merge_cells(start_row=layout.row, start_column=1, end_row=layout.row, end_column=11)
    layout.ws.row_dimensions[layout.row].height = 18
    layout.row += 1
    layout.skip()


def _excel_write_value_entry_reconciliation(layout: _InventoryExcelLayout, summary: dict) -> None:
    ve = summary.get("value_entry_reconciliation") or {}
    if not ve.get("available"):
        return

    account = (
        f"{summary.get('gl_account_no', '2110')} "
        f"{summary.get('gl_account_name', 'Resale Items')}"
    ).strip()
    layout.write_section("ValueEntry comparison", subsection=True)
    layout.write_line(
        f"Headline amounts are from G/L account {account} (same as Sales dashboard).",
    )
    layout.write_table_headers(
        ["", "G/L (report)", "ValueEntry", "Variance (G/L − VE)"]
    )
    def _write_recon_row(label: str, gl_key: str, ve_key: str, var_key: str, *, bold: bool = False) -> None:
        r = layout.row
        c1 = layout.ws.cell(row=r, column=1, value=label)
        if bold:
            c1.font = Font(bold=True)
        layout._write_amount(r, 2, summary.get(gl_key, 0))
        if bold:
            layout.ws.cell(row=r, column=2).font = Font(bold=True)
        layout._write_amount(r, 3, ve.get(ve_key, 0))
        if bold:
            layout.ws.cell(row=r, column=3).font = Font(bold=True)
        layout._write_amount(r, 4, ve.get(var_key, 0))
        if bold:
            layout.ws.cell(row=r, column=4).font = Font(bold=True)
        layout.row += 1

    def _write_breakdown_rows(breakdown: list) -> None:
        subtle = PatternFill("solid", fgColor="F8FAFC")
        for line in breakdown or []:
            r = layout.row
            layout.ws.cell(
                row=r,
                column=1,
                value=f"  {line.get('label', '')} ({line.get('transaction_count', 0)})",
            )
            layout.ws.cell(row=r, column=2, value="—")
            layout._write_amount(r, 3, line.get("amount", 0))
            layout.ws.cell(row=r, column=4, value="—")
            for col in range(1, 5):
                layout.ws.cell(row=r, column=col).fill = subtle
            layout.row += 1

    _write_recon_row("Opening", "opening_value", "opening_balance", "opening_variance")
    _write_recon_row("Closing", "closing_value", "closing_balance", "closing_variance")
    _write_recon_row(
        "Stock in (period)",
        "inbound_value",
        "inbound_balance",
        "period_inbound_variance",
        bold=True,
    )
    _write_breakdown_rows(ve.get("stock_in_breakdown"))
    _write_recon_row(
        "Stock out (period)",
        "outbound_value",
        "outbound_balance",
        "period_outbound_variance",
        bold=True,
    )
    _write_breakdown_rows(ve.get("stock_out_breakdown"))
    layout.skip(2)


def _excel_write_movement_table(
    layout: _InventoryExcelLayout, buckets: list, summary: dict
) -> None:
    layout.write_section("Inventory value movement")
    headers = [
        "Period",
        "Opening Value",
        "Stock In (+)",
        "Stock Out (−)",
        "Net Change",
        "Closing Value",
    ]
    layout.write_table_headers(headers)

    row_nets = []
    for bucket in buckets:
        net = float(bucket.get("net_change", 0))
        row_nets.append(net)
        r = layout.row
        layout.ws.cell(row=r, column=1, value=bucket.get("period", ""))
        layout._write_amount(r, 2, bucket.get("opening_value", 0))
        layout._write_amount(r, 3, bucket.get("inbound_value", 0))
        layout._write_amount(r, 4, bucket.get("outbound_value", 0))
        net_cell = layout.ws.cell(row=r, column=5, value=net)
        _style_net_change_cell(net_cell, net)
        layout._write_amount(r, 6, bucket.get("closing_value", 0))
        layout.row += 1

    if not buckets:
        r = layout.row
        layout.ws.cell(row=r, column=1, value="No movement data")
        layout.row += 1
    else:
        r = layout.row
        layout.ws.cell(row=r, column=1, value="TOTAL")
        layout._write_amount(r, 2, summary.get("opening_value", 0), bold=True)
        layout._write_amount(r, 3, summary.get("inbound_value", 0), bold=True)
        layout._write_amount(r, 4, summary.get("outbound_value", 0), bold=True)
        net_val = float(summary.get("net_change", 0))
        net_cell = layout.ws.cell(row=r, column=5, value=net_val)
        _style_net_change_cell(net_cell, net_val)
        if net_val > 0:
            net_cell.font = Font(bold=True, color="276221")
        elif net_val < 0:
            net_cell.font = Font(bold=True, color="9C0006")
        else:
            net_cell.font = Font(bold=True)
        layout._write_amount(r, 6, summary.get("closing_value", 0), bold=True)
        for col in range(1, 7):
            layout.ws.cell(row=r, column=col).fill = layout.totals_fill
        layout.row += 1
    layout.skip(2)


def _excel_write_entry_type_table(layout: _InventoryExcelLayout, export_details: dict) -> None:
    rows = export_details.get("movement_by_entry_type") or []
    if not rows:
        return

    layout.write_section("Movement by transaction type")
    layout.write_table_headers(
        [
            "Entry Type",
            "Category",
            "Transactions",
            "Stock In",
            "Stock Out",
            "Net Value",
            "Sales Revenue",
        ]
    )
    for row in rows:
        r = layout.row
        is_sales = row.get("entry_type") == "Sales"
        layout.ws.cell(row=r, column=1, value=row.get("entry_type_label", ""))
        layout.ws.cell(row=r, column=2, value=row.get("category", ""))
        layout.ws.cell(row=r, column=3, value=row.get("transaction_count", 0))
        layout._dash_or_amount(r, 4, float(row.get("stock_in_value", 0)))
        layout._dash_or_amount(r, 5, abs(float(row.get("stock_out_value", 0))))
        layout._write_amount(r, 6, float(row.get("net_value", 0)))
        layout._dash_or_amount(r, 7, float(row.get("sales_revenue", 0)))
        if is_sales:
            for col in range(1, 8):
                layout.ws.cell(row=r, column=col).fill = layout.sales_highlight
                layout.ws.cell(row=r, column=col).font = Font(bold=True)
        layout.row += 1
    layout.skip(2)


def _excel_append_detail_section(layout: _InventoryExcelLayout, section: dict) -> None:
    entry_type = section.get("entry_type", "")
    is_sales = entry_type == "Sales"
    title = section.get("title", entry_type)
    count = section.get("transaction_count", 0)

    layout.write_section(f"{title} — {count} line(s)", subsection=True)
    summary_parts = [
        f"Category: {section.get('category', '')}",
        f"Stock In: {_format_ugx(section.get('stock_in_total', 0))}",
        f"Stock Out: {_format_ugx(section.get('stock_out_total', 0))}",
        f"Net: {_format_ugx(section.get('net_cost', 0))}",
    ]
    if is_sales:
        summary_parts.append(
            f"Total revenue: {_format_ugx(section.get('sales_revenue', 0))}"
        )
    layout.write_line(" | ".join(summary_parts))

    if is_sales:
        headers = [
            "Date",
            "Invoice",
            "Item",
            "Description",
            "Qty",
            "COGS",
            "Revenue",
            "Margin",
            "Balance",
            "Doc Type",
            "Notes",
        ]
    else:
        headers = [
            "Date",
            "Document",
            "Item",
            "Description",
            "Qty",
            "Direction",
            "Cost",
            "Balance",
            "Doc Type",
            "Entry",
            "Notes",
        ]
    layout.write_table_headers(headers)
    if is_sales:
        for col in range(1, len(headers) + 1):
            layout.ws.cell(row=layout.row - 1, column=col).fill = layout.sales_highlight

    data_start = layout.row
    for row in section.get("rows") or []:
        r = layout.row
        layout.ws.cell(row=r, column=1, value=row.get("posting_date", ""))
        layout.ws.cell(row=r, column=2, value=row.get("document_no", ""))
        layout.ws.cell(row=r, column=3, value=row.get("item_no", ""))
        layout.ws.cell(row=r, column=4, value=(row.get("item_description") or "")[:40])
        if is_sales:
            layout.ws.cell(row=r, column=5, value=float(abs(row.get("quantity", 0))))
            cogs = float(row.get("cost_amount", 0))
            rev = float(row.get("sales_amount", 0))
            layout._write_amount(r, 6, cogs)
            layout._write_amount(r, 7, rev)
            layout._write_amount(r, 8, rev - cogs)
            layout._write_amount(r, 9, row.get("running_balance", 0))
            layout.ws.cell(row=r, column=10, value=row.get("document_type", "") or "—")
            layout.ws.cell(row=r, column=11, value=(row.get("description") or "")[:80])
        else:
            layout.ws.cell(row=r, column=5, value=float(row.get("quantity", 0)))
            layout.ws.cell(row=r, column=6, value=row.get("direction", ""))
            layout._write_amount(r, 7, row.get("cost_amount", 0))
            layout._write_amount(r, 8, row.get("running_balance", 0))
            layout.ws.cell(row=r, column=9, value=row.get("document_type", "") or "—")
            layout.ws.cell(row=r, column=10, value=row.get("entry_type_label", ""))
            layout.ws.cell(row=r, column=11, value=(row.get("description") or "")[:80])
        layout.row += 1

    if layout.row == data_start:
        layout.ws.cell(row=layout.row, column=1, value="No lines")
        layout.row += 1
    else:
        r = layout.row
        layout.ws.cell(row=r, column=1, value="TOTAL")
        if is_sales:
            layout._write_amount(r, 6, section.get("stock_out_total", 0), bold=True)
            layout._write_amount(r, 7, section.get("sales_revenue", 0), bold=True)
            layout._write_amount(
                r,
                8,
                round(
                    section.get("sales_revenue", 0) - section.get("stock_out_total", 0),
                    2,
                ),
                bold=True,
            )
        else:
            layout._write_amount(r, 7, section.get("net_cost", 0), bold=True)
        for col in range(1, len(headers) + 1):
            cell = layout.ws.cell(row=r, column=col)
            cell.fill = layout.totals_fill
            cell.font = Font(bold=True)
        layout.row += 1
    layout.skip(2)


def _excel_write_detail_sections(layout: _InventoryExcelLayout, export_details: dict) -> None:
    sections = export_details.get("detail_sections") or []
    if not sections:
        return

    layout.write_section("Transaction detail by type")
    layout.write_line(
        "Every Value Entry in the period is listed below, grouped by transaction type. "
        "Sales show COGS and revenue; purchases and adjustments show how stock value changed.",
    )
    layout.skip()
    for section in sections:
        _excel_append_detail_section(layout, section)


def _excel_write_grouped_ledger(layout: _InventoryExcelLayout, export_details: dict) -> None:
    groups = export_details.get("ledger_groups") or []
    layout.write_section("Transaction detail (all Value Entries)")
    if not groups:
        layout.write_line("No transactions in selected period.")
        layout.skip(2)
        return

    headers = [
        "Date",
        "Document",
        "Entry Type",
        "Item",
        "Qty",
        "Direction",
        _cur("Cost"),
        _cur("Revenue"),
        _cur("Balance"),
        "Description",
    ]
    for group in groups:
        sep_text = (
            f"─── {group.get('period_label', group.get('period', ''))} ─── "
            f"Opening: {_format_ugx(group.get('opening_value', 0))} │ "
            f"In: +{_format_ugx(group.get('stock_in', 0))} │ "
            f"Out: −{_format_ugx(group.get('stock_out', 0))} │ "
            f"Closing: {_format_ugx(group.get('closing_value', 0))} ───"
        )
        header_cell = layout.ws.cell(row=layout.row, column=1, value=sep_text)
        header_cell.font = Font(bold=True)
        header_cell.fill = layout.subtle_fill
        layout.ws.merge_cells(
            start_row=layout.row,
            start_column=1,
            end_row=layout.row,
            end_column=len(headers),
        )
        layout.row += 1

        layout.write_table_headers(headers)
        data_start = layout.row
        for txn in group.get("transactions") or []:
            r = layout.row
            item_label = txn.get("item_no", "")
            if txn.get("item_description"):
                item_label = f"{item_label} {(txn.get('item_description') or '')[:24]}"
            layout.ws.cell(row=r, column=1, value=txn.get("posting_date", ""))
            layout.ws.cell(row=r, column=2, value=txn.get("document_no", ""))
            layout.ws.cell(row=r, column=3, value=txn.get("entry_type_label", ""))
            layout.ws.cell(row=r, column=4, value=item_label)
            layout.ws.cell(row=r, column=5, value=float(txn.get("quantity", 0)))
            layout.ws.cell(row=r, column=6, value=txn.get("direction", ""))
            layout._write_amount(r, 7, txn.get("cost_amount", 0))
            rev = float(txn.get("revenue", 0))
            if rev:
                layout._write_amount(r, 8, rev)
            else:
                layout.ws.cell(row=r, column=8, value="—")
            layout._write_amount(r, 9, txn.get("running_balance", 0))
            layout.ws.cell(row=r, column=10, value=(txn.get("description") or "")[:80])
            layout.row += 1

        if layout.row == data_start:
            layout.ws.cell(row=layout.row, column=1, value="No transactions")
            layout.row += 1
        layout.skip()

    layout.skip()


def _excel_write_item_breakdown(layout: _InventoryExcelLayout, buckets: list) -> None:
    has_items = any(bucket.get("items") for bucket in buckets)
    if not has_items:
        return

    for bucket in buckets:
        items = bucket.get("items") or []
        if not items:
            continue
        period = bucket.get("period", "")
        layout.write_section(f"Items moved — {period}", subsection=True)
        layout.write_table_headers(
            ["Item No", "Description", "Stock In", "Stock Out", "Net Qty"]
        )
        for item in items:
            r = layout.row
            layout.ws.cell(row=r, column=1, value=item.get("item_no", ""))
            layout.ws.cell(row=r, column=2, value=(item.get("item_description") or "")[:40])
            layout._write_amount(r, 3, item.get("inbound_value", 0))
            layout._write_amount(r, 4, item.get("outbound_value", 0))
            layout.ws.cell(row=r, column=5, value=float(item.get("net_qty", 0)))
            layout.row += 1
        layout.skip()


def _write_transactions_sheet(ws, export_details, header_font, header_fill, right_align):
    headers = [
        "Date",
        "Document",
        "Entry Type",
        "Item No",
        "Description",
        "Qty",
        "Direction",
        _cur("Cost"),
        _cur("Revenue"),
        _cur("Balance"),
    ]
    for col_idx, header in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=1, column=col_idx, value=header), header_font, header_fill)

    row_idx = 2
    subtotal_fill = PatternFill("solid", fgColor="E8ECF8")
    bold = Font(bold=True)

    for group in export_details.get("ledger_groups") or []:
        sub_row = row_idx
        period_label = group.get("period_label", group.get("period", ""))
        header_text = (
            f"── {period_label} ──  "
            f"Opening {_format_ugx(group.get('opening_value', 0))} | "
            f"Closing {_format_ugx(group.get('closing_value', 0))}"
        )
        header_cell = ws.cell(row=sub_row, column=1, value=header_text)
        ws.merge_cells(start_row=sub_row, start_column=1, end_row=sub_row, end_column=10)
        header_cell.font = bold
        header_cell.fill = subtotal_fill
        row_idx += 1

        group_start = row_idx
        for txn in group.get("transactions") or []:
            ws.cell(row=row_idx, column=1, value=txn.get("posting_date", ""))
            ws.cell(row=row_idx, column=2, value=txn.get("document_no", ""))
            ws.cell(row=row_idx, column=3, value=txn.get("entry_type_label", ""))
            ws.cell(row=row_idx, column=4, value=txn.get("item_no", ""))
            ws.cell(row=row_idx, column=5, value=txn.get("item_description", ""))
            ws.cell(row=row_idx, column=6, value=float(txn.get("quantity", 0)))
            ws.cell(row=row_idx, column=7, value=txn.get("direction", ""))
            for col, key in ((8, "cost_amount"), (10, "running_balance")):
                c = ws.cell(row=row_idx, column=col, value=float(txn.get(key, 0)))
                c.number_format = "#,##0.00"
                c.alignment = right_align
            rev = float(txn.get("revenue", 0))
            rev_cell = ws.cell(row=row_idx, column=9, value=rev if rev else "—")
            if rev:
                rev_cell.number_format = "#,##0.00"
                rev_cell.alignment = right_align
            row_idx += 1

        for r in range(group_start, row_idx):
            ws.row_dimensions[r].outline_level = 1

        summary_row = row_idx
        ws.cell(row=summary_row, column=1, value="Period subtotal")
        ws.cell(row=summary_row, column=3, value=group.get("period_label", ""))
        ws.cell(
            row=summary_row,
            column=8,
            value=float(group.get("stock_in", 0)),
        ).number_format = "#,##0.00"
        ws.cell(
            row=summary_row,
            column=9,
            value=float(group.get("stock_out", 0)),
        ).number_format = "#,##0.00"
        ws.cell(
            row=summary_row,
            column=10,
            value=float(group.get("closing_value", 0)),
        ).number_format = "#,##0.00"
        for col in range(1, 11):
            ws.cell(row=summary_row, column=col).font = bold
        row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No transactions in selected period")

    ws.auto_filter.ref = f"A1:J{max(row_idx - 1, 1)}"
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = True

    for letter, width in zip("ABCDEFGHIJ", [12, 16, 22, 14, 32, 8, 12, 14, 14, 16]):
        ws.column_dimensions[letter].width = width


def build_inventory_value_movement_excel(
    report_payload: dict, branch_label: str = ""
) -> HttpResponse:
    """Single worksheet stacked in the same order as the PDF export."""
    summary = _summary(report_payload)
    buckets = _movement_rows(report_payload)
    period = _period_meta(report_payload)
    export_details = _export_details(report_payload)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_properties.tabColor = "1F3FFF"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.sheet_view.showGridLines = True
    layout = _InventoryExcelLayout(ws)

    _excel_write_report_header(layout, period, branch_label, summary)
    _excel_write_value_entry_reconciliation(layout, summary)

    txn_count = export_details.get("transaction_count", 0)
    if txn_count:
        layout.write_line(
            f"This export documents {txn_count} inventory value entries, "
            "including every sale, purchase, and adjustment in the period.",
        )
        layout.skip()

    _excel_write_movement_table(layout, buckets, summary)
    _excel_write_entry_type_table(layout, export_details)
    _excel_write_detail_sections(layout, export_details)
    _excel_write_grouped_ledger(layout, export_details)
    _excel_write_item_breakdown(layout, buckets)

    ws.freeze_panes = "A3"
    column_widths = {
        "A": 14,
        "B": 18,
        "C": 16,
        "D": 32,
        "E": 10,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 40,
    }
    for letter, width in column_widths.items():
        ws.column_dimensions[letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    start = period.get("start_date", "")
    end = period.get("end_date", "")
    filename = f"inventory-value-movement-{start}-{end}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
