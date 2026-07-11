import json
import xlsxwriter
import uuid
import os
import pandas as pd
import time

from helpers.helpers import to_camel_case
from django.shortcuts import render, get_object_or_404
from django.db import models
from django.apps import apps
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.core.cache import cache
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

from io import BytesIO
from django.db import connection
from datetime import datetime
from dateutil import parser
from openpyxl import Workbook

from openpyxl import load_workbook


from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings

from .models import UploadTemplates, ConfigPackage, ConfigPackageTable
from items.tasks import process_item_import, process_journal_import

from utils.utils import ExcelProcessor

from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass

from items.models import Item, UnitOfMeasure, ItemCategory
from postings.models import GeneralProductPostingGroup, InventoryPostingGroup

from celery.result import AsyncResult
from celery.states import PENDING, SUCCESS, FAILURE, STARTED
from celery import chain, group
from threading import Thread

from rest_framework import viewsets, status
from rest_framework.authentication import SessionAuthentication
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from authentication.authentication import JWTAuthenticationWithRevocationChecks as JWTAuthentication
from .serializers import (
    ConfigPackageSerializer,
    ConfigPackageTableSerializer,
    ConfigPackageDetailSerializer,
)
from base.models import Objects
from rest_framework.exceptions import ValidationError
from django.db import transaction

import logging
from django_tenants.utils import tenant_context

from .import_handlers import process_import_data
from .tasks import import_tables_background

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    valid: bool
    message: str
    details: List[str] = None
    imported_count: int = 0


class DataImporter:
    EXCLUDED_FIELDS = {
        "id",
        "created_at",
        "updated_at",
        "system_id",
        "tree_id",
        "lft",
        "rght",
        "level",  # MPTT fields
        "costing_method",
        "general_product_posting_group",
        "inventory_posting_group",
    }

    def __init__(self, file, table_id: str, table_name: str, package_code: str):
        self.file = file
        self.table_id = table_id
        self.table_name = table_name
        self.package_code = package_code
        self.model = None
        self.field_mapping = {}
        self.errors = []

    def get_model(self) -> bool:
        """Get the Django model for the given table_id."""
        for app_config in apps.get_app_configs():
            try:
                self.model = apps.get_model(app_config.label, self.table_id)
                return True
            except LookupError:
                continue
        return False

    def get_model_fields(self) -> Dict[str, Any]:
        """Get available fields for the model."""
        model_fields = {}
        for field in self.model._meta.get_fields():
            if (
                not field.is_relation or field.many_to_one
            ) and field.name not in self.EXCLUDED_FIELDS:
                field_name = field.name
                verbose_name = (
                    field.verbose_name.title()
                    if hasattr(field, "verbose_name")
                    else field.name
                )
                self.field_mapping[verbose_name] = field_name
                model_fields[field_name] = field
        return model_fields

    def read_excel(self) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
        """Read and validate Excel file."""
        try:
            df = pd.read_excel(self.file)
            if df.empty:
                return None, "File is empty"

            # Get headers from row 4 and data starting from row 5
            headers = df.iloc[3].fillna("").str.strip()
            data_df = df.iloc[4:]
            data_df.columns = headers

            return data_df, None
        except Exception as e:
            return None, f"Error reading Excel file: {str(e)}"

    def process_record(self, row: dict, model) -> Dict:
        """Convert row data to model fields based on header names"""
        record = {}
        model_fields = {field.name: field for field in model._meta.fields}

        # Map the actual Excel headers to model fields
        header_mapping = {
            "Item Name": "item_name",
            "Bar Code No": "bar_code_no",
            "Type": "type",
            "Costing Method": "costing_method",
            "Unit of Measure": "unit_of_measure",
            "Item Category": "item_category",
            "General Product Posting Group": "general_product_posting_group",
            "Inventory Posting Group": "inventory_posting_group",
            # "Shelf No": "shelf_no",
            "Unit Price": "unit_price",
            "Description": "description",
            "Blocked": "blocked",
        }

        # First create the base record with mapped fields
        for header, value in row.items():
            if header in header_mapping:
                field_name = header_mapping[header]
                record[field_name] = value

        # Special handling for Items model
        if model._meta.model_name == "item":
            from items.models import UnitOfMeasure, ItemCategory
            from postings.models import (
                InventoryPostingSetup,
                GeneralProductPostingGroup,
            )

            # Handle Unit of Measure
            uom_code = record.get("unit_of_measure", "PCS")
            try:
                uom, _ = UnitOfMeasure.objects.get_or_create(
                    code=uom_code,
                    defaults={"description": f"Auto-created for {uom_code}"},
                )
                record["unit_of_measure"] = uom
            except Exception as e:
                print(f"Error creating UnitOfMeasure: {e}")
                record["unit_of_measure"] = None

            # Handle Item Category
            category_code = record.get("item_category", "UNCATEGORIZED")
            try:
                category, _ = ItemCategory.objects.get_or_create(
                    code=category_code,
                    defaults={"description": f"Auto-created for {category_code}"},
                )
                record["item_category"] = category
            except Exception as e:
                print(f"Error creating ItemCategory: {e}")
                record["item_category"] = None

            # Handle General Product Posting Group
            gpg_code = record.get("general_product_posting_group", "RETAIL")
            try:
                gpg, _ = GeneralProductPostingGroup.objects.get_or_create(
                    code=gpg_code,
                    defaults={"description": f"Auto-created for {gpg_code}"},
                )
                record["general_product_posting_group"] = gpg
            except Exception as e:
                print(f"Error creating GeneralProductPostingGroup: {e}")
                record["general_product_posting_group"] = None

            # Handle Inventory Posting Group
            ipg_code = record.get("inventory_posting_group", "RETAIL")
            try:
                ipg, _ = InventoryPostingSetup.objects.get_or_create(
                    code=ipg_code,
                    defaults={
                        "description": f"Auto-created for {ipg_code}",
                        "inventory_account": "1234",
                        "inventory_account_interim": "1234",
                    },
                )
                record["inventory_posting_group"] = ipg
            except Exception as e:
                print(f"Error creating InventoryPostingSetup: {e}")
                record["inventory_posting_group"] = None

            # Set default values for required fields
            record["type"] = "Inventory"
            record["costing_method"] = "FIFO"
            record["blocked"] = record.get("blocked", False)
            record["unit_price"] = (
                float(record.get("unit_price", 0)) if record.get("unit_price") else 0
            )

        # Final pass: convert values based on field types
        final_record = {}
        for field_name, value in record.items():
            if field_name in model_fields:
                field = model_fields[field_name]
                if value is not None:
                    if isinstance(
                        field,
                        (models.IntegerField, models.FloatField, models.DecimalField),
                    ):
                        try:
                            final_record[field_name] = (
                                float(value) if value not in (None, "") else 0
                            )
                        except (ValueError, TypeError):
                            final_record[field_name] = 0
                    else:
                        final_record[field_name] = value

        print("Final processed record:", final_record)
        return final_record

    def import_data(self, data_df: pd.DataFrame) -> ImportResult:
        """Import data from DataFrame to Django model."""
        imported_count = 0
        records = data_df.to_dict("records")

        for index, record in enumerate(records, start=1):
            try:
                data = self.process_record(record, self.model)
                if data:
                    self.model.objects.create(**data)
                    imported_count += 1
            except Exception as e:
                self.errors.append(f"Row {index}: {str(e)}")

        if self.errors:
            return ImportResult(
                valid=False, message="Import errors occurred", details=self.errors
            )

        return ImportResult(
            valid=True,
            message=f"Successfully imported {imported_count} records",
            imported_count=imported_count,
        )


class UploadDataView(View):
    template_name = "config_packages/upload-data.html"

    def get(self, request, *args, **kwargs):
        context = {"upload_templates": UploadTemplates.objects.all()}
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        print(file)
        process_id = str(uuid.uuid4())

        cache.set(
            f"upload_status_{process_id}",
            {
                "status": "processing",
                "progress": 0,
                "message": "Starting file processing...",
            },
            timeout=3600,
        )  # The cache timeout is set to 3600 seconds (1 hour).

        ExcelProcessor.initialize_background_process(file, process_id, request)

        return render(
            request, "partials/process-upload-results.html", {"process_id": process_id}
        )


def get_model_template(request):
    template_id = request.GET.get("template_id")
    if template_id:
        template = UploadTemplates.objects.get(name=template_id)
        print(template.template_file.url)
        return JsonResponse(
            {"template": template.template_file.url, "model_name": template.name}
        )
    return JsonResponse({"error": "Template not found"}, status=404)


def process_upload_status(request, process_id):
    status = cache.get(f"upload_status_{process_id}")
    return JsonResponse(status)


class ConfigPackageListView(View):
    template_name = "config_packages/config-packages-list.html"

    def get(self, request, *args, **kwargs):
        context = {"packages": ConfigPackage.objects.all()}
        return render(request, self.template_name, context)


@require_http_methods(["POST"])
def save_package_header(request):
    try:
        data = json.loads(request.body)
        code = data.get("code")
        package_name = data.get("package_name")
        current_code = data.get("current_code")

        if current_code and current_code != code:
            # Handle code change - update existing package
            package = ConfigPackage.objects.get(code=current_code)
            package.code = code
            package.package_name = package_name
            package.save()
        else:
            # Create new package
            package, created = ConfigPackage.objects.get_or_create(
                code=code, defaults={"package_name": package_name, "status": "DRAFT"}
            )
            if not created:
                package.package_name = package_name
                package.save()

        return JsonResponse({"code": package.code, "message": "Saved successfully"})
    except Exception as e:
        return JsonResponse({"message": str(e)}, status=400)


@require_http_methods(["POST"])
def save_package_table(request):
    try:
        data = json.loads(request.body)
        package_code = data.get("package_code")
        table_id = data.get("table_id")
        table_name = data.get("table_name")

        package = ConfigPackage.objects.get(code=package_code)
        table, created = ConfigPackageTable.objects.get_or_create(
            package_code=package, table_id=table_id, defaults={"table_name": table_name}
        )
        if not created:
            table.table_name = table_name
            table.save()

        return JsonResponse({"message": "Table saved successfully"})
    except Exception as e:
        return JsonResponse({"message": str(e)}, status=400)


def get_table_list(request):
    """
    Returns a list of available tables for the select2 dropdown
    """
    search_term = request.GET.get("search", "").lower()
    tables = []

    for model in apps.get_models():
        # Skip Django's built-in models
        if model._meta.app_label in [
            "admin",
            "auth",
            "contenttypes",
            "sessions",
            "messages",
        ]:
            continue

        # Skip abstract models
        if model._meta.abstract:
            continue

        model_name = model._meta.verbose_name.title()
        app_label = model._meta.app_label

        # Apply search filter
        if (
            search_term
            and search_term not in model_name.lower()
            and search_term not in app_label.lower()
        ):
            continue

        # Create table entry
        table = {
            "id": model._meta.model_name,
            "text": model_name,
            "name": model._meta.model_name,
            "app": app_label,
        }
        tables.append(table)

    # Add custom entries if needed
    if not search_term or "items" in search_term.lower():
        tables.append(
            {
                "id": "itemsonly",
                "text": "Items Only",
                "name": "Items_Only",
                "app": "custom",
            }
        )

    # Sort tables by display name
    tables.sort(key=lambda x: x["text"])

    return JsonResponse(
        {
            "results": tables,
            "pagination": {"more": False},  # Set to true if you implement pagination
        }
    )


def delete_package(request, package_code):
    package = ConfigPackage.objects.get(code=package_code)
    package.delete()
    return JsonResponse({"message": "Package deleted successfully"})


def get_package_details(request, package_code):
    try:
        package = ConfigPackage.objects.get(code=package_code)
        tables = ConfigPackageTable.objects.filter(package_code=package)

        return JsonResponse(
            {
                "code": package.code,
                "package_name": package.package_name,
                "tables": [
                    {
                        "id": table.table_id,
                        "text": table.table_name,  # This will show in the input
                        "name": table.table_name,  # This will be used for the readonly field
                    }
                    for table in tables
                ],
            }
        )
    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)


def export_package(request, package_code):
    try:
        package = ConfigPackage.objects.get(code=package_code)
        tables = ConfigPackageTable.objects.filter(package_code=package)

        # Create Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()

        # Add headers
        headers = ["Table ID", "Table Name"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # Add data
        for row, table in enumerate(tables, start=1):
            worksheet.write(row, 0, table.table_id)
            worksheet.write(row, 1, table.table_name)

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename=package-{package_code}.xlsx"
        )
        return response

    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)


@require_http_methods(["POST"])
def import_package(request):
    if "file" not in request.FILES:
        return JsonResponse({"message": "No file uploaded"}, status=400)

    # Add your Excel import logic here
    # Return success/error response
    return JsonResponse({"message": "Import successful"})


@require_http_methods(["POST"])
@csrf_exempt
def export_tables(request):
    try:
        data = json.loads(request.body)
        tables = data.get("tables", [])

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)

        # Add formats
        header_format = workbook.add_format(
            {"bold": True, "bg_color": "#f0f0f0", "border": 1}
        )
        title_format = workbook.add_format(
            {"bold": True, "font_size": 14, "align": "left"}
        )
        data_format = workbook.add_format({"border": 1})
        datetime_format = workbook.add_format(
            {"border": 1, "num_format": "yyyy-mm-dd hh:mm:ss"}
        )
        date_format = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd"})

        for table in tables:
            table_id = table["table_id"]
            table_name = table["table_name"]
            package_code = table["package_code"]  # Get package code from each table

            # Create worksheet for each table
            worksheet = workbook.add_worksheet(table_name[:31])

            try:
                # Write package code and table name at the top
                worksheet.write(0, 0, "Package Code:", header_format)
                worksheet.write(
                    0, 1, package_code
                )  # Use the package code from the table
                worksheet.write(1, 0, "Table Name:", header_format)
                worksheet.write(1, 1, table_name)

                # Add some spacing
                current_row = 3  # Start table data after the header

                # Get the model class
                model = None
                for app_config in apps.get_app_configs():
                    try:
                        model_name = to_camel_case(table_name)
                        model = apps.get_model(app_config.label, model_name)
                        break
                    except LookupError:
                        continue

                if not model:
                    raise Exception(f"Model not found for table: {table_name}")

                # Get field configuration from ConfigPackageTable
                try:
                    package = ConfigPackage.objects.get(code=package_code)
                    config_table = ConfigPackageTable.objects.get(
                        package_code=package, table_id=table_id
                    )

                    # Ensure field_config is populated
                    if (
                        not config_table.field_config
                        or "fields" not in config_table.field_config
                    ):
                        config_table.populate_field_config()
                        config_table.save()

                    # Get export fields from configuration
                    export_fields = config_table.get_export_fields()
                    print(export_fields)

                    # Get field details from configuration
                    field_config = config_table.field_config.get("fields", [])
                    field_map = {field["name"]: field for field in field_config}

                except (ConfigPackage.DoesNotExist, ConfigPackageTable.DoesNotExist):
                    # Fallback to old method if configuration not found
                    export_fields = None
                    field_map = {}

                # If no field configuration, use all fields (fallback)
                if not export_fields:
                    fields = model._meta.get_fields()
                    exclude_fields = {
                        "system_id",
                        "created_at",
                        "updated_at",
                        "id",
                        "password",
                    }
                    filtered_fields = [
                        field for field in fields if field.name not in exclude_fields
                    ]

                    columns = []
                    for field in filtered_fields:
                        if not field.is_relation or field.many_to_one:
                            columns.append(
                                {
                                    "name": field.name,
                                    "verbose_name": (
                                        field.verbose_name.title()
                                        if hasattr(field, "verbose_name")
                                        else field.name
                                    ),
                                    "is_datetime": isinstance(
                                        field, models.DateTimeField
                                    ),
                                    "is_date": isinstance(field, models.DateField),
                                }
                            )
                else:
                    # Use configured fields
                    columns = []
                    for field_name in export_fields:
                        try:
                            field = model._meta.get_field(field_name)
                            if not field.is_relation or field.many_to_one:
                                columns.append(
                                    {
                                        "name": field.name,
                                        "verbose_name": (
                                            field.verbose_name.title()
                                            if hasattr(field, "verbose_name")
                                            else field.name
                                        ),
                                        "is_datetime": isinstance(
                                            field, models.DateTimeField
                                        ),
                                        "is_date": isinstance(field, models.DateField),
                                    }
                                )
                        except Exception:
                            # Skip fields that don't exist in the model
                            continue

                # Write headers
                for col_idx, column in enumerate(columns):
                    worksheet.write(
                        current_row, col_idx, column["verbose_name"], header_format
                    )
                    worksheet.set_column(col_idx, col_idx, 15)
                current_row += 1

                # Get data using Django ORM
                queryset = model.objects.all()

                # Write data
                for row_idx, obj in enumerate(queryset, start=current_row):
                    for col_idx, column in enumerate(columns):
                        value = getattr(obj, column["name"])

                        # Handle datetime and date fields
                        if column["is_datetime"] and value is not None:
                            if value.tzinfo is not None:
                                value = value.replace(tzinfo=None)
                            worksheet.write_datetime(
                                row_idx, col_idx, value, datetime_format
                            )
                        elif column["is_date"] and value is not None:
                            worksheet.write_datetime(
                                row_idx, col_idx, value, date_format
                            )
                        else:
                            if hasattr(value, "_meta"):
                                value = str(value)
                            worksheet.write(row_idx, col_idx, value, data_format)

            except Exception as e:
                worksheet.write(0, 0, f"Error exporting table: {str(e)}")

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=table_data_export.xlsx"
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@require_http_methods(["POST"])
def validate_import(request):
    try:
        if "file" not in request.FILES:
            return JsonResponse({"valid": False, "error": "No file uploaded"})

        file = request.FILES["file"]

        try:
            # First read the table name from row 2
            df_info = pd.read_excel(file, header=None)
            table_name = (
                str(df_info.iloc[1, 1]).strip().lower() if len(df_info) > 1 else ""
            )

            # Now read the actual data starting from the header row (row 4)
            file.seek(0)  # Reset file pointer
            df = pd.read_excel(file, header=3)  # This means the header is in row 4

            # Convert DataFrame to JSON string
            serialized_df = df.to_json(orient="records")

        except Exception as e:
            return JsonResponse(
                {"valid": False, "error": f"Error reading Excel file: {str(e)}"}
            )

        if table_name == "item":
            task = process_item_import.delay(serialized_df, request.tenant.schema_name)
        elif table_name == "itemjournal":
            task = process_journal_import.delay(
                serialized_df,
                request.tenant.schema_name,
                request.user.id,  # Pass the user ID
            )
        else:
            return JsonResponse(
                {"valid": False, "error": f"Unsupported table type: {table_name}"}
            )

        return JsonResponse(
            {
                "valid": True,
                "message": f"Import started for {table_name}",
                "task_id": task.id,
            }
        )

    except Exception as e:
        logger.error(f"Import validation error: {str(e)}", exc_info=True)
        return JsonResponse({"valid": False, "error": str(e)})


@require_http_methods(["POST"])
@csrf_exempt
def import_tables(request):
    try:
        # Get tenant from request
        tenant = request.tenant
        print(f"Tenant from request: {tenant}")
        print(f"Tenant schema_name: {tenant.schema_name if tenant else 'None'}")

        # Validate tenant exists
        if tenant is None:
            return JsonResponse(
                {"message": "Tenant not found in request", "success": False}, status=400
            )

        # Get the file and parameters from request
        excel_file = request.FILES.get("file")
        package_code = request.POST.get("package_code")
        table_id = request.POST.get("table_id")
        table_name = request.POST.get("table_name")

        if not all([excel_file, package_code, table_id, table_name]):
            return JsonResponse(
                {"message": "Missing required parameters", "success": False}, status=400
            )

        # Save the uploaded file temporarily
        import tempfile
        import os

        # Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            for chunk in excel_file.chunks():
                temp_file.write(chunk)
            temp_file_path = temp_file.name

        try:
            # Start the background task
            logger.info(
                f"DEBUG: View - About to call import_tables_background with user_id: {request.user.id}"
            )
            task = import_tables_background.delay(
                tenant_schema_name=tenant.schema_name,
                excel_file_path=temp_file_path,
                package_code=package_code,
                table_id=table_id,
                table_name=table_name,
                user_id=request.user.id,
            )

            return JsonResponse(
                {
                    "message": "Import started in background",
                    "success": True,
                    "task_id": task.id,
                    "status": "started",
                }
            )

        except Exception as e:
            # Clean up temp file if task creation fails
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            raise e

        # Note: The old synchronous processing logic below is kept for fallback
        # but the background task above should handle the import

        # Get model fields
        exclude_fields = {
            "id",
            "created_at",
            "updated_at",
            "system_id",
            "tree_id",
            "lft",
            "rght",
            "level",  # MPTT fields
            "password",
            "last_login",
            "start_date",
            "schema_name",  # Sensitive and tenant fields
        }

        model_fields = []
        for field in model._meta.get_fields():
            if (
                not field.is_relation or field.many_to_one
            ) and field.name not in exclude_fields:
                model_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": (
                            field.verbose_name.title()
                            if hasattr(field, "verbose_name")
                            else field.name
                        ),
                        "type": field.get_internal_type(),
                    }
                )

        # Read Excel file - get headers from row 4 and data starting from row 5
        df = pd.read_excel(excel_file, header=None)
        headers = df.iloc[3].fillna("").str.strip()
        data_df = df.iloc[4:]
        data_df.columns = headers

        # Validate columns match model fields
        excel_columns = set(headers)
        backend_default_fields = {
            "Costing Method",
            "Type",
            "Blocked",
            "Inventory Posting Group",
            "General Product Posting Group",
            "Tracking Code",
        }
        required_fields = {
            field["verbose_name"]
            for field in model_fields
            if field["verbose_name"] not in backend_default_fields
        }

        missing_fields = required_fields - excel_columns
        if missing_fields:
            return JsonResponse(
                {
                    "message": f"Missing required fields in Excel: {', '.join(missing_fields)}",
                    "success": False,
                },
                status=400,
            )

        # Initialize counters
        stats = {
            "created": 0,
            "updated": 0,
            "failed": 0,
            "total_rows": len(data_df),
            "errors": [],  # Add errors list to track detailed errors
        }

        # Get or create config package table - no need to specify using() anymore
        config_package = ConfigPackage.objects.get(code=package_code)
        package_table, _ = ConfigPackageTable.objects.get_or_create(
            package_code=config_package,
            table_id=table_id,
            defaults={"table_name": table_name},
        )

        field_mapping = {field["verbose_name"]: field["name"] for field in model_fields}

        # Define dependent fields that need special processing
        dependent_fields = [
            "sales_unit_of_measure",
            "purchase_unit_of_measure",
        ]

        # Process each row
        for row_index, row in data_df.iterrows():
            try:
                row_data = row.to_dict()
                row_data = {k: ("" if pd.isna(v) else v) for k, v in row_data.items()}

                mapped_data = {}
                for excel_col, value in row_data.items():
                    if excel_col in field_mapping:
                        # For Item model, handle 'no' field specially
                        if (
                            model._meta.model_name == "item"
                            and field_mapping[excel_col] == "no"
                        ):
                            # Only include 'no' field if it has a value, otherwise let save method generate it
                            if value != "" and value is not None:
                                mapped_data[field_mapping[excel_col]] = value
                            # If empty, don't include it in mapped_data - save method will generate it
                        else:
                            # Only add non-empty values to mapped_data for other fields
                            if value != "" and value is not None:
                                mapped_data[field_mapping[excel_col]] = value

                # Add tenant-specific fields if model supports it
                if hasattr(model, "schema_name"):
                    mapped_data["schema_name"] = tenant.schema_name

                # Use only 'no' as identifier for Item model, else fallback to previous logic
                if model._meta.model_name == "item":
                    identifier_fields = ["no"]
                else:
                    identifier_fields = ["id", "code", "no", "name", "email"]
                identifier_field = next(
                    (field for field in identifier_fields if field in mapped_data),
                    None,
                )

                # For Item model, if no identifier field found but we have item_name, treat as new item creation
                if (
                    not identifier_field
                    and model._meta.model_name == "item"
                    and "item_name" in mapped_data
                ):
                    # This is a new item creation - no item number provided, will be auto-generated
                    identifier_field = (
                        "item_name"  # Use item_name as identifier for new items
                    )
                    print(
                        f"Row {row_index + 5}: New item creation - item_name: {mapped_data['item_name']}"
                    )
                elif not identifier_field:
                    error_msg = f"Row {row_index + 5}: No identifier field found (expected one of: {', '.join(identifier_fields)}). Row data: {row_data}"
                    stats["errors"].append(error_msg)
                    stats["failed"] += 1
                    logger.warning(error_msg)
                    continue

                # Validate and process foreign key fields before saving
                validation_errors = []

                # Store dependent field values for processing after item creation
                dependent_field_values = {}

                # First pass: Process basic fields that don't depend on other fields
                fields_to_remove = []
                for field_name, value in mapped_data.items():
                    if (
                        value == ""
                        or value is None
                        or field_name in dependent_fields
                        or field_name == "id"  # Exclude id field from processing
                    ):
                        if field_name in dependent_fields and value:
                            # Store dependent field values for later processing
                            dependent_field_values[field_name] = value
                            # Mark for removal to prevent it from being passed to model creation
                            fields_to_remove.append(field_name)
                        continue

                    # Get the field object
                    try:
                        field = model._meta.get_field(field_name)

                        # Handle foreign key fields
                        if field.is_relation and field.many_to_one:
                            if value == "" or value is None:
                                # Skip empty foreign key values
                                mapped_data[field_name] = None
                                continue

                            # Try to find the related object
                            try:
                                # For UnitOfMeasure, try to get by code first
                                if field_name == "unit_of_measure":
                                    from items.models import UnitOfMeasure

                                    related_obj = UnitOfMeasure.objects.filter(
                                        code=value
                                    ).first()
                                    if not related_obj:
                                        validation_errors.append(
                                            f"Row {row_index + 5}: Unit of Measure '{value}' not found"
                                        )
                                        continue
                                    mapped_data[field_name] = related_obj

                                # For ItemCategory, try to get by code first
                                elif field_name == "item_category":
                                    from items.models import ItemCategory

                                    related_obj = ItemCategory.objects.filter(
                                        code=value
                                    ).first()
                                    if not related_obj:
                                        validation_errors.append(
                                            f"Row {row_index + 5}: Item Category '{value}' not found"
                                        )
                                        continue
                                    mapped_data[field_name] = related_obj

                                # For GeneralProductPostingGroup, try to get by code first
                                elif field_name == "general_product_posting_group":
                                    from postings.models import (
                                        GeneralProductPostingGroup,
                                    )

                                    related_obj = (
                                        GeneralProductPostingGroup.objects.filter(
                                            code=value
                                        ).first()
                                    )
                                    if not related_obj:
                                        validation_errors.append(
                                            f"Row {row_index + 5}: General Product Posting Group '{value}' not found"
                                        )
                                        continue
                                    mapped_data[field_name] = related_obj

                                # For InventoryPostingGroup, try to get by code first
                                elif field_name == "inventory_posting_group":
                                    from postings.models import (
                                        InventoryPostingGroup,
                                    )

                                    related_obj = InventoryPostingGroup.objects.filter(
                                        code=value
                                    ).first()
                                    if not related_obj:
                                        validation_errors.append(
                                            f"Row {row_index + 5}: Inventory Posting Group '{value}' not found"
                                        )
                                        continue
                                    mapped_data[field_name] = related_obj

                                # For other foreign keys, try to get by ID or name
                                else:
                                    # Try to get by ID first
                                    try:
                                        related_obj = field.related_model.objects.get(
                                            id=value
                                        )
                                        mapped_data[field_name] = related_obj
                                    except (
                                        ValueError,
                                        field.related_model.DoesNotExist,
                                    ):
                                        # Try to get by name or code
                                        if hasattr(field.related_model, "name"):
                                            related_obj = (
                                                field.related_model.objects.filter(
                                                    name=value
                                                ).first()
                                            )
                                        elif hasattr(field.related_model, "code"):
                                            related_obj = (
                                                field.related_model.objects.filter(
                                                    code=value
                                                ).first()
                                            )
                                        else:
                                            related_obj = None

                                        if not related_obj:
                                            validation_errors.append(
                                                f"Row {row_index + 5}: {field_name} '{value}' not found"
                                            )
                                            continue
                                        mapped_data[field_name] = related_obj

                            except Exception as e:
                                validation_errors.append(
                                    f"Row {row_index + 5}: Error processing {field_name} '{value}': {str(e)}"
                                )
                                continue

                        # Special handling for unit_of_measure field even if not detected as foreign key
                        elif field_name == "unit_of_measure":
                            from items.models import UnitOfMeasure

                            related_obj = UnitOfMeasure.objects.filter(
                                code=value
                            ).first()
                            if not related_obj:
                                validation_errors.append(
                                    f"Row {row_index + 5}: Unit of Measure '{value}' not found"
                                )
                                continue
                            mapped_data[field_name] = related_obj

                    except Exception as e:
                        validation_errors.append(
                            f"Row {row_index + 5}: Error validating field {field_name}: {str(e)}"
                        )
                        continue

                # Remove dependent fields from mapped_data after iteration is complete
                for field_name in fields_to_remove:
                    if field_name in mapped_data:
                        del mapped_data[field_name]

                # Clean up any remaining empty string values for foreign key fields
                for field_name, value in list(mapped_data.items()):
                    if value == "" or value is None:
                        try:
                            field = model._meta.get_field(field_name)
                            if field.is_relation and field.many_to_one:
                                mapped_data[field_name] = None
                            elif value == "":
                                # Remove empty string values for non-foreign key fields
                                del mapped_data[field_name]
                        except:
                            # If field doesn't exist, remove it
                            if field_name in mapped_data:
                                del mapped_data[field_name]

                # Note: Handler functions are now handled in the new import_handlers system

                # If there are validation errors, skip this row
                if validation_errors:
                    stats["errors"].extend(validation_errors)
                    stats["failed"] += 1
                    continue

                # Save to actual model table
                try:
                    # Check if record exists in the actual model
                    if (
                        identifier_field == "item_name"
                        and model._meta.model_name == "item"
                    ):
                        # This is a new item creation - check if item with same name exists
                        from items.models import Item

                        existing_item_by_name = Item.objects.filter(
                            item_name=mapped_data["item_name"]
                        ).first()
                        if existing_item_by_name:
                            # Item exists with this name, update it instead of creating new
                            obj = existing_item_by_name
                            for key, value in mapped_data.items():
                                if key != "no":  # Don't update the item number
                                    setattr(obj, key, value)
                            obj.save()
                            stats["updated"] += 1
                            created = False
                        else:
                            # No existing item found, create new one
                            # If item number is provided, use it; otherwise let the save method generate it
                            if "no" in mapped_data and mapped_data["no"]:
                                # Use the provided item number - this will create the item with that number
                                pass
                            # Note: If 'no' is not in mapped_data, the save method will generate it automatically

                            # Create new item
                            obj = model.objects.create(**mapped_data)
                            stats["created"] += 1
                            created = True
                    else:
                        # Normal flow for existing items or other models
                        filter_kwargs = {
                            identifier_field: mapped_data[identifier_field]
                        }

                        # Try to get existing record first
                        try:
                            obj = model.objects.get(**filter_kwargs)
                            # Record exists, update it
                            for key, value in mapped_data.items():
                                setattr(obj, key, value)
                            obj.save()
                            stats["updated"] += 1
                            created = False
                        except model.DoesNotExist:
                            # Record doesn't exist, create it
                            # For Item model, check for existing items by name first
                            if model._meta.model_name == "item":
                                # Ensure we have the required item_name field
                                if (
                                    "item_name" not in mapped_data
                                    or not mapped_data["item_name"]
                                ):
                                    error_msg = f"Row {row_index + 5}: Item Name is required for new item creation."
                                    stats["errors"].append(error_msg)
                                    stats["failed"] += 1
                                    logger.warning(error_msg)
                                    continue

                                # Check for existing items by name - if found, update instead of creating
                                from items.models import Item

                                existing_item_by_name = Item.objects.filter(
                                    item_name=mapped_data["item_name"]
                                ).first()
                                if existing_item_by_name:
                                    # Item exists with this name, update it instead of creating new
                                    obj = existing_item_by_name
                                    for key, value in mapped_data.items():
                                        if key != "no":  # Don't update the item number
                                            setattr(obj, key, value)
                                    obj.save()
                                    stats["updated"] += 1
                                    created = False
                                else:
                                    # No existing item found, create new one
                                    # If item number is provided, use it; otherwise let the save method generate it
                                    if "no" in mapped_data and mapped_data["no"]:
                                        # Use the provided item number - this will create the item with that number
                                        pass
                                    # Note: If 'no' is not in mapped_data, the save method will generate it automatically

                                    # Create new item
                                    obj = model.objects.create(**mapped_data)
                                    stats["created"] += 1
                                    created = True
                            else:
                                # For non-Item models, create normally
                                obj = model.objects.create(**mapped_data)
                                stats["created"] += 1
                                created = True

                    # Process dependent fields after item creation/update
                    if dependent_field_values and model._meta.model_name == "item":
                        from items.models import UnitOfMeasure, ItemUnitOfMeasure

                        for field_name, value in dependent_field_values.items():
                            try:
                                # Get the UnitOfMeasure by code
                                unit_of_measure = UnitOfMeasure.objects.filter(
                                    code=value
                                ).first()
                                if not unit_of_measure:
                                    logger.warning(
                                        f"Row {row_index + 5}: Unit of Measure '{value}' not found for {field_name}"
                                    )
                                    continue

                                # Create or get ItemUnitOfMeasure
                                item_uom, created = (
                                    ItemUnitOfMeasure.objects.get_or_create(
                                        item=obj,
                                        unit_of_measure=unit_of_measure,
                                        defaults={"quantity_per_unit": 1},
                                    )
                                )

                                # Set the field on the object
                                setattr(obj, field_name, item_uom)
                                obj.save()

                            except Exception as e:
                                logger.warning(
                                    f"Row {row_index + 5}: Error processing {field_name} '{value}': {str(e)}"
                                )

                    # Note: Relationship handling is now done in the new import_handlers system

                    # Store in ConfigPackageTable's data field for tracking
                    existing_records = package_table.data or []
                    record_exists = False

                    # Create a serializable version of mapped_data for JSON storage
                    serializable_data = {}
                    for key, value in mapped_data.items():
                        if hasattr(value, "__str__"):
                            # Convert Django model instances to string representation
                            serializable_data[key] = str(value)
                        else:
                            # Keep primitive values as is
                            serializable_data[key] = value

                    for i, record in enumerate(existing_records):
                        if (
                            record.get(identifier_field)
                            == serializable_data[identifier_field]
                        ):
                            existing_records[i].update(serializable_data)
                            record_exists = True
                            break

                    if not record_exists:
                        existing_records.append(serializable_data)

                    # Save updated data to ConfigPackageTable
                    package_table.data = existing_records
                    package_table.save()

                except Exception as e:
                    error_msg = f"Row {row_index + 5}: Error saving record: {str(e)}"
                    stats["errors"].append(error_msg)
                    stats["failed"] += 1
                    logger.error(error_msg)
                    continue

            except Exception as e:
                error_msg = f"Row {row_index + 5}: Error processing row: {str(e)}"
                stats["errors"].append(error_msg)
                stats["failed"] += 1
                logger.error(error_msg)

        # Prepare response message based on results
        if stats["failed"] > 0:
            message = f"Import completed with {stats['failed']} errors. {stats['created']} records created, {stats['updated']} records updated."
        else:
            message = f"Import completed successfully. {stats['created']} records created, {stats['updated']} records updated."

        return JsonResponse(
            {
                "message": message,
                "success": stats["failed"] == 0,
                "statistics": stats,
            }
        )

    except Exception as e:
        logger.error(f"Import failed for tenant {request.tenant.schema_name}: {str(e)}")
        return JsonResponse(
            {"message": f"Import failed: {str(e)}", "success": False}, status=500
        )


def export_package_to_excel(request, package_id):
    try:
        package = ConfigPackage.objects.get(id=package_id)

        # Create a response with Excel mime type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{package.name}_export.xlsx"'
        )

        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        for table in package.tables.all():
            ws = wb.create_sheet(title=table.table_name)

            # Get the data
            data = json.loads(table.data)

            # Process datetime fields to remove timezone
            for row in data:
                for key, value in row.items():
                    # Check if the value looks like a datetime string
                    if isinstance(value, str) and ("T" in value or "+" in value):
                        try:
                            # Parse the datetime string
                            dt = parser.parse(value)
                            # Remove timezone info
                            if dt.tzinfo is not None:
                                dt = dt.replace(tzinfo=None)
                            # Format back to string
                            row[key] = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            continue

            # Write headers
            if data:
                headers = list(data[0].keys())
                ws.append(headers)

                # Write data
                for row in data:
                    ws.append([row.get(header) for header in headers])

        # Save the workbook to the response
        wb.save(response)
        return response

    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@require_http_methods(["GET"])
@csrf_exempt
def check_import_status(request, task_id):
    """Check the status of a background import task"""
    try:
        from celery.result import AsyncResult

        # Get the task result
        task_result = AsyncResult(task_id)

        # Get task state and info
        state = task_result.state
        info = task_result.info if task_result.info else {}

        # Map Celery states to our status format
        if state == "PENDING":
            status_data = {
                "state": "PENDING",
                "progress": 0,
                "message": "Task is waiting to be processed",
                "status": "pending",
            }
        elif state == "STARTED":
            status_data = {
                "state": "STARTED",
                "progress": 10,
                "message": "Task has started processing",
                "status": "processing",
            }
        elif state == "PROGRESS":
            # Extract progress from task info
            progress = info.get("progress", 0)
            message = info.get("message", "Processing data...")
            status_data = {
                "state": "PROGRESS",
                "progress": progress,
                "message": message,
                "status": "processing",
                "meta": {"progress": progress, "message": message},
            }
        elif state == "SUCCESS":
            # Extract results from task info
            result = task_result.result if task_result.result else {}
            status_data = {
                "state": "SUCCESS",
                "progress": 100,
                "message": result.get("message", "Import completed successfully"),
                "status": "completed",
                "success": True,
                "statistics": result.get("statistics", {}),
                "meta": {"progress": 100, "message": "Import completed"},
            }
        elif state == "FAILURE":
            # Extract error from task info
            error_message = str(task_result.info) if task_result.info else "Task failed"
            status_data = {
                "state": "FAILURE",
                "progress": 0,
                "message": "Import failed",
                "status": "failed",
                "success": False,
                "error": error_message,
            }
        else:
            status_data = {
                "state": "UNKNOWN",
                "progress": 0,
                "message": f"Unknown task state: {state}",
                "status": "unknown",
            }

        return JsonResponse(status_data)

    except Exception as e:
        print(f"Error checking import status: {e}")
        return JsonResponse(
            {
                "state": "ERROR",
                "progress": 0,
                "message": f"Error checking task status: {str(e)}",
                "status": "error",
                "error": str(e),
            },
            status=500,
        )


@require_http_methods(["POST"])
def validate_journal_import(request):
    try:
        if "file" not in request.FILES:
            return JsonResponse({"valid": False, "error": "No file uploaded"})

        file = request.FILES["file"]

        # Create temp directory if it doesn't exist
        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)

        # Generate unique filename
        filename = f"journal_import_{uuid.uuid4()}.xlsx"
        temp_path = os.path.join("temp", filename)
        abs_path = os.path.join(settings.MEDIA_ROOT, temp_path)

        # Ensure directory exists
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)

        # Save file
        with open(abs_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # Start celery task
        task = process_journal_import.delay(
            os.path.normpath(abs_path),
            schema_name=request.tenant.schema_name,
            user_id=request.user.id,  # Pass the user ID to the task
        )

        return JsonResponse(
            {"valid": True, "message": "Import started", "task_id": task.id}
        )

    except Exception as e:
        return JsonResponse(
            {"valid": False, "error": f"Error starting import: {str(e)}"}
        )


class ConfigPackageViewSet(viewsets.ModelViewSet):
    # Match ItemJournalViewSet / sales ViewSets: DEFAULT CustomJWTAuthentication can 401
    # Bearer tokens that SimpleJWT accepts (SPA + tenant middleware already scope the tenant).
    queryset = ConfigPackage.objects.all()
    serializer_class = ConfigPackageSerializer
    lookup_field = "code"
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ConfigPackageDetailSerializer
        return ConfigPackageSerializer

    @action(detail=True, methods=["get"])
    def tables(self, request, code=None):
        package = self.get_object()
        tables = ConfigPackageTable.objects.filter(package_code=package)
        serializer = ConfigPackageTableSerializer(tables, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def upsert(self, request):
        code = request.data.get("code")
        system_id = request.data.get("system_id")
        package_name = request.data.get("package_name")

        if not code:
            raise ValidationError({"code": "Code is required"})
        if not package_name:
            raise ValidationError({"package_name": "Package name is required"})

        try:
            if system_id:
                # Find by system_id
                package = ConfigPackage.objects.filter(system_id=system_id).first()
                if not package:
                    raise ConfigPackage.DoesNotExist

                old_code = package.code

                # If code is changing, handle the primary key update
                if old_code != code:
                    # Check if new code already exists
                    if ConfigPackage.objects.filter(code=code).exists():
                        raise ValidationError({"code": "This code is already in use"})

                    # Create new record with new code
                    new_package_data = {
                        "code": code,
                        "package_name": package_name,
                        "status": package.status,
                        "system_id": system_id,
                        # Copy any other fields you need to preserve
                    }

                    # Create new record
                    serializer = self.get_serializer(data=new_package_data)
                    if serializer.is_valid(raise_exception=True):
                        with transaction.atomic():
                            # Delete old record
                            package.delete()
                            # Save new record
                            new_package = serializer.save()
                            return Response(serializer.data, status=status.HTTP_200_OK)
                else:
                    # Just updating package_name
                    serializer = self.get_serializer(
                        package, data={"package_name": package_name}, partial=True
                    )
                    if serializer.is_valid(raise_exception=True):
                        package = serializer.save()
                        return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                # Handle create/update without system_id
                try:
                    package = ConfigPackage.objects.get(code=code)
                    serializer = self.get_serializer(
                        package, data={"package_name": package_name}, partial=True
                    )
                except ConfigPackage.DoesNotExist:
                    serializer = self.get_serializer(
                        data={"code": code, "package_name": package_name}
                    )

                if serializer.is_valid(raise_exception=True):
                    package = serializer.save()
                    return Response(
                        serializer.data,
                        status=(
                            status.HTTP_200_OK if package else status.HTTP_201_CREATED
                        ),
                    )

        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(f"Error in upsert: {str(e)}")
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ConfigPackageTableViewSet(viewsets.ModelViewSet):
    queryset = ConfigPackageTable.objects.all()
    serializer_class = ConfigPackageTableSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    http_method_names = ["get", "post", "put", "delete"]
    pagination_class = None
    # filter_backends = [DjangoFilterBackend]
    filterset_fields = ["package_code"]

    def get_queryset(self):
        queryset = ConfigPackageTable.objects.all()
        package_code = self.request.query_params.get("package_code", None)
        if package_code is not None:
            queryset = queryset.filter(package_code=package_code)
        return queryset


@require_http_methods(["GET"])
def get_field_config(request, package_code, table_id):
    """
    Get field configuration for a specific table in a package
    """
    try:

        # get tenaat
        with tenant_context(request.tenant):
            package = ConfigPackage.objects.get(code=package_code)
            table = ConfigPackageTable.objects.get(
                package_code=package, table_id=table_id
            )
            print(f"DEBUG: Table: {table}")
            print(f"DEBUG: Table field_config: {table.field_config}")
            # Ensure field_config is populated
            if (
                not table.field_config
                or "fields" not in table.field_config
                or len(table.field_config.get("fields", [])) == 0
            ):
                print(f"DEBUG: Populating field config")
            table.populate_field_config()
            table.save()

        return JsonResponse(
            {
                "success": True,
                "field_config": table.field_config,
                "export_fields": table.get_export_fields(),
            }
        )

    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)
    except ConfigPackageTable.DoesNotExist:
        return JsonResponse({"error": "Table not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@require_http_methods(["POST"])
@csrf_exempt
def update_field_config(request, package_code, table_id):
    """
    Update field configuration for a specific table in a package
    """
    try:
        data = json.loads(request.body)
        package = ConfigPackage.objects.get(code=package_code)
        table = ConfigPackageTable.objects.get(package_code=package, table_id=table_id)

        # Update field configuration
        table.update_field_config(
            primary_fields=data.get("primary_fields"),
            default_fields=data.get("default_fields"),
            custom_fields=data.get("custom_fields"),
            excluded_fields=data.get("excluded_fields"),
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Field configuration updated successfully",
                "field_config": table.field_config,
                "export_fields": table.get_export_fields(),
            }
        )

    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)
    except ConfigPackageTable.DoesNotExist:
        return JsonResponse({"error": "Table not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@require_http_methods(["GET"])
def get_export_fields(request, package_code, table_id):
    """
    Get the list of fields that will be included in exports
    """
    try:
        package = ConfigPackage.objects.get(code=package_code)
        table = ConfigPackageTable.objects.get(package_code=package, table_id=table_id)

        export_fields = table.get_export_fields()

        return JsonResponse(
            {
                "success": True,
                "export_fields": export_fields,
                "field_config": table.field_config,
            }
        )

    except ConfigPackage.DoesNotExist:
        return JsonResponse({"error": "Package not found"}, status=404)
    except ConfigPackageTable.DoesNotExist:
        return JsonResponse({"error": "Table not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
