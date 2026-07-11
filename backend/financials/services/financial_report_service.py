from __future__ import annotations

import html
import re
from datetime import date, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.db.models import QuerySet, Sum
from django.utils import timezone

from financials import enums
from financials.currency import get_local_currency_code
from financials.models import (
    FinancialReport,
    FinancialReportColumnLine,
    FinancialReportRowLine,
    G_LAccount,
    GeneralLedgerEntry,
)
from financials.services.period_formula import ReportPeriod, resolve_comparison_period
from reports.utils.formatters import format_currency


def _column_key(line_no: int) -> str:
    return f"col_{line_no}"


def _resolve_period(report: FinancialReport, today: Optional[date] = None) -> ReportPeriod:
    today = today or timezone.now().date()
    period_type = report.period_type or enums.FinancialReportPeriodType.Month.value

    if period_type == enums.FinancialReportPeriodType.Day.value:
        return ReportPeriod(today, today, today.strftime("%b %d, %Y"))
    if period_type == enums.FinancialReportPeriodType.Week.value:
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return ReportPeriod(start, end, f"{start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}")
    if period_type == enums.FinancialReportPeriodType.Quarter.value:
        quarter = (today.month - 1) // 3
        start = date(today.year, quarter * 3 + 1, 1)
        if quarter == 3:
            end = date(today.year, 12, 31)
        else:
            end = date(today.year, (quarter + 1) * 3 + 1, 1) - timedelta(days=1)
        return ReportPeriod(start, end, f"Q{quarter + 1} {today.year}")
    if period_type == enums.FinancialReportPeriodType.Year.value:
        start = date(today.year, 1, 1)
        end = date(today.year, 12, 31)
        return ReportPeriod(start, end, str(today.year))
    if period_type == enums.FinancialReportPeriodType.Accounting_Period.value:
        return ReportPeriod(date(1900, 1, 1), today, "All periods")

    # Default: Month
    start = date(today.year, today.month, 1)
    if today.month == 12:
        end = date(today.year, 12, 31)
    else:
        end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    return ReportPeriod(start, end, today.strftime("%B %Y"))


def _period_from_dates(start_date: date, end_date: date) -> ReportPeriod:
    if start_date > end_date:
        raise ValueError("Start date must be on or before end date.")
    if start_date == end_date:
        label = start_date.strftime("%b %d, %Y")
    else:
        label = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
    return ReportPeriod(start_date, end_date, label)


def _parse_dimension_filter(value: str) -> List[str]:
    if not value or not str(value).strip():
        return []
    return [part.strip() for part in str(value).split("|") if part.strip()]


def _apply_report_filters(
    queryset: QuerySet,
    report: FinancialReport,
) -> QuerySet:
    qs = queryset
    dim_filters = (
        ("dimension_1_filter", "global_dimension_1__code"),
        ("dimension_2_filter", "global_dimension_2__code"),
    )
    for field_name, lookup in dim_filters:
        codes = _parse_dimension_filter(getattr(report, field_name, "") or "")
        if codes:
            qs = qs.filter(**{f"{lookup}__in": codes})
    return qs


def _parse_account_numbers(totaling: str) -> List[str]:
    if not totaling or not str(totaling).strip():
        return []
    raw = str(totaling).strip()
    if "|" in raw:
        start, end = [part.strip() for part in raw.split("|", 1)]
        if start and end:
            return list(
                G_LAccount.objects.filter(no__gte=start, no__lte=end)
                .order_by("no")
                .values_list("no", flat=True)
            )
    return [raw]


def _aggregate_amount(
    queryset: QuerySet,
    amount_type: str,
) -> float:
    if amount_type == enums.FinancialReportAmountType.Debits.value:
        result = queryset.filter(amount__gt=0).aggregate(total=Sum("amount"))
        return float(result["total"] or 0)
    if amount_type == enums.FinancialReportAmountType.Credits.value:
        result = queryset.filter(amount__lt=0).aggregate(total=Sum("amount"))
        return abs(float(result["total"] or 0))
    if amount_type == enums.FinancialReportAmountType.Debits_Minus_Credits.value:
        debits = queryset.filter(amount__gt=0).aggregate(total=Sum("amount"))
        credits = queryset.filter(amount__lt=0).aggregate(total=Sum("amount"))
        return float(debits["total"] or 0) - abs(float(credits["total"] or 0))
    if amount_type == enums.FinancialReportAmountType.Credits_Minus_Debits.value:
        debits = queryset.filter(amount__gt=0).aggregate(total=Sum("amount"))
        credits = queryset.filter(amount__lt=0).aggregate(total=Sum("amount"))
        return abs(float(credits["total"] or 0)) - float(debits["total"] or 0)

    result = queryset.aggregate(total=Sum("amount"))
    return float(result["total"] or 0)


def _column_queryset(
    base_qs: QuerySet,
    column: FinancialReportColumnLine,
    period: ReportPeriod,
) -> QuerySet:
    column_type = column.column_type or enums.FinancialReportColumnType.Net_Change.value
    if column_type == enums.FinancialReportColumnType.Balance_at_Date.value:
        return base_qs.filter(posting_date__lte=period.end_date)
    if column_type == enums.FinancialReportColumnType.Beginning_Balance.value:
        return base_qs.filter(posting_date__lt=period.start_date)
    return base_qs.filter(
        posting_date__gte=period.start_date,
        posting_date__lte=period.end_date,
    )


def _eval_formula(formula: str, amounts_by_row: Dict[str, float]) -> float:
    if not formula or not str(formula).strip():
        return 0.0
    expr = re.sub(r"\s+", "", str(formula))
    if not expr:
        return 0.0

    total: Optional[float] = None
    op = "+"
    idx = 0
    while idx < len(expr):
        char = expr[idx]
        if char in "+-":
            op = char
            idx += 1
            continue
        match = re.match(r"\d+", expr[idx:])
        if not match:
            idx += 1
            continue
        token = match.group(0)
        idx += len(token)
        value = float(amounts_by_row.get(token, 0.0))
        if total is None:
            total = value if op == "+" else -value
        elif op == "+":
            total += value
        else:
            total -= value
        op = "+"
    return float(total or 0.0)


def _should_show_row(
    line: FinancialReportRowLine,
    amounts: Dict[str, Optional[float]],
    show_all_lines: bool,
) -> bool:
    if show_all_lines:
        return True
    show = line.show or enums.FinancialReportShowLine.Yes.value
    if show == enums.FinancialReportShowLine.No.value:
        return False
    if show in (
        enums.FinancialReportShowLine.If_Amount_Not_Zero.value,
        enums.FinancialReportShowLine.If_Any_Column_Not_Zero.value,
    ):
        return any(abs(v or 0) > 0.00001 for v in amounts.values())
    if show == enums.FinancialReportShowLine.When_Positive_Balance.value:
        return any((v or 0) > 0.00001 for v in amounts.values())
    if show == enums.FinancialReportShowLine.When_Negative_Balance.value:
        return any((v or 0) < -0.00001 for v in amounts.values())
    return True


def _amounts_have_value(amounts: Dict[str, Optional[float]]) -> bool:
    return any(abs(v or 0) > 0.00001 for v in amounts.values())


def _apply_section_visibility(
    row_lines: List[FinancialReportRowLine],
    visibility: Dict[int, bool],
    amounts_by_line: Dict[int, Dict[str, float]],
) -> None:
    """
    Hide Begin-Total / End-Total blocks when every posting line in the section
    is hidden and the end total has no amounts (BC-style section collapsing).
    """
    i = 0
    while i < len(row_lines):
        line = row_lines[i]
        if (line.row_type or "") != enums.FinancialReportRowType.Begin_Total.value:
            i += 1
            continue

        end_idx = None
        for j in range(i + 1, len(row_lines)):
            if (row_lines[j].row_type or "") == enums.FinancialReportRowType.End_Total.value:
                end_idx = j
                break
        if end_idx is None:
            i += 1
            continue

        section_postings = [
            section_line
            for section_line in row_lines[i + 1:end_idx]
            if (section_line.row_type or "") == enums.FinancialReportRowType.Posting.value
        ]
        end_line = row_lines[end_idx]
        end_amounts = amounts_by_line.get(end_line.line_no, {})
        posting_visible = any(visibility.get(posting.line_no, False) for posting in section_postings)
        end_visible = visibility.get(end_line.line_no, False) and _amounts_have_value(end_amounts)
        section_active = posting_visible or end_visible

        if not section_active:
            visibility[line.line_no] = False
            visibility[end_line.line_no] = False
            for posting in section_postings:
                visibility[posting.line_no] = False
        else:
            visibility[line.line_no] = True

        i = end_idx + 1


INDENT_PT = 14
HEADER_BG = "#1a5276"
SECTION_BG = "#f4f7fa"
TOTAL_BG = "#e8eef4"
GRID_COLOR = "#d0d7de"
ACCENT_LINE = "#8899aa"


def _visible_export_rows(data: dict) -> List[dict]:
    return [row for row in (data.get("rows") or []) if row.get("visible", True)]


def _row_is_bold(row: dict) -> bool:
    if row.get("bold"):
        return True
    row_type = row.get("row_type") or ""
    return row_type in (
        enums.FinancialReportRowType.Header.value,
        enums.FinancialReportRowType.End_Total.value,
        enums.FinancialReportRowType.Total.value,
    )


def _row_bg_hex(row: dict) -> Optional[str]:
    row_type = row.get("row_type") or ""
    if row_type == enums.FinancialReportRowType.Header.value:
        return HEADER_BG
    if row_type in (
        enums.FinancialReportRowType.End_Total.value,
        enums.FinancialReportRowType.Total.value,
    ):
        return TOTAL_BG
    if row_type == enums.FinancialReportRowType.Begin_Total.value:
        return SECTION_BG
    return None


def _row_fg_hex(row: dict) -> Optional[str]:
    if (row.get("row_type") or "") == enums.FinancialReportRowType.Header.value:
        return "#ffffff"
    return None


def _row_has_top_line(row: dict) -> bool:
    row_type = row.get("row_type") or ""
    return row_type in (
        enums.FinancialReportRowType.End_Total.value,
        enums.FinancialReportRowType.Total.value,
    )


def _pdf_font_name(row: dict) -> str:
    bold = _row_is_bold(row)
    italic = bool(row.get("italic"))
    if bold and italic:
        return "Helvetica-BoldOblique"
    if bold:
        return "Helvetica-Bold"
    if italic:
        return "Helvetica-Oblique"
    return "Helvetica"


def _excel_hex(color: Optional[str]) -> Optional[str]:
    if not color:
        return None
    return color.lstrip("#").upper()


def _excel_font(row: dict) -> Font:
    kwargs: dict = {}
    if _row_is_bold(row):
        kwargs["bold"] = True
    if row.get("italic"):
        kwargs["italic"] = True
    if row.get("underline"):
        kwargs["underline"] = "single"
    fg = _excel_hex(_row_fg_hex(row))
    if fg:
        kwargs["color"] = fg
    return Font(**kwargs)


def _excel_fill(row: dict) -> Optional[PatternFill]:
    bg = _excel_hex(_row_bg_hex(row))
    if not bg:
        return None
    return PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def _apply_excel_row_style(ws, excel_row: int, row: dict, amount_col_count: int) -> None:
    font = _excel_font(row)
    fill = _excel_fill(row)
    top_line = _row_has_top_line(row)
    border = Border(top=Side(style="thin", color=_excel_hex(ACCENT_LINE))) if top_line else None
    indent = int(row.get("indentation") or 0)

    for col_idx in range(1, 2 + amount_col_count + 1):
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if col_idx == 2:
            cell.alignment = Alignment(indent=indent, vertical="center")
        elif col_idx >= 3:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if cell.value is not None:
                cell.number_format = "#,##0.00"
        else:
            cell.alignment = Alignment(vertical="center")


class FinancialReportService:
    def __init__(self, queryset: Optional[QuerySet] = None):
        self.base_queryset = queryset or GeneralLedgerEntry.objects.all()

    def generate(
        self,
        report_name: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> dict:
        report = FinancialReport.objects.select_related(
            "row_definition",
            "column_definition",
        ).get(name=report_name)

        if not report.row_definition_id:
            raise ValueError("Financial report has no row definition.")
        if not report.column_definition_id:
            raise ValueError("Financial report has no column definition.")

        row_lines = list(
            FinancialReportRowLine.objects.filter(row_group=report.row_definition)
            .order_by("line_no")
        )
        column_lines = list(
            FinancialReportColumnLine.objects.filter(column_group=report.column_definition)
            .order_by("line_no")
        )
        if not row_lines:
            raise ValueError("Row definition has no lines.")
        if not column_lines:
            raise ValueError("Column definition has no lines.")

        period_type = report.period_type or enums.FinancialReportPeriodType.Month.value
        if start_date is not None and end_date is not None:
            period = _period_from_dates(start_date, end_date)
        elif report.start_date and report.end_date:
            period = _period_from_dates(report.start_date, report.end_date)
        else:
            period = _resolve_period(report)
        filtered_qs = _apply_report_filters(self.base_queryset, report)

        columns = [
            {
                "key": _column_key(col.line_no),
                "line_no": col.line_no,
                "column_no": col.column_no or str(col.line_no),
                "header": col.column_header or f"Column {col.line_no}",
                "column_type": col.column_type,
                "comparison_period_formula": col.comparison_period_formula or "0M",
            }
            for col in column_lines
        ]

        formula_columns = {
            col.line_no
            for col in column_lines
            if (col.formula or "").strip()
        }
        column_periods = {
            col.line_no: resolve_comparison_period(
                period,
                col.comparison_period_formula,
                period_type=period_type,
            )
            for col in column_lines
        }

        raw_amounts: Dict[int, Dict[str, float]] = {
            line.line_no: {_column_key(col.line_no): 0.0 for col in column_lines}
            for line in row_lines
        }
        row_no_amounts: Dict[str, float] = {}
        subtotal_stacks: Dict[str, List[float]] = {
            _column_key(col.line_no): [] for col in column_lines
        }

        for line in row_lines:
            row_type = line.row_type or enums.FinancialReportRowType.Posting.value
            totaling_type = line.totaling_type or enums.FinancialReportTotalingType.Posting_Accounts.value

            for col in column_lines:
                col_key = _column_key(col.line_no)
                amount = 0.0

                if col.line_no in formula_columns:
                    raw_amounts[line.line_no][col_key] = 0.0
                    continue

                col_period = column_periods[col.line_no]
                col_qs = _column_queryset(filtered_qs, col, col_period)
                effective_amount_type = (
                    col.amount_type
                    or line.amount_type
                    or enums.FinancialReportAmountType.Net_Amount.value
                )

                if row_type == enums.FinancialReportRowType.Header.value:
                    amount = 0.0
                elif totaling_type == enums.FinancialReportTotalingType.Formula.value:
                    amount = 0.0
                elif row_type in (
                    enums.FinancialReportRowType.Begin_Total.value,
                    enums.FinancialReportRowType.End_Total.value,
                ):
                    if row_type == enums.FinancialReportRowType.Begin_Total.value:
                        subtotal_stacks[col_key].append(0.0)
                        amount = 0.0
                    else:
                        stack = subtotal_stacks[col_key]
                        amount = stack.pop() if stack else 0.0
                elif totaling_type == enums.FinancialReportTotalingType.Posting_Accounts.value:
                    account_nos = _parse_account_numbers(line.totaling or "")
                    if account_nos:
                        line_qs = col_qs.filter(gl_account_id__in=account_nos)
                        amount = _aggregate_amount(line_qs, effective_amount_type)
                    else:
                        amount = 0.0
                else:
                    amount = 0.0

                if line.show_opposite_sign:
                    amount *= -1
                if col.show_opposite_sign:
                    amount *= -1

                raw_amounts[line.line_no][col_key] = amount

                if row_type == enums.FinancialReportRowType.Posting.value and subtotal_stacks[col_key]:
                    subtotal_stacks[col_key][-1] += amount

        for col in column_lines:
            if col.line_no not in formula_columns:
                continue
            col_key = _column_key(col.line_no)
            for line in row_lines:
                amounts_by_col_no = {
                    (other.column_no or str(other.line_no)): raw_amounts[line.line_no][
                        _column_key(other.line_no)
                    ]
                    for other in column_lines
                    if other.line_no not in formula_columns
                }
                amount = _eval_formula(col.formula or "", amounts_by_col_no)
                if col.show_opposite_sign:
                    amount *= -1
                if line.show_opposite_sign:
                    amount *= -1
                raw_amounts[line.line_no][col_key] = amount

        for line in row_lines:
            if (line.totaling_type or "") != enums.FinancialReportTotalingType.Formula.value:
                continue
            for col in column_lines:
                col_key = _column_key(col.line_no)
                row_no_amounts_for_col = {
                    (rl.row_no or str(rl.line_no)): raw_amounts[rl.line_no][col_key]
                    for rl in row_lines
                }
                amount = _eval_formula(line.totaling or "", row_no_amounts_for_col)
                if line.show_opposite_sign:
                    amount *= -1
                raw_amounts[line.line_no][col_key] = amount

        for line in row_lines:
            row_key = line.row_no or str(line.line_no)
            if line.line_no in raw_amounts:
                row_no_amounts[row_key] = raw_amounts[line.line_no][columns[0]["key"]]

        rows: List[dict] = []
        visibility: Dict[int, bool] = {}
        rounded_amounts: Dict[int, Dict[str, float]] = {}
        for line in row_lines:
            amounts = {
                key: round(value, 2)
                for key, value in raw_amounts[line.line_no].items()
            }
            rounded_amounts[line.line_no] = amounts
            visibility[line.line_no] = _should_show_row(
                line,
                amounts,
                report.show_all_lines,
            )

        if not report.show_all_lines:
            _apply_section_visibility(row_lines, visibility, rounded_amounts)

        for line in row_lines:
            amounts = rounded_amounts[line.line_no]
            rows.append(
                {
                    "line_no": line.line_no,
                    "row_no": line.row_no or str(line.line_no),
                    "description": line.description,
                    "row_type": line.row_type,
                    "bold": line.bold,
                    "italic": line.italic,
                    "underline": line.underline,
                    "indentation": line.indentation,
                    "amounts": amounts,
                    "visible": visibility[line.line_no],
                }
            )

        return {
            "report_name": report.name,
            "description": report.description,
            "period_type": report.period_type,
            "period_label": period.label,
            "start_date": period.start_date.isoformat(),
            "end_date": period.end_date.isoformat(),
            "currency_code": get_local_currency_code(),
            "columns": columns,
            "rows": rows,
        }

    @staticmethod
    def render_html(data: dict) -> str:
        title = html.escape(data.get("description") or data.get("report_name") or "Financial Report")
        period_label = html.escape(data.get("period_label") or "")
        currency = html.escape(data.get("currency_code") or "")
        columns = data.get("columns") or []
        rows = _visible_export_rows(data)

        header_cells = "".join(
            f"<th class='amount'>{html.escape(str(col.get('header') or ''))}</th>"
            for col in columns
        )
        body_rows = []
        for row in rows:
            indent = int(row.get("indentation") or 0)
            description = html.escape(str(row.get("description") or ""))
            bg = _row_bg_hex(row)
            fg = _row_fg_hex(row)
            style_parts = [f"padding-left:{12 + indent * 16}px"]
            if _row_is_bold(row):
                style_parts.append("font-weight:700")
            if row.get("italic"):
                style_parts.append("font-style:italic")
            if row.get("underline"):
                style_parts.append("text-decoration:underline")
            if fg:
                style_parts.append(f"color:{fg}")
            if bg:
                style_parts.append(f"background:{bg}")
            if _row_has_top_line(row):
                style_parts.append(f"border-top:2px solid {ACCENT_LINE}")
            style_attr = f' style="{";".join(style_parts)}"'
            amount_cells = []
            amounts = row.get("amounts") or {}
            for col in columns:
                key = col.get("key")
                value = amounts.get(key)
                if value is None:
                    display = ""
                else:
                    display = html.escape(format_currency(float(value)))
                amount_cells.append(f"<td class='amount'{style_attr}>{display}</td>")
            body_rows.append(
                f"<tr{style_attr}>"
                f"<td{style_attr}>{html.escape(str(row.get('row_no') or ''))}</td>"
                f"<td{style_attr}>{description}</td>"
                f"{''.join(amount_cells)}"
                f"</tr>"
            )

        return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8" />
  <title>{title}</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; color: #111; }}
    h1 {{ margin: 0 0 4px; font-size: 22px; color: {HEADER_BG}; }}
    .meta {{ color: #555; margin-bottom: 18px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border-bottom: 1px solid {GRID_COLOR}; padding: 8px 10px; text-align: left; }}
    th {{ background: {HEADER_BG}; color: #fff; font-weight: 600; }}
    td.amount, th.amount {{ text-align: right; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <div class="meta">{period_label} · {currency}</div>
  <table>
    <thead>
      <tr>
        <th>Row No.</th>
        <th>Description</th>
        {header_cells}
      </tr>
    </thead>
    <tbody>
      {''.join(body_rows)}
    </tbody>
  </table>
</body>
</html>"""

    @staticmethod
    def generate_pdf(data: dict) -> bytes:
        title = data.get("description") or data.get("report_name") or "Financial Report"
        period_label = data.get("period_label") or ""
        currency = data.get("currency_code") or ""
        columns = data.get("columns") or []
        rows = _visible_export_rows(data)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(LETTER),
            leftMargin=36,
            rightMargin=36,
            topMargin=40,
            bottomMargin=36,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            textColor=colors.HexColor(HEADER_BG),
            spaceAfter=4,
        )
        elements = [
            Paragraph(html.escape(str(title)), title_style),
            Paragraph(
                html.escape(f"{period_label} · {currency}".strip(" ·")),
                styles["Normal"],
            ),
            Spacer(1, 12),
        ]

        def _desc_para(row: dict) -> Paragraph:
            indent = int(row.get("indentation") or 0)
            fg = _row_fg_hex(row) or "#111111"
            ps = ParagraphStyle(
                f"Desc{row.get('line_no')}",
                parent=styles["Normal"],
                leftIndent=indent * INDENT_PT,
                fontName=_pdf_font_name(row),
                fontSize=9,
                textColor=colors.HexColor(fg),
                underline=bool(row.get("underline")),
            )
            return Paragraph(html.escape(str(row.get("description") or "")), ps)

        def _amount_para(row: dict, value: Optional[float]) -> Paragraph:
            fg = _row_fg_hex(row) or "#111111"
            text = "" if value is None else format_currency(float(value))
            ps = ParagraphStyle(
                f"Amt{row.get('line_no')}",
                parent=styles["Normal"],
                alignment=TA_RIGHT,
                fontName=_pdf_font_name(row),
                fontSize=9,
                textColor=colors.HexColor(fg),
            )
            return Paragraph(html.escape(text), ps)

        def _row_no_para(row: dict) -> Paragraph:
            fg = _row_fg_hex(row) or "#111111"
            ps = ParagraphStyle(
                f"RowNo{row.get('line_no')}",
                parent=styles["Normal"],
                fontName=_pdf_font_name(row),
                fontSize=9,
                textColor=colors.HexColor(fg),
            )
            return Paragraph(html.escape(str(row.get("row_no") or "")), ps)

        table_data = [
            ["Row No.", "Description"]
            + [str(col.get("header") or "") for col in columns]
        ]
        for row in rows:
            amounts = row.get("amounts") or {}
            amount_cells = [
                _amount_para(row, amounts.get(col.get("key")))
                for col in columns
            ]
            table_data.append(
                [_row_no_para(row), _desc_para(row)] + amount_cells
            )

        col_widths = [55, 220] + [90] * len(columns)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(HEADER_BG)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(GRID_COLOR)),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        for i, row in enumerate(rows, start=1):
            bg = _row_bg_hex(row)
            if bg:
                style_commands.append(
                    ("BACKGROUND", (0, i), (-1, i), colors.HexColor(bg))
                )
            if _row_has_top_line(row):
                style_commands.append(
                    ("LINEABOVE", (0, i), (-1, i), 1, colors.HexColor(ACCENT_LINE))
                )
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_excel(data: dict) -> bytes:
        title = data.get("description") or data.get("report_name") or "Financial Report"
        period_label = data.get("period_label") or ""
        currency = data.get("currency_code") or ""
        columns = data.get("columns") or []
        rows = _visible_export_rows(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=14, color=_excel_hex(HEADER_BG))
        ws.append([f"{period_label} · {currency}".strip(" ·")])
        ws["A2"].font = Font(color="555555")
        ws.append([])
        ws.append(
            ["Row No.", "Description"]
            + [str(col.get("header") or "") for col in columns]
        )
        header_row = ws.max_row
        header_fill = PatternFill(
            start_color=_excel_hex(HEADER_BG),
            end_color=_excel_hex(HEADER_BG),
            fill_type="solid",
        )
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            if cell.column >= 3:
                cell.alignment = Alignment(horizontal="right", vertical="center")

        for row in rows:
            amounts = row.get("amounts") or {}
            amount_cells = []
            for col in columns:
                value = amounts.get(col.get("key"))
                amount_cells.append(float(value) if value is not None else None)
            ws.append(
                [str(row.get("row_no") or ""), str(row.get("description") or "")]
                + amount_cells
            )
            _apply_excel_row_style(ws, ws.max_row, row, len(columns))

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 44
        for idx in range(len(columns)):
            col_letter = chr(ord("C") + idx)
            ws.column_dimensions[col_letter].width = 18

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def generate_financial_report(
    report_name: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> dict:
    return FinancialReportService().generate(report_name, start_date, end_date)


def print_financial_report(
    report_name: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> Tuple[dict, str]:
    data = generate_financial_report(report_name, start_date, end_date)
    return data, FinancialReportService.render_html(data)
